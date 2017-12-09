import json
import re
from random import choice

from discord import ChannelType, Embed
from discord.utils import find

from musicbot.bookmarks import bookmark
from musicbot.entry import Entry
from musicbot.logger import OnlineLogger
from musicbot.settings import Settings
from musicbot.utils import (Response, block_user, command_info, owner_only,
                            parse_timestamp, to_timestamp)
from musicbot.web_author import WebAuthor
from openpyxl import Workbook


class ToolCommands:

    @command_info("2.0.3", 1485516420, {
        "3.7.5": (1481827320, "The command finally works like it should"),
        "3.9.9": (1499977057, "moving Giesela too"),
        "4.1.8": (1500882643, "Updating to new player model")
    })
    async def cmd_moveus(self, server, channel, author, user_mentions, leftover_args):
        """
        ///|Usage
        `{command_prefix}moveus <channel name>`
        ///|Explanation
        Move everyone in your current channel to another one!
        """

        target_channel = None
        target = " ".join(leftover_args)

        if user_mentions:
            target_channel = user_mentions[0].voice.voice_channel

        if not target_channel and target:
            target_channel = find(
                lambda vc: vc.type == ChannelType.voice and target.lower().strip() in vc.name.lower().strip(),
                server.channels
            )

        if target_channel is None:
            return Response("Can't resolve the target channel!")

        author_channel = author.voice.voice_channel
        voice_members = author_channel.voice_members

        move_myself = False
        if server.me in voice_members:
            voice_members.remove(server.me)
            move_myself = True

        for voice_member in voice_members:
            await self.move_member(voice_member, target_channel)

        if move_myself:
            await self.get_player(server, target_channel)

    @command_info("1.0.0", 1477180800, {
        "2.0.2": (1481827560, "Can now use @mentions to \"goto\" a user"),
        "4.1.8": (1500881315, "Merging with old goto command")
    })
    async def cmd_summon(self, server, author, user_mentions, leftover_args):
        """
        Usage:
            {command_prefix}summon [@mention | name]

        Call the bot to the summoner's voice channel.
        """

        target_channel = None
        target = " ".join(leftover_args)

        if user_mentions:
            target_channel = user_mentions[0].voice.voice_channel

        if not target_channel and target:
            target_channel = find(lambda vc: vc.type == ChannelType.voice and target.lower().strip() in vc.name.lower().strip(), server.channels)

        if not target_channel:
            target_channel = author.voice_channel

        if not target_channel:
            return Response("Couldn't find voic channel")

        player = await self.get_player(server, target_channel)

        if player.is_stopped:
            player.play()

    @owner_only
    async def cmd_countmsgs(self, server, author, channel_id, number):
        """
        ///|Usage
        `{command_prefix}countmsgs <channel> <number>`
        ///|Explanation
        Count up to <number> messages in <channel> and return stats by user.
        """
        alphabet = list("abcdefghijklmnopqrstuvwxyz")

        def index_to_alphabet(ind):
            if ind < len(alphabet):
                return alphabet[ind].upper()

            remainder = ind % len(alphabet)
            return index_to_alphabet(ind -
                                     remainder) + alphabet[remainder].upper()

        msgs_by_member = {}
        msgs_by_date = OrderedDict()
        answers_by_date = OrderedDict()
        channel = server.get_channel(channel_id)
        last_msg = None
        last_answer = None
        spam = 0

        async for msg in self.logs_from(channel, limit=int(number)):
            increment = 1
            if last_msg is not None and msg.author.id == last_msg.author.id and abs(
                    (last_msg.timestamp - msg.timestamp).total_seconds()) < 10:
                spam += 1
                last_msg = msg
                increment = 0

            if last_answer is None or last_answer.author != msg.author:
                dt = answers_by_date.get(
                    "{0.day:0>2}/{0.month:0>2}/{0.year:0>4}".format(
                        msg.timestamp), {})
                dt[msg.author.id] = dt.get(msg.author.id, 0) + increment
                answers_by_date["{0.day:0>2}/{0.month:0>2}/{0.year:0>4}".
                                format(msg.timestamp)] = dt
                last_answer = msg

            existing_msgs = msgs_by_member.get(msg.author.id, [0, 0])
            existing_msgs[0] += increment
            existing_msgs[1] += len(re.sub(r"\W", r"", msg.content))
            msgs_by_member[msg.author.id] = existing_msgs
            dt = msgs_by_date.get(
                "{0.day:0>2}/{0.month:0>2}/{0.year:0>4}".format(msg.timestamp),
                {})
            dt[msg.author.id] = dt.get(msg.author.id, 0) + increment
            msgs_by_date["{0.day:0>2}/{0.month:0>2}/{0.year:0>4}".format(
                msg.timestamp)] = dt
            last_msg = msg

        wb = Workbook()
        ws = wb.active
        ws.title = "Messages"
        ws2 = wb.create_sheet("Answers")
        ws["A2"] = "TOTAL"
        sorted_user_index = {}
        i = 1
        for member in sorted(msgs_by_member):
            data = msgs_by_member[member]
            ws["{}{}".format("A", i)] = server.get_member(
                member
            ).name if server.get_member(member) is not None else "Unknown"
            ws["{}{}".format("B", i)] = data[0]
            ws["{}{}".format("C", i)] = data[1]
            sorted_user_index[member] = index_to_alphabet(i)
            i += 1

        i += 1
        for date in reversed(msgs_by_date.keys()):
            ws["A" + str(i)] = date
            for mem in msgs_by_date[date]:
                ws["{}{}".format(sorted_user_index.get(mem),
                                 i)] = msgs_by_date[date][mem]
            i += 1

        i = 1
        for date in reversed(answers_by_date.keys()):
            ws2["A" + str(i)] = date
            for mem in answers_by_date[date]:
                ws2["{}{}".format(sorted_user_index.get(mem),
                                  i)] = answers_by_date[date][mem]
            i += 1

        wb.save("cache/last_data.xlsx")

        await self.send_file(
            author,
            open("cache/last_data.xlsx", "rb"),
            filename="%s-msgs.xlsx" % (server.name.replace(" ", "_")))

    async def cmd_archivechat(self, server, author, message, placeholder=None, number=1000000):
        if message.channel_mentions is None or len(
                message.channel_mentions) < 1:
            return Response("Stupid duck")

        channel = message.channel_mentions[0]
        msgs = []
        async for msg in self.logs_from(channel, limit=int(number)):
            msg_data = {
                "name": msg.author.name,
                "timestamp": str(round(msg.timestamp.timestamp())),
                "content": msg.content,
                "attachments": msg.attachments
            }
            msgs.append(msg_data)

        json.dump(msgs[::-1], open("cache/last_message_archive.json", "w+"))
        await self.send_file(
            author,
            open("cache/last_message_archive.json", "rb"),
            filename="%s-msg-archive.json" % (server.name.replace(" ", "_")))

    @owner_only
    async def cmd_surveyserver(self, server):
        if self.online_loggers.get(server.id, None) is not None:
            return Response("I'm already looking at this server")
        else:
            online_logger = OnlineLogger(self)
            self.online_loggers[server.id] = online_logger
            Settings["online_loggers"] = list(self.online_loggers.keys())
            return Response("okay, okay!")

    def load_online_loggers(self):
        for server_id in Settings.get_setting("online_loggers", default=[]):
            online_logger = OnlineLogger(self)
            self.online_loggers[server_id] = online_logger
            for listener in Settings.get_setting(
                    "online_logger_listeners_" + server_id, default=[]):
                online_logger.add_listener(listener)

    @owner_only
    async def cmd_evalsurvey(self, server, author):
        online_logger = self.online_loggers.get(server.id, None)
        if online_logger is None:
            return Response("I'm not even spying here")
        online_logger.create_output()
        await self.send_file(
            author,
            open("cache/last_survey_data.xlsx", "rb"),
            filename="%s-survey.xlsx" % (server.name.replace(" ", "_")))
        return Response("There you go, fam")

    @owner_only
    async def cmd_resetsurvey(self, server):
        online_logger = self.online_loggers.get(server.id, None)
        if online_logger is None:
            return Response("I'm not even spying here")
        online_logger.reset()
        return Response("Well then")

    async def cmd_notifyme(self, server, author):
        """
        Usage:
            {command_prefix}notifyme

        Get notified when someone starts playing
        """
        online_logger = self.online_loggers.get(server.id, None)
        if online_logger is None:
            return Response("I'm not even spying here")
        if online_logger.add_listener(author.id):
            Settings["online_logger_listeners_" + server.id] = [
                *Settings.get_setting(
                    "online_logger_listeners_" + server.id, default=[]),
                author.id
            ]
            return Response("Got'cha!")
        else:
            try:
                Settings["online_logger_listeners_" + server.id] = [
                    x
                    for x in Settings.get_setting(
                        "online_logger_listeners_" + server.id, default=[])
                    if x != author.id
                ]
            except ValueError:
                pass

            return Response("Nevermore you shall be annoyed!")

    @command_info("2.2.1", 1493757540, {
        "3.7.8": (1499019245, "Fixed quoting by content.")
    })
    async def cmd_quote(self, author, channel, message, leftover_args):
        """
        ///|Usage
        `{command_prefix}quote [#channel] <message id> [message id...]`
        `{command_prefix}quote [#channel] [@mention] \"<message content>\"`
        ///|Explanation
        Quote a message
        """

        quote_to_channel = channel
        target_author = None

        if message.channel_mentions:
            channel = message.channel_mentions[0]
            leftover_args = leftover_args[1:]

        if message.mentions:
            target_author = message.mentions[0]
            leftover_args = leftover_args[1:]

        if len(leftover_args) < 1:
            return Response("Please specify the message you want to quote")

        message_content = " ".join(leftover_args)
        if (message_content[0] == "\"" and message_content[-1] == "\"") or re.search(r"\D", message_content) is not None:
            message_content = message_content.replace("\"", "")
            async for msg in self.logs_from(channel, limit=3000):
                if msg.id != message.id and message_content.lower().strip() in msg.content.lower().strip():
                    if target_author is None or target_author.id == msg.author.id:
                        leftover_args = [msg.id, ]
                        break
            else:
                if target_author is not None:
                    return Response("Didn't find a message with that content from {}".format(target_author.mention))
                else:
                    return Response("Didn't find a message with that content")

        await self.safe_delete_message(message)
        for message_id in leftover_args:
            try:
                quote_message = await self.get_message(channel, message_id)
            except:
                return Response("Didn't find a message with the id `{}`".
                                format(message_id))

            author_data = {
                "name": quote_message.author.display_name,
                "icon_url": quote_message.author.avatar_url
            }
            embed_data = {
                "description": quote_message.content,
                "timestamp": quote_message.timestamp,
                "colour": quote_message.author.colour
            }
            em = Embed(**embed_data)
            em.set_author(**author_data)
            await self.send_message(quote_to_channel, embed=em)
        return

    @command_info("3.2.5", 1496428380, {
        "3.3.9": (1497521393, "Added edit sub-command"),
        "3.4.1": (1497550771, "Added the filter \"mine\" to the listing function"),
        "3.4.6": (1497617827, "when listing bookmarks, they musn't be \"inline\"."),
        "3.5.8": (1497827057, "Editing bookmarks now works as expected"),
        "4.6.1": (1502582759, "Updated to new entry model"),
        "4.7.8": (1504105817, "Using the new copy method for entries to make sure that they're \"clean\"")
    })
    async def cmd_bookmark(self, author, player, leftover_args):
        """
        ///|Creation
        {command_prefix}bookmark [name] [timestamp]
        ///|Explanation
        Create a new bookmark for the current entry. If no name is provided the entry's title will be used and if there's no timestamp provided the current timestamp will be used.
        ///|Using
        {command_prefix}bookmark <id | name>
        ///|Editing
        {command_prefix}bookmark edit <id> [new name] [new timestamp]
        ///|Listing
        {command_prefix}bookmark list [mine]
        ///|Removal
        {command_prefix}bookmark remove <id | name>
        """
        if len(leftover_args) > 0:
            arg = leftover_args[0].lower()

            if arg in ["list", "showall"]:
                em = Embed(title="Bookmarks")
                bookmarks = bookmark.all_bookmarks

                if "mine" in leftover_args:
                    bookmarks = filter(lambda x: bookmark.get_bookmark(x)["author"]["id"] == author.id, bookmarks)

                for bm in bookmarks:
                    bm_name = bm["name"]
                    bm_author = WebAuthor.from_dict(bm["author"]).display_name
                    bm_timestamp = to_timestamp(bm["timestamp"])
                    bm_id = bm["id"]

                    t = "**{}**".format(bm_name)
                    v = "`{}` starting at `{}` *by* **{}**".format(bm_id, bm_timestamp, bm_author)

                    em.add_field(name=t, value=v, inline=False)

                return Response(embed=em)

            elif arg in ["remove", "delete"]:
                if len(leftover_args) < 2:
                    return Response("Please provide an id or a name")
                bm = bookmark.get_bookmark(" ".join(leftover_args[1:]))
                if not bm:
                    return Response("Didn't find a bookmark with that query")
                if bookmark.remove_bookmark(bm["id"]):
                    return Response("Removed bookmark `{}`".format(bm["name"]))
                else:
                    return Response("Something went wrong")

            elif arg in ["edit", "change"]:
                if len(leftover_args) < 2:
                    return Response("Please provide an id")

                bm_id = leftover_args[1]
                if bm_id not in bookmark:
                    return Response("No bookmark with id `{}` found".format(bm_id))

                if len(leftover_args) < 3:
                    return Response("Please also specify what you want to change")

                new_timestamp = parse_timestamp(leftover_args[-1])
                if new_timestamp is not None:  # 0 evaluates to false so I need to check this oldschool-like
                    new_name = " ".join(leftover_args[2:-1]) if len(leftover_args) > 3 else None
                else:
                    new_name = " ".join(leftover_args[2:])

                if bookmark.edit_bookmark(bm_id, new_name, new_timestamp):
                    return Response("Successfully edited bookmark `{}`".format(bm_id))
                else:
                    return Response("Something went wrong while editing `{}`".format(bm_id))

            else:
                bm = bookmark.get_bookmark(" ".join(leftover_args))
                if bm:
                    entry = Entry.from_dict(player.queue, bm["entry"])
                    entry.seek(bm["timestamp"])

                    player.queue._add_entry(entry)

                    author = WebAuthor.from_dict(bm["author"])
                    return Response("Loaded bookmark `{0}` by **{1}**".format(bm["name"], author.display_name))

                elif player.current_entry:
                    bm_timestamp = player.progress
                    bm_name = None

                    if len(leftover_args) > 1:
                        timestamp = parse_timestamp(leftover_args[-1])
                        if timestamp:
                            bm_timestamp = timestamp

                        bm_name = " ".join(
                            leftover_args[:-1]) if timestamp else " ".join(leftover_args)
                    else:
                        timestamp = parse_timestamp(leftover_args[-1])
                        if timestamp:
                            bm_timestamp = timestamp
                        else:
                            bm_name = " ".join(leftover_args)

                    id = bookmark.add_bookmark(player.current_entry, bm_timestamp, author.id, bm_name)
                    return Response("Created a new bookmark with the id `{0}` (\"{2}\", `{3}`)\nUse `{1}bookmark {0}` to load it ".format(id, self.config.command_prefix, bm_name, to_timestamp(bm_timestamp)))
                else:
                    return Response("There's no such bookmark and there's nothing playing either")

        else:
            if player.current_entry:
                id = bookmark.add_bookmark(player.current_entry.copy(),
                                           player.progress, author.id)
                return Response(
                    "Created a new bookmark with the id `{0}`\nUse `{1}bookmark {0}` to load it ".
                    format(id, self.config.command_prefix))
            else:
                return await self.cmd_bookmark(author, player, [
                    "list",
                ])

    @block_user
    @command_info("2.0.3", 1486054560, {
        "3.7.2": (1498252803, "no arguments provided crash Fixed")
    })
    async def cmd_random(self, channel, author, leftover_args):
        """
        ///|Basic
        `{command_prefix}random <item1>, <item2>, [item3], [item4]`
        ///|Use an existing set
        `{command_prefix}random <setname>`
        ///|List all the existing sets
        `{command_prefix}random list`
        ///|Creation
        `{command_prefix}random create <name>, <option1>, <option2>, [option3], [option4]`
        ///|Editing
        `{command_prefix}random edit <name>, [add | remove | replace], <item> [, item2, item3]`
        ///|Removal
        `{command_prefix}random remove <name>`
        ///|Explanation
        Choose a random item out of a list or use a pre-defined list.
        """

        if not leftover_args:
            return Response("Why u gotta be stupid?")

        items = [x.strip()
                 for x in " ".join(leftover_args).split(",") if x is not ""]

        if items[0].split()[0].lower().strip() == "create":
            if len(items) < 2:
                return Response(
                    "Can't create a set with the given arguments",
                    delete_after=20)

            set_name = "_".join(items[0].split()[1:]).lower().strip()
            set_items = items[1:]
            if self.random_sets.create_set(set_name, set_items):
                return Response(
                    "Created set **{0}**\nUse `{1}random {0}` to use it!".format(
                        set_name, self.config.command_prefix),
                    delete_after=60)
            else:
                return Response(
                    "OMG, shit went bad quickly! Everything's burning!\nDUCK there he goes again, the dragon's coming. Eat HIM not me. PLEEEEEEEEEEEEEASE!"
                )
        elif items[0].split()[0].lower().strip() == "list":
            return_string = ""
            for s in self.random_sets.get_sets():
                return_string += "**{}**\n```\n{}```\n\n".format(
                    s[0], ", ".join(s[1]))

            return Response(return_string)
        elif items[0].split()[0].lower().strip() == "edit":
            if len(items[0].split()) < 2:
                return Response(
                    "Please provide the name of the list you wish to edit!",
                    delete_after=20)

            set_name = "_".join(items[0].split()[1:]).lower().strip()

            existing_items = self.random_sets.get_set(set_name)
            if existing_items is None:
                return Response("This set does not exist!")

            edit_mode = items[1].strip().lower() if len(items) > 1 else None
            if edit_mode is None:
                return Response(
                    "You need to provide the way you want to edit the list",
                    delete_after=20)

            if len(items) < 3:
                return Response(
                    "You have to specify the items you want to add/remove or set as the new items"
                )

            if edit_mode == "add":
                for option in items[2:]:
                    self.random_sets.add_option(set_name, option)
            elif edit_mode == "remove":
                for option in items[2:]:
                    self.random_sets.remove_option(set_name, option)
            elif edit_mode == "replace":
                self.random_sets.replace_options(set_name, items[2:])
            else:
                return Response(
                    "This is not a valid edit mode!")

            return Response("Edited your set!")
        elif items[0].split()[0].lower().strip() == "remove":
            set_name = "_".join(items[0].split()[1:]).lower().strip()
            set_items = items[1:]
            res = self.random_sets.remove_set(set_name, set_items)
            if res:
                return Response("Removed set!")
            elif res is None:
                return Response("No such set!")
            else:
                return Response(
                    "OMG, shit went bad quickly! Everything's burning!\nDUCK there he goes again, the dragon's coming. Eat HIM not me. PLEEEEEEEEEEEEEASE!"
                )

        if len(items) <= 0 or items is None:
            return Response(
                "Is your name \"{0}\" by any chance?\n(This is not how this command works. Use `{1}help random` to find out how not to be a stupid **{0}** anymore)".
                format(author.name, self.config.command_prefix),
                delete_after=30)

        if len(items) <= 1:
            # return Response("Only you could use `{1}random` for one item...
            # Well done, {0}!".format(author.name, self.config.command_prefix),
            # delete_after=30)

            query = "_".join(items[0].split())
            items = self.random_sets.get_set(query.lower().strip())
            if items is None:
                return Response("Something went wrong")

        await self.safe_send_message(channel,
                                     "I choose **" + choice(items) + "**")
