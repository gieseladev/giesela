import asyncio
import configparser
import datetime
import inspect
import json
import operator
import os
import re
import shlex
import shutil
import subprocess
import sys
import time
import traceback
from collections import OrderedDict, defaultdict
from datetime import date, datetime, timedelta
from functools import wraps
from io import BytesIO
from random import choice, random, shuffle
from textwrap import dedent, indent

import aiohttp
import discord
import requests
from discord import Embed, utils
from discord.enums import ChannelType
from discord.ext.commands.bot import _get_variable
from discord.object import Object
from discord.utils import find
from discord.voice_client import VoiceClient
from moviepy import editor, video
from openpyxl import Workbook

from . import downloader, exceptions
from .bookmarks import bookmark
from .cleverbot import CleverWrap
from .config import Config, ConfigDefaults
from .constants import VERSION as BOTVERSION
from .constants import AUDIO_CACHE_PATH, DISCORD_MSG_CHAR_LIMIT
from .entry import (RadioSongEntry, RadioStationEntry, SpotifyEntry,
                    StreamEntry, TimestampEntry, YoutubeEntry)
from .games.game_2048 import Game2048
from .games.game_cah import GameCAH
from .games.game_hangman import GameHangman
from .logger import OnlineLogger
from .lyrics import search_for_lyrics
from .nine_gag import ContentType, get_post
from .opus_loader import load_opus_lib
from .player import MusicPlayer
from .radio import RadioStations
from .random_sets import RandomSets
from .reminder import Action, Calendar
from .saved_playlists import Playlists
from .settings import Settings
from .tungsten import Tungsten
from .utils import (create_bar, escape_dis, format_time, get_dev_changelog,
                    get_dev_version, get_master_version, get_related_videos,
                    hex_to_dec, load_file, nice_cut, ordinal, paginate,
                    parse_timestamp, prettydate, random_line,
                    run_function_every, to_timestamp, write_file)
from .web_socket_server import GieselaServer

load_opus_lib()


class Response:

    def __init__(self, content=None, reply=False, delete_after=0, embed=None):
        self.content = content
        self.reply = reply
        self.delete_after = delete_after
        self.embed = embed


class send_typing(run_function_every):

    def __init__(self, bot, channel):
        self.channel = channel
        self.bot = bot
        super().__init__(self._func, 9)

    def _func(self):
        asyncio.run_coroutine_threadsafe(
            self.bot.send_typing(self.channel), self.bot.loop)


class MusicBot(discord.Client):

    def __init__(self):
        self.players = {}
        self.the_voice_clients = {}
        self.locks = defaultdict(asyncio.Lock)
        self.voice_client_connect_lock = asyncio.Lock()
        self.voice_client_move_lock = asyncio.Lock()

        self.config = Config(ConfigDefaults.options_file)
        self.playlists = Playlists(ConfigDefaults.playlists_file)
        self.random_sets = RandomSets(ConfigDefaults.random_sets)
        self.online_loggers = {}
        self.cah = GameCAH(self)

        self.blacklist = set(load_file(self.config.blacklist_file))
        self.autoplaylist = load_file(self.config.auto_playlist_file)
        self.downloader = downloader.Downloader(download_folder='audio_cache')
        self.calendar = Calendar(self)

        self.exit_signal = None
        self.init_ok = False
        self.cached_client_id = None
        self.chatters = {}
        self.blocked_commands = Settings.get_setting(
            "blocked_commands", default={})
        self.users_in_menu = set()

        if not self.autoplaylist:
            print("Warning: Autoplaylist is empty, disabling.")
            self.config.auto_playlist = False

        ssd_defaults = {'last_np_msg': None, 'auto_paused': False}
        self.server_specific_data = defaultdict(lambda: dict(ssd_defaults))

        super().__init__()
        self.aiosession = aiohttp.ClientSession(loop=self.loop)
        self.http.user_agent += ' MusicBot/%s' % BOTVERSION
        self.instant_translate = False
        self.instant_translate_mode = 1
        self.instant_translate_certainty = .7

        self.load_online_loggers()

    def owner_only(func):
        @wraps(func)
        async def wrapper(self, *args, **kwargs):
            # Only allow the owner to use these commands
            orig_msg = _get_variable('message')

            if not orig_msg or orig_msg.author.id == self.config.owner_id:
                return await func(self, *args, **kwargs)
            else:
                return Response("only the owner can use this command")

        return wrapper

    def command_info(version, timestamp, changelog={}):
        def function_decorator(func):
            func.version = version
            func.timestamp = datetime.fromtimestamp(timestamp)
            func.changelog = [(ver, datetime.fromtimestamp(time), log)
                              for ver, (time, log) in changelog.items()]

            return func

        return function_decorator

    def block_user(func):
        @wraps(func)
        async def wrapper(self, *args, **kwargs):
            orig_msg = _get_variable("message")

            self.users_in_menu.add(orig_msg.author.id)
            print("Now blocking " + str(orig_msg.author))
            try:
                res = await func(self, *args, **kwargs)
                self.users_in_menu.remove(orig_msg.author.id)
                print("Unblocking " + str(orig_msg.author))
                return res
            except Exception as e:  # just making sure that no one gets stuck in a menu and can't use any commands anymore
                self.users_in_menu.remove(orig_msg.author.id)
                raise e

        return wrapper

    @staticmethod
    def _fixg(x, dp=2):
        return ('{:.%sf}' % dp).format(x).rstrip('0').rstrip('.')

    def _get_owner(self, voice=False):
        if voice:
            for server in self.servers:
                for channel in server.channels:
                    for m in channel.voice_members:
                        if m.id == self.config.owner_id:
                            return m
        else:
            return discord.utils.find(lambda m: m.id == self.config.owner_id,
                                      self.get_all_members())

    def _delete_old_audiocache(self, path=AUDIO_CACHE_PATH):
        try:
            shutil.rmtree(path)
            return True
        except:
            try:
                os.rename(path, path + '__')
            except:
                return False
            try:
                shutil.rmtree(path)
            except:
                os.rename(path + '__', path)
                return False

        return True

    async def _auto_summon(self):
        owner = self._get_owner(voice=True)
        if owner:
            print("Found owner in \"%s\", attempting to join..." %
                  owner.voice_channel.name)
            await self.cmd_summon(owner.voice_channel, owner, None)
            return owner.voice_channel

    async def _autojoin_channels(self, channels):
        joined_servers = []

        for channel in channels:
            if channel.server in joined_servers:
                print("Already joined a channel in %s, skipping" %
                      channel.server.name)
                continue

            if channel and channel.type == discord.ChannelType.voice:
                print("Attempting to autojoin %s in %s" %
                      (channel.name, channel.server.name))

                chperms = channel.permissions_for(channel.server.me)

                if not chperms.connect:
                    print("Cannot join channel \"%s\", no permission." %
                          channel.name)
                    continue

                elif not chperms.speak:
                    print(
                        "Will not join channel \"%s\", no permission to speak."
                        % channel.name)
                    continue

                try:
                    player = await self.get_player(channel, create=True)

                    if player.is_stopped:
                        player.play()

                    if self.config.auto_playlist:
                        await self.on_player_finished_playing(player)

                    joined_servers.append(channel.server)
                except Exception as e:
                    if self.config.debug_mode:
                        traceback.print_exc()
                    print("Failed to join", channel.name)

            elif channel:
                print("Not joining %s on %s, that's a text channel." %
                      (channel.name, channel.server.name))

            else:
                print("Invalid channel thing: " + channel)

    async def _wait_delete_msg(self, message, after):
        await asyncio.sleep(after)
        await self.safe_delete_message(message)

    async def _manual_delete_check(self, message, *, quiet=False):
        if self.config.delete_invoking:
            await self.safe_delete_message(message, quiet=quiet)

    async def generate_invite_link(self, *, permissions=None, server=None):
        if not self.cached_client_id:
            appinfo = await self.application_info()
            self.cached_client_id = appinfo.id

        return discord.utils.oauth_url(
            self.cached_client_id, permissions=permissions, server=server)

    async def get_voice_client(self, channel):
        if isinstance(channel, Object):
            channel = self.get_channel(channel.id)

        if getattr(channel, 'type', ChannelType.text) != ChannelType.voice:
            raise AttributeError('Channel passed must be a voice channel')

        with await self.voice_client_connect_lock:
            server = channel.server
            if server.id in self.the_voice_clients:
                return self.the_voice_clients[server.id]

            s_id = self.ws.wait_for('VOICE_STATE_UPDATE',
                                    lambda d: d.get('user_id') == self.user.id)
            _voice_data = self.ws.wait_for('VOICE_SERVER_UPDATE',
                                           lambda d: True)

            await self.ws.voice_state(server.id, channel.id)

            s_id_data = await asyncio.wait_for(
                s_id, timeout=10, loop=self.loop)
            voice_data = await asyncio.wait_for(
                _voice_data, timeout=10, loop=self.loop)
            session_id = s_id_data.get('session_id')

            kwargs = {
                'user': self.user,
                'channel': channel,
                'data': voice_data,
                'loop': self.loop,
                'session_id': session_id,
                'main_ws': self.ws
            }
            voice_client = VoiceClient(**kwargs)
            self.the_voice_clients[server.id] = voice_client

            retries = 3
            for x in range(retries):
                try:
                    print("Attempting connection...")
                    await asyncio.wait_for(
                        voice_client.connect(), timeout=10, loop=self.loop)
                    print("Connection established.")
                    break
                except:
                    traceback.print_exc()
                    print("Failed to connect, retrying (%s/%s)..." %
                          (x + 1, retries))
                    await asyncio.sleep(1)
                    await self.ws.voice_state(server.id, None, self_mute=True)
                    await asyncio.sleep(1)

                    if x == retries - 1:
                        raise exceptions.HelpfulError(
                            "Cannot establish connection to voice chat.  "
                            "Something may be blocking outgoing UDP connections.",
                            "This may be an issue with a firewall blocking UDP.  "
                            "Figure out what is blocking UDP and disable it.  "
                            "It's most likely a system firewall or overbearing anti-virus firewall.  "
                        )

            return voice_client

    def get_global_user(self, user_id):
        for server in self.servers:
            mem = server.get_member(user_id)
            if mem is not None:
                return mem

        return None

    async def mute_voice_client(self, channel, mute):
        await self._update_voice_state(channel, mute=mute)

    async def deafen_voice_client(self, channel, deaf):
        await self._update_voice_state(channel, deaf=deaf)

    async def move_voice_client(self, channel):
        await self._update_voice_state(channel)

    async def reconnect_voice_client(self, server):
        if server.id not in self.the_voice_clients:
            return

        vc = self.the_voice_clients.pop(server.id)
        _paused = False

        player = None
        if server.id in self.players:
            player = self.players[server.id]
            if player.is_playing:
                player.pause()
                _paused = True

        try:
            await vc.disconnect()
        except:
            print("Error disconnecting during reconnect")
            traceback.print_exc()

        await asyncio.sleep(0.1)

        if player:
            new_vc = await self.get_voice_client(vc.channel)
            player.reload_voice(new_vc)

            if player.is_paused and _paused:
                player.resume()

    async def disconnect_voice_client(self, server):
        if server.id not in self.the_voice_clients:
            return

        if server.id in self.players:
            self.players.pop(server.id).kill()

        await self.the_voice_clients.pop(server.id).disconnect()

    async def disconnect_all_voice_clients(self):
        for vc in self.the_voice_clients.copy().values():
            await self.disconnect_voice_client(vc.channel.server)

    async def _update_voice_state(self, channel, *, mute=False, deaf=False):
        if isinstance(channel, Object):
            channel = self.get_channel(channel.id)

        if getattr(channel, 'type', ChannelType.text) != ChannelType.voice:
            raise AttributeError('Channel passed must be a voice channel')

        # I'm not sure if this lock is actually needed
        with await self.voice_client_move_lock:
            server = channel.server

            payload = {
                'op': 4,
                'd': {
                    'guild_id': server.id,
                    'channel_id': channel.id,
                    'self_mute': mute,
                    'self_deaf': deaf
                }
            }

            await self.ws.send(utils.to_json(payload))
            self.the_voice_clients[server.id].channel = channel

    async def get_player(self, channel=None, create=False, auto_summon=True, server_id=None):
        server = channel.server if channel else self.get_server(server_id)

        if server.id not in self.players:
            if not create:
                if auto_summon:
                    channel = await self.goto_home(
                        channel.server)
                else:
                    raise exceptions.CommandError(
                        'The bot is not in a voice channel.  '
                        'Use %ssummon to summon it to your voice channel.' %
                        self.config.command_prefix)

            voice_client = await self.get_voice_client(channel)

            player = MusicPlayer(self, voice_client) \
                .on('play', self.on_player_play) \
                .on('resume', self.on_player_resume) \
                .on('pause', self.on_player_pause) \
                .on('stop', self.on_player_stop) \
                .on('finished-playing', self.on_player_finished_playing) \
                .on('entry-added', self.on_player_entry_added)

            self.players[server.id] = player

        return self.players[server.id]

    async def on_player_play(self, player, entry):
        GieselaServer.send_player_information_update(
            player.voice_client.server.id)
        await self.update_now_playing(entry)

        channel = entry.meta.get("channel", None)

        if channel:
            last_np_msg = self.server_specific_data[channel.server][
                "last_np_msg"]
            if last_np_msg and last_np_msg.channel == channel:

                # if the last np message isn't the last message in the channel;
                # delete it
                async for lmsg in self.logs_from(channel, limit=1):
                    if lmsg != last_np_msg and last_np_msg:
                        await self.safe_delete_message(last_np_msg)
                        self.server_specific_data[channel.server][
                            'last_np_msg'] = None
                    break  # This is probably redundant

            if isinstance(entry, TimestampEntry):
                sub_entry = entry.current_sub_entry
                sub_title = sub_entry["name"]
                sub_index = sub_entry["index"] + 1
                newmsg = "Now playing **{0}** ({1}{2} entry) from \"{3}\"".format(
                    sub_title, sub_index, ordinal(sub_index), entry.whole_title)
            elif isinstance(entry, RadioSongEntry):
                newmsg = "Now playing **{}**".format(
                    " - ".join((entry.artist, entry.title)))
            else:
                newmsg = "Now playing **{}**".format(entry.title)

            if self.server_specific_data[channel.server]["last_np_msg"]:
                self.server_specific_data[channel.server][
                    "last_np_msg"] = await self.safe_edit_message(
                        last_np_msg, newmsg, send_if_fail=True)
            else:
                self.server_specific_data[channel.server][
                    "last_np_msg"] = await self.safe_send_message(
                        channel, newmsg)

    async def on_player_resume(self, player, entry, **_):
        await self.update_now_playing(entry)
        GieselaServer.send_player_information_update(
            player.voice_client.server.id)

    async def on_player_pause(self, player, entry, **_):
        await self.update_now_playing(entry, True)
        GieselaServer.send_player_information_update(
            player.voice_client.server.id)

    async def on_player_stop(self, player, **_):
        await self.update_now_playing()
        GieselaServer.send_player_information_update(
            player.voice_client.server.id)

    async def on_player_finished_playing(self, player, **_):
        if not player.playlist.entries and not player.current_entry:
            GieselaServer.send_player_information_update(
                player.voice_client.server.id)

        if not player.playlist.entries and not player.current_entry and self.config.auto_playlist:
            if self.config.auto_playlist:
                while self.autoplaylist:
                    song_url = choice(self.autoplaylist)
                    info = await self.downloader.safe_extract_info(
                        player.playlist.loop,
                        song_url,
                        download=False,
                        process=False)

                    if not info:
                        self.autoplaylist.remove(song_url)
                        print(
                            "[Info] Removing unplayable song from autoplaylist: %s"
                            % song_url)
                        write_file(self.config.auto_playlist_file,
                                   self.autoplaylist)
                        continue

                    if info.get('entries', None):  # or .get('_type', '') == 'playlist'
                        pass  # Wooo playlist
                        # Blarg how do I want to do this

                    try:
                        await player.playlist.add_entry(
                            song_url, channel=None, author=None)
                    except exceptions.ExtractionError as e:
                        print("Error adding song from autoplaylist:", e)
                        continue

                    break

                if not self.autoplaylist:
                    print(
                        "[Warning] No playable songs in the autoplaylist, disabling."
                    )
                    self.config.auto_playlist = False

    async def on_player_entry_added(self, playlist, entry, **_):
        pass

    async def on_server_join(self, server):
        for channel in server.channels:
            if channel.type is not ChannelType.text:
                continue

            msg = await self.safe_send_message(
                channel,
                "Hello there,\nMy name is {}!\n\n*Type {}help to find out more.*".
                format(self.user.mention, self.config.command_prefix))
            if msg is not None:
                return

    async def update_now_playing(self, entry=None, is_paused=False):
        game = None

        if self.user.bot:
            activeplayers = sum(1 for p in self.players.values()
                                if p.is_playing)
            if activeplayers > 1:
                game = discord.Game(name="Music")
                entry = None

            elif activeplayers == 1:
                player = discord.utils.get(
                    self.players.values(), is_playing=True)
                entry = player.current_entry

        if entry:
            prefix = "\u275A\u275A" if is_paused else ""

            if isinstance(entry, StreamEntry):
                prefix += "\u25CE"
                name = entry.title
            elif isinstance(entry, TimestampEntry):
                prefix += "\u1F4DC"
                name = entry.title
            else:
                name = entry.title

            name = u"{} {}".format(prefix, name)[:128]
            game = discord.Game(name=name)

        await self.change_presence(game=game)

    async def safe_send_message(self,
                                dest,
                                content=None,
                                *,
                                max_letters=DISCORD_MSG_CHAR_LIMIT,
                                split_message=True,
                                tts=False,
                                expire_in=0,
                                also_delete=None,
                                quiet=False,
                                embed=None):
        msg = None
        try:
            if split_message and content and len(content) > max_letters:
                print("Message too long, splitting it up")
                msgs = paginate(content, length=DISCORD_MSG_CHAR_LIMIT)

                for msg in msgs:
                    nmsg = await self.send_message(dest, msg, tts=tts)

                    if nmsg and expire_in:
                        asyncio.ensure_future(
                            self._wait_delete_msg(nmsg, expire_in))

                    if also_delete and isinstance(also_delete,
                                                  discord.Message):
                        asyncio.ensure_future(
                            self._wait_delete_msg(also_delete, expire_in))
            else:
                msg = await self.send_message(
                    dest, content, tts=tts, embed=embed)

                if msg and expire_in:
                    asyncio.ensure_future(
                        self._wait_delete_msg(msg, expire_in))

                if also_delete and isinstance(also_delete, discord.Message):
                    asyncio.ensure_future(
                        self._wait_delete_msg(also_delete, expire_in))

        except discord.Forbidden:
            if not quiet:
                print("Warning: Cannot send message to %s, no permission" %
                      dest.name)

        except discord.NotFound:
            if not quiet:
                print("Warning: Cannot send message to %s, invalid channel?"
                      % dest.name)

        return msg

    async def safe_delete_message(self, message, *, quiet=False):
        try:
            return await self.delete_message(message)

        except discord.Forbidden:
            if not quiet:
                print("Warning: Cannot delete message \"%s\", no permission"
                      % message.clean_content)

        except discord.NotFound:
            if not quiet:
                print(
                    "Warning: Cannot delete message \"%s\", message not found"
                    % message.clean_content)

    async def safe_edit_message(self, message, new, *, send_if_fail=False, quiet=False, keep_at_bottom=False):
        if keep_at_bottom:
            async for lmsg in self.logs_from(message.channel, limit=5):
                if lmsg.id == message.id:
                    break
            else:
                await self.safe_delete_message(message)
                return await self.safe_send_message(message.channel, new)
                return

        try:
            return await self.edit_message(message, new)

        except discord.NotFound:
            if not quiet:
                print(
                    "Warning: Cannot edit message \"%s\", message not found" %
                    message.clean_content)
            if send_if_fail:
                if not quiet:
                    print("Sending instead")
                return await self.safe_send_message(message.channel, new)

    def log(self, content="\n", *, end='\n', flush=True):
        sys.stdout.buffer.write((content + end).encode('utf-8', 'replace'))
        if flush:
            sys.stdout.flush()

    async def send_typing(self, destination):
        try:
            return await super().send_typing(destination)
        except discord.Forbidden:
            if self.config.debug_mode:
                print(
                    "Could not send typing to %s, no permssion" % destination)

    async def edit_profile(self, **fields):
        if self.user.bot:
            return await super().edit_profile(**fields)
        else:
            return await super().edit_profile(self.config._password, **fields)

    def _cleanup(self):
        try:
            self.loop.run_until_complete(self.logout())
        except:  # Can be ignored
            pass

        pending = asyncio.Task.all_tasks()
        gathered = asyncio.gather(*pending)

        try:
            gathered.cancel()
            self.loop.run_until_complete(gathered)
            gathered.exception()
        except:  # Can be ignored
            pass

    def run(self):
        try:
            self.loop.run_until_complete(self.start(*self.config.auth))

        except discord.errors.LoginFailure:
            # Add if token, else
            raise exceptions.HelpfulError(
                "Bot cannot login, bad credentials.",
                "Fix your Email or Password or Token in the options file.  "
                "Remember that each field should be on their own line.")

        finally:
            try:
                self._cleanup()
            except Exception as e:
                print("Error in cleanup:", e)

            self.loop.close()
            if self.exit_signal:
                raise self.exit_signal

    async def logout(self):
        await self.disconnect_all_voice_clients()
        return await super().logout()

    async def on_error(self, event, *args, **kwargs):
        ex_type, ex, stack = sys.exc_info()

        if ex_type == exceptions.HelpfulError:
            print("Exception in " + str(event))
            print(ex.message)

            await asyncio.sleep(2)  # don't ask
            await self.logout()

        elif issubclass(ex_type, exceptions.Signal):
            self.exit_signal = ex_type
            await self.logout()

        else:
            traceback.print_exc()

    async def on_resumed(self):
        for vc in self.the_voice_clients.values():
            vc.main_ws = self.ws

    async def on_ready(self):
        print('\rConnected!  Musicbot v%s\n' % BOTVERSION)

        if self.config.owner_id == self.user.id:
            raise exceptions.HelpfulError(
                "Your OwnerID is incorrect or you've used the wrong credentials.",
                "The bot needs its own account to function.  "
                "The OwnerID is the id of the owner, not the bot.  "
                "Figure out which one is which and use the correct information."
            )

        self.init_ok = True

        print("Bot:   %s/%s#%s" % (self.user.id, self.user.name,
                                   self.user.discriminator))

        owner = self._get_owner(voice=True) or self._get_owner()
        if owner and self.servers:
            print("Owner: %s/%s#%s\n" % (owner.id, owner.name,
                                         owner.discriminator))

            print('Server List:')
            [print(' - ' + s.name) for s in self.servers]

        elif self.servers:
            print("Owner could not be found on any server (id: %s)\n" %
                  self.config.owner_id)

            print('Server List:')
            [print(' - ' + s.name) for s in self.servers]

        else:
            print("Owner unknown, bot is not on any servers.")
            if self.user.bot:
                print(
                    "\nTo make the bot join a server, paste this link in your browser."
                )
                print(
                    "Note: You should be logged into your main account and have \n"
                    "manage server permissions on the server you want the bot to join.\n"
                )
                print("    " + await self.generate_invite_link())

        print()

        if self.config.bound_channels:
            chlist = set(
                self.get_channel(i) for i in self.config.bound_channels if i)
            chlist.discard(None)
            invalids = set()

            invalids.update(c for c in chlist
                            if c.type == discord.ChannelType.voice)
            chlist.difference_update(invalids)
            self.config.bound_channels.difference_update(invalids)

            print("Bound to text channels:")
            [
                print(' - %s/%s' % (ch.server.name.strip(),
                                    ch.name.strip())) for ch in chlist if ch
            ]

            if invalids and self.config.debug_mode:
                print("\nNot binding to voice channels:")
                [
                    print(' - %s/%s' % (ch.server.name.strip(),
                                        ch.name.strip())) for ch in invalids
                    if ch
                ]

            print()

        else:
            print("Not bound to any text channels")

        if self.config.autojoin_channels:
            chlist = set(
                self.get_channel(i) for i in self.config.autojoin_channels
                if i)
            chlist.discard(None)
            invalids = set()

            invalids.update(c for c in chlist
                            if c.type == discord.ChannelType.text)
            chlist.difference_update(invalids)
            self.config.autojoin_channels.difference_update(invalids)

            print("Autojoining voice chanels:")
            [
                print(' - %s/%s' % (ch.server.name.strip(),
                                    ch.name.strip())) for ch in chlist if ch
            ]

            if invalids and self.config.debug_mode:
                print("\nCannot join text channels:")
                [
                    print(' - %s/%s' % (ch.server.name.strip(),
                                        ch.name.strip())) for ch in invalids
                    if ch
                ]

            autojoin_channels = chlist

        else:
            print("Not autojoining any voice channels")
            autojoin_channels = set()

        print()
        print("Options:")

        print("  Command prefix: " + self.config.command_prefix)
        print(
            "  Default volume: %s%%" % int(self.config.default_volume * 100))
        print("  Auto-Summon: " + ['Disabled', 'Enabled'
                                   ][self.config.auto_summon])
        print("  Auto-Playlist: " + ['Disabled', 'Enabled'
                                     ][self.config.auto_playlist])
        print("  Auto-Pause: " + ['Disabled', 'Enabled'
                                  ][self.config.auto_pause])
        print("  Delete Messages: " + ['Disabled', 'Enabled'
                                       ][self.config.delete_messages])
        if self.config.delete_messages:
            print("    Delete Invoking: " + ['Disabled', 'Enabled'
                                             ][self.config.delete_invoking])
        print("  Debug Mode: " + ['Disabled', 'Enabled'
                                  ][self.config.debug_mode])
        print("  Downloaded songs will be %s" % ['deleted', 'saved'
                                                 ][self.config.save_videos])
        print()

        # maybe option to leave the ownerid blank and generate a random command for the owner to use
        # wait_for_message is pretty neato

        if not self.config.save_videos and os.path.isdir(AUDIO_CACHE_PATH):
            if self._delete_old_audiocache():
                print("Deleting old audio cache")
            else:
                print("Could not delete old audio cache, moving on.")

        if self.config.autojoin_channels:
            await self._autojoin_channels(autojoin_channels)

        elif self.config.auto_summon:
            print("Attempting to autosummon...", flush=True)

            # waitfor + get value
            owner_vc = await self._auto_summon()

            if owner_vc:
                print("Done!", flush=True)
                if self.config.auto_playlist:
                    print("Starting auto-playlist")
                    await self.on_player_finished_playing(
                        await self.get_player(owner_vc))
            else:
                print(
                    "Owner not found in a voice channel, could not autosummon."
                )

        print()
        # t-t-th-th-that's all folks!

        GieselaServer.run(self)

    @command_info("1.9.5", 1477774380, {
        "3.4.5": (1497616203, "Improved default help message using embeds"),
        "3.6.0": (1497904733, "Fixed weird indent of some help texts"),
        "3.7.0": (1498233256, "Some better help texts"),
        "3.7.1": (1498237739, "Added interactive help"),
        "3.7.4": (1498318916, "Added \"lyrics\" function help text")
    })
    async def cmd_help(self, channel, leftover_args):
        """
        ///|Usage
        `{command_prefix}help [command]`
        ///|Explanation
        Logs a help message.
        ///|Interactive
        `{command_prefix}help <query>`
        """
        command = None

        if len(leftover_args) > 0:
            command = " ".join(leftover_args)

        if command:
            cmd = getattr(self, 'cmd_' + command, None)
            if cmd:
                documentation = cmd.__doc__.format(
                    command_prefix=self.config.command_prefix)
                em = Embed(title="**{}**".format(command.upper()))
                fields = documentation.split("///")
                if len(fields) < 2:  # backward compatibility
                    return Response(
                        "```\n{}```".format(dedent(cmd.__doc__).format(command_prefix=self.config.command_prefix)))

                for field in fields:
                    if field is None or field is "":
                        continue
                    inline = True
                    if field.startswith("(NL)"):
                        inline = False
                        field = field[4:]
                        # print(field)

                    match = re.match(r"\|(.+)\n((?:.|\n)+)", field)
                    if match is None:
                        continue
                    title, text = match.group(1, 2)

                    em.add_field(
                        name="**{}**".format(title), value=dedent(text), inline=inline)
                await self.send_message(channel, embed=em)
                return
            else:
                await self.send_typing(channel)
                params = {
                    "v": date.today().strftime("%d/%m/%y"),
                    "q": command}
                headers = {
                    "Authorization": "Bearer CU4UAUCKWN37QLXHMBOYZ425NOGBMIYK"}
                resp = requests.get("https://api.wit.ai/message",
                                    params=params, headers=headers)
                data = resp.json()
                entities = data["entities"]

                return Response(json.dumps(entities, indent=4))

        else:
            em = Embed(
                title="GIESELA HELP",
                url="http://siku2.github.io/Giesela/",
                colour=hex_to_dec("#828c51"),
                description="Here are some of the most useful commands,\nYou can always use `{0}help <cmd>` to get more detailed information on a command".
                format(self.config.command_prefix))

            music_commands = "\n".join([
                "`{0}play` play music",
                "`{0}search` search for music",
                "`{0}stream` enqueue a livestream",
                "`{0}pause` pause playback",
                "`{0}resume` resume playback"
                "`{0}volume` change volume",
                "`{0}seek` seek to a timestamp",
                "`{0}fwd` forward time",
                "`{0}rwd` rewind time"
            ]).format(self.config.command_prefix)
            em.add_field(name="Music", value=music_commands, inline=False)

            queue_commands = "\n".join([
                "`{0}queue` show the queue",
                "`{0}history` show playback history",
                "`{0}np` more information on the current entry",
                "`{0}skip` skip to the next entry in queue",
                "`{0}replay` replay the current entry",
                "`{0}repeat` change repeat mode",
                "`{0}remove` remove entry from queue",
                "`{0}clear` remove all entries from queue",
                "`{0}shuffle` shuffle the queue",
                "`{0}promote` promote entry to front"
            ]).format(self.config.command_prefix)
            em.add_field(name="Queue", value=queue_commands, inline=False)

            playlist_commands = "\n".join([
                "`{0}playlist` create/edit/list playlists",
                "`{0}addtoplaylist` add entry to playlist",
                "`{0}removefromplaylist` remove entry from playlist"
            ]).format(self.config.command_prefix)
            em.add_field(name="Playlist",
                         value=playlist_commands, inline=False)

            misc_commands = "\n".join([
                "`{0}lyrics` retrieve lyrics for the current song",
                "`{0}random` choose between items",
                "`{0}game` play a game",
                "`{0}ask` ask a question",
                "`{0}c` chat with Giesela"
            ]).format(self.config.command_prefix)
            em.add_field(name="Misc", value=misc_commands, inline=False)

            return Response(embed=em)

    async def cmd_blacklist(self, message, user_mentions, option, something):
        """
        ///|Usage
        {command_prefix}blacklist [ + | - | add | remove ] @UserName [@UserName2 ...]
        ///|Explanation
        Add or remove users to the blacklist.
        """

        if not user_mentions:
            raise exceptions.CommandError("No users listed.", expire_in=20)

        if option not in ['+', '-', 'add', 'remove']:
            raise exceptions.CommandError(
                'Invalid option "%s" specified, use +, -, add, or remove' %
                option,
                expire_in=20)

        for user in user_mentions.copy():
            if user.id == self.config.owner_id:
                print(
                    "[Commands:Blacklist] The owner cannot be blacklisted.")
                user_mentions.remove(user)

        old_len = len(self.blacklist)

        if option in ['+', 'add']:
            self.blacklist.update(user.id for user in user_mentions)

            write_file(self.config.blacklist_file, self.blacklist)

            return Response(
                '%s users have been added to the blacklist' %
                (len(self.blacklist) - old_len),
                reply=True,
                delete_after=10)

        else:
            if self.blacklist.isdisjoint(user.id for user in user_mentions):
                return Response(
                    'none of those users are in the blacklist.',
                    reply=True,
                    delete_after=10)

            else:
                self.blacklist.difference_update(user.id
                                                 for user in user_mentions)
                write_file(self.config.blacklist_file, self.blacklist)

                return Response(
                    '%s users have been removed from the blacklist' %
                    (old_len - len(self.blacklist)),
                    reply=True,
                    delete_after=10)

    async def cmd_id(self, author, user_mentions):
        """
        ///|Usage
        {command_prefix}id [@user]
        ///|Explanation
        Tells the user their id or the id of another user.
        """
        if not user_mentions:
            return Response(
                'your id is `%s`' % author.id, reply=True)
        else:
            usr = user_mentions[0]
            return Response(
                "%s's id is `%s`" % (usr.name, usr.id),
                reply=True,
                delete_after=35)

    @owner_only
    async def cmd_joinserver(self, message, server_link=None):
        """
        Usage:
            {command_prefix}joinserver invite_link

        Asks the bot to join a server.  Note: Bot accounts cannot use invite links.
        """

        if self.user.bot:
            url = await self.generate_invite_link()
            return Response(
                "Bot accounts can't use invite links!  Click here to invite me: \n{}".
                format(url),
                reply=True,
                delete_after=30)

        try:
            if server_link:
                await self.accept_invite(server_link)
                return Response(":+1:")

        except:
            raise exceptions.CommandError(
                'Invalid URL provided:\n{}\n'.format(server_link),
                expire_in=30)

    @command_info("1.0.0", 1477180800, {
        "3.5.2": (1497712233, "Updated documentaion for this command"),
        "3.8.9": (1499461104, "Part of the `Giesenesis` rewrite"),
        "3.9.6": (1499879464, "Better error handling"),
        "3.9.7": (1499968174, "Added a placement parameter."),
        "4.0.0": (1499981166, "Added \"random\" as a possible placement parameter")
    })
    async def cmd_play(self, player, channel, author, leftover_args, song_url):
        """
        ///|Usage
        `{command_prefix}play <song link | query> [index | "last" | "next" | "random"]`
        ///|Explanation
        Adds the song to the queue.  If no link is provided, the first
        result from a youtube search is added to the queue.
        """

        placement = None

        last_arg = leftover_args.pop() if leftover_args else ""

        if last_arg.lower() in ["next", "now", "first"]:
            placement = 0
            song_url = ""
        elif last_arg.lower() in ["anywhere", "random"]:
            placement = "random"
        elif last_arg.isnumeric():
            placement = int(last_arg) - 1
            song_url = ""
        else:
            leftover_args.append(last_arg)

        with send_typing(self, channel):
            query = " ".join([*leftover_args, song_url.strip("<>")])

            try:
                entry = await player.playlist.get_entry_from_query(query, author=author, channel=channel)
            except BaseException as e:
                return Response("There was a tiny problem with your request:\n```\n{}\n```".format(e))

        if not entry:
            return Response("Couldn't find anything for me to add")

        if isinstance(entry, list):
            print("[PLAY] This is a playlist!")
            # playlist handling
            entries_added = 0
            entries_not_added = 0

            entry_generator = player.playlist.get_entries_from_urls_gen(
                *entry, author=author, channel=channel)

            total_entries = len(entry)
            progress_message = await self.safe_send_message(
                channel,
                "Parsing {} entries\n{} [0%]".format(
                    total_entries,
                    create_bar(0, length=20)
                )
            )
            times = []
            abs_start = time.time()
            start_time = abs_start

            progress_message_future = None

            async for ind, entry in entry_generator:
                if entry:
                    player.playlist._add_entry(entry, placement)
                    entries_added += 1
                else:
                    entries_not_added += 1

                times.append(time.time() - start_time)
                start_time = time.time()

                if not progress_message_future or progress_message_future.done():
                    avg_time = sum(times) / float(len(times))
                    entries_left = total_entries - ind - 1
                    expected_time = format_time(
                        avg_time * entries_left,
                        max_specifications=1,
                        unit_length=1
                    )
                    completion_ratio = (ind + 1) / total_entries

                    if progress_message_future:
                        progress_message = progress_message_future.result()

                    progress_message_future = asyncio.ensure_future(self.safe_edit_message(
                        progress_message,
                        "Parsing {} entr{} at {} entries/min\n{} [{}%]\n{} remaining".format(
                            entries_left,
                            "y" if entries_left == 1 else "ies",
                            round(60 / avg_time, 1),
                            create_bar(completion_ratio, length=20),
                            round(100 * completion_ratio),
                            expected_time
                        ),
                        keep_at_bottom=True
                    ))

            delta_time = time.time() - abs_start

            progress_message_future.cancel()
            await self.safe_delete_message(progress_message)
            return Response("Added {} entries to the queue\nSkipped {} entries\nIt took {} to add all entries".format(
                entries_added,
                entries_not_added,
                format_time(delta_time, unit_length=1)
            ))
        else:
            player.playlist._add_entry(entry, placement)
            return Response("Enqueued `{}`".format(entry.title))

    @command_info("2.0.2", 1482252120, {
        "3.5.2": (1497712808, "Updated help text")
    })
    async def cmd_stream(self, player, channel, author, song_url):
        """
        ///|Usage
        `{command_prefix}stream <media link>`
        ///|Explanation
        Enqueue a media stream.
        This could mean an actual stream like Twitch, Youtube Gaming or even a radio stream, or simply streaming
        media without predownloading it.
        """

        song_url = song_url.strip('<>')

        await self.send_typing(channel)
        await player.playlist.add_stream_entry(
            song_url, channel=channel, author=author)

        return Response(":+1:")

    @block_user
    @command_info("2.0.3", 1485523740, {
        "3.7.7": (1499018088, "radio selection looks good again"),
        "3.8.9": (1499535312, "Part of the `Giesenesis` rewrite")
    })
    async def cmd_radio(self, player, channel, author, leftover_args):
        """
        ///|Usage
        `{command_prefix}radio [station name]`
        ///|Random station
        `{command_prefix}radio random`
        ///|Explanation
        Play live radio.
        You can leave the parameters blank in order to get a tour around all the channels,
        you can specify the station you want to listen to or you can let the bot choose for you by entering \"random\"
        """
        if leftover_args:
            if leftover_args[0].lower().strip() == "random":
                station_info = RadioStations.get_random_station()
                await player.playlist.add_radio_entry(station_info, channel=channel, author=author)
                return Response(
                    "I choose\n**{.name}**".format(station_info.name))
            else:
                # try to find the radio station
                search_name = " ".join(leftover_args)
                station_info = RadioStations.get_station(
                    search_name.lower().strip())
                if station_info:
                    await player.playlist.add_radio_entry(station_info, channel=channel, author=author)
                    return Response("Your favourite:\n**{.name}**".format(station_info))

        # help the user find the right station

        def check(m):
            t = ["y", "yes", "yeah", "yep", "sure"]
            f = ["n", "no", "nope", "never"]

            m = m.content.lower().strip()
            return m in t or m in f

        possible_stations = RadioStations.get_all_stations()
        shuffle(possible_stations)

        interface_string = "**{0.name}**\n\nType `yes` or `no`"

        for station in possible_stations:
            msg = await self.safe_send_message(
                channel, interface_string.format(station))
            response = await self.wait_for_message(
                author=author, channel=channel, check=check)
            await self.safe_delete_message(msg)
            play_station = response.content.lower().strip() in [
                "y", "yes", "yeah", "yep", "sure"
            ]
            await self.safe_delete_message(response)

            if play_station:
                await player.playlist.add_radio_entry(station_info, channel=channel, author=author)
                return Response(
                    "There you go fam!\n**{.name}**".format(station))
            else:
                continue

    @block_user
    @command_info("1.0.0", 1477180800, {
        "3.5.2": (1497712233, "Updated documentaion for this command"),
        "3.5.9": (1497890999, "Revamped design and functions making this command more useful"),
        "3.6.1": (1497967505, "deleting messages when leaving search")
    })
    async def cmd_search(self, player, channel, author, leftover_args):
        """
        ///|Usage
        `{command_prefix}search [number] <query>`
        ///|Explanation
        Searches for a video and adds the one you choose.
        """

        if not leftover_args:
            return Response("Please specify a search query.")

        try:
            number = int(leftover_args[0])
            if number > 20:
                return Response("You musn't search for more than 20 videos")

            query = " ".join(leftover_args[1:])

            if not query:
                return Response("You have to specify the query too.")
        except:
            number = 5
            query = " ".join(leftover_args)

        search_query = "ytsearch{}:{}".format(number, query)

        search_msg = await self.safe_send_message(channel, "Searching for videos...")
        await self.send_typing(channel)

        try:
            info = await self.downloader.extract_info(
                player.playlist.loop,
                search_query,
                download=False,
                process=True)

        except Exception as e:
            await self.safe_edit_message(search_msg, str(e), send_if_fail=True)
            return
        else:
            await self.safe_delete_message(search_msg)

        if not info:
            return Response("No videos found.")

        result_string = "**Result {0}/{1}**\n{2}"
        interface_string = "**Commands:**\n`play` play this result\n`addtoplaylist <playlist name>` add this result to a playlist\n\n`n` next result\n`p` previous result\n`exit` abort and exit"

        current_result_index = 0
        total_results = len(info["entries"])

        while True:
            current_result = info["entries"][current_result_index]

            result_message = await self.safe_send_message(channel, result_string.format(current_result_index + 1, total_results, current_result["webpage_url"]))
            interface_message = await self.safe_send_message(channel, interface_string)
            response_message = await self.wait_for_message(100, author=author, channel=channel, check=lambda msg: msg.content.strip().lower().split()[0] in ("play", "addtoplaylist", "n", "p", "exit"))

            if not response_message:
                await self.safe_delete_message(result_message)
                await self.safe_delete_message(interface_message)
                await self.safe_delete_message(response_message)
                return Response("Aborting search. [Timeout]")

            content = response_message.content.strip()
            command, *args = content.lower().split()

            if command == "exit":
                await self.safe_delete_message(result_message)
                await self.safe_delete_message(interface_message)
                await self.safe_delete_message(response_message)
                return Response("Okay then. Search again soon")
            elif command in "np":
                # feels hacky but is actully genius
                current_result_index += {"n": 1, "p": -1}[command]
                current_result_index %= total_results
            elif command == "play":
                await self.send_typing(channel)
                await self.cmd_play(player, channel, author, [], current_result["webpage_url"])
                await self.safe_delete_message(result_message)
                await self.safe_delete_message(interface_message)
                await self.safe_delete_message(response_message)
                return Response("Alright, coming right up!")
            elif command == "addtoplaylist":
                await self.send_typing(channel)
                if len(args) < 1:
                    err_msg = await self.safe_send_message(channel, "You have to specify the playlist which you would like to add this result to")
                    await asyncio.sleep(3)
                    await self.safe_delete_message(err_msg)
                    await self.safe_delete_message(result_message)
                    await self.safe_delete_message(interface_message)
                    await self.safe_delete_message(response_message)
                    continue

                playlistname = args[0]
                add_entry = (await self.get_play_entry(player, current_result["webpage_url"], channel=channel, author=author))[0]

                if playlistname not in self.playlists.saved_playlists:
                    if len(playlistname) < 3:
                        err_msg = await self.safe_send_message(channel, "Your name is too short. Please choose one with at least three letters.")
                        await asyncio.sleep(3)
                        await self.safe_delete_message(err_msg)
                        await self.safe_delete_message(result_message)
                        await self.safe_delete_message(interface_message)
                        await self.safe_delete_message(response_message)
                        continue

                    self.playlists.set_playlist(
                        [add_entry], playlistname, author.id)
                    await self.safe_delete_message(result_message)
                    await self.safe_delete_message(interface_message)
                    await self.safe_delete_message(response_message)
                    return Response("Created a new playlist \"{}\" and added `{}`.".format(playlistname.title(),
                                                                                           add_entry.title))

                self.playlists.edit_playlist(
                    playlistname, player.playlist, new_entries=[add_entry])
                await self.safe_delete_message(result_message)
                await self.safe_delete_message(interface_message)
                await self.safe_delete_message(response_message)
                return Response("Added `{}` to playlist \"{}\".".format(add_entry.title, playlistname.title()))

            await self.safe_delete_message(result_message)
            await self.safe_delete_message(interface_message)
            await self.safe_delete_message(response_message)

    @block_user
    @command_info("3.8.4", 1499188226, {
        "3.8.9": (1499461647, "Part of the `Giesenesis` rewrite")
    })
    async def cmd_suggest(self, player, channel, author):
        """
        ///|Usage
        `{command_prefix}suggest`
        ///|Explanation
        Find similar videos to the current one
        """

        if not player.current_entry:
            return Response("Can't give you any suggestions when there's nothing playing.")

        if not isinstance(player.current_entry, YoutubeEntry):
            return Response("Can't provide any suggestions for this entry type")

        vidId = player.current_entry.video_id

        videos = get_related_videos(vidId)

        if not videos:
            return Response("Couldn't find anything.")

        result_string = "**Result {0}/{1}**\n{2}"
        interface_string = "**Commands:**\n`play` play this result\n\n`n` next result\n`p` previous result\n`exit` abort and exit"

        current_result_index = 0
        total_results = len(videos)

        while True:
            current_result = videos[current_result_index]

            result_message = await self.safe_send_message(channel, result_string.format(current_result_index + 1, total_results, current_result["url"]))
            interface_message = await self.safe_send_message(channel, interface_string)
            response_message = await self.wait_for_message(100, author=author, channel=channel, check=lambda msg: msg.content.strip().lower().split()[0] in ("play", "n", "p", "exit"))

            if not response_message:
                await self.safe_delete_message(result_message)
                await self.safe_delete_message(interface_message)
                await self.safe_delete_message(response_message)
                return Response("Aborting. [Timeout]")

            content = response_message.content.strip()
            command, *args = content.lower().split()

            if command == "exit":
                await self.safe_delete_message(result_message)
                await self.safe_delete_message(interface_message)
                await self.safe_delete_message(response_message)
                return Response("Okay then. Suggest again soon *(Sorry, I couldn't resist)*")
            elif command in "np":
                # feels hacky but is actully genius
                current_result_index += {"n": 1, "p": -1}[command]
                current_result_index %= total_results
            elif command == "play":
                await self.send_typing(channel)
                await self.cmd_play(player, channel, author, [], current_result["url"])
                await self.safe_delete_message(result_message)
                await self.safe_delete_message(interface_message)
                await self.safe_delete_message(response_message)
                return Response("Alright, coming right up!")

            await self.safe_delete_message(result_message)
            await self.safe_delete_message(interface_message)
            await self.safe_delete_message(response_message)

    @command_info("1.0.0", 1477180800, {
        "3.5.4": (1497721686, "Updating the looks of the \"now playing\" message and a bit of cleanup"),
        "3.6.2": (1498143480, "Updated design of default entry and included a link to the video"),
        "3.6.5": (1498152579, "Timestamp-entries now also include a thumbnail"),
        "3.8.9": (1499461647, "Part of the `Giesenesis` rewrite")
    })
    async def cmd_np(self, player, channel, server, message):
        """
        ///|Usage
        {command_prefix}np
        ///|Explanation
        Displays the current song in chat.
        """

        if player.current_entry:
            if self.server_specific_data[server]["last_np_msg"]:
                await self.safe_delete_message(
                    self.server_specific_data[server]["last_np_msg"])
                self.server_specific_data[server]["last_np_msg"] = None

            entry = player.current_entry
            em = None

            if isinstance(entry, RadioSongEntry):
                progress_ratio = entry.song_progress / \
                    (entry.song_duration or 1)
                desc = "{} `[{}/{}]`".format(
                    create_bar(progress_ratio, length=20),
                    to_timestamp(entry.song_progress),
                    to_timestamp(entry.song_duration)
                )
                foot = "🔴 Live from {}".format(entry.station_name)

                em = Embed(
                    title=entry.title,
                    description=desc,
                    url=entry.link,
                    colour=hex_to_dec("#a23dd1")
                )

                em.set_footer(text=foot)
                em.set_thumbnail(url=entry.cover)
                em.set_author(
                    name=entry.artist
                )
            elif isinstance(entry, RadioStationEntry):
                desc = "`{}`".format(
                    to_timestamp(player.progress)
                )
                foot = "🔴 Live from {}".format(entry.station_name)

                em = Embed(
                    title=entry.title,
                    description=desc,
                    url=entry.link,
                    colour=hex_to_dec("#be7621")
                )

                em.set_footer(text=foot)
                em.set_thumbnail(url=entry.cover)
            elif isinstance(entry, StreamEntry):
                desc = "🔴 Live [`{}`]".format(to_timestamp(player.progress))

                em = Embed(
                    title=entry.title,
                    description=desc,
                    colour=hex_to_dec("#a23dd1")
                )

            if isinstance(entry, SpotifyEntry):
                artist_name = " & ".join(
                    artist.name for artist in entry.artists[:2])
                artist_avatar = choice(entry.artists[:2]).image
                progress_ratio = player.progress / entry.end_seconds
                desc = "{} `[{}/{}]`".format(
                    create_bar(progress_ratio, length=20),
                    to_timestamp(player.progress),
                    to_timestamp(entry.end_seconds)
                )

                em = Embed(
                    title=entry.song_name,
                    description=desc,
                    url=entry.url,
                    colour=hex_to_dec("#F9FF6E")
                )

                em.set_thumbnail(url=entry.cover)
                em.set_author(
                    name=artist_name,
                    icon_url=artist_avatar
                )
                em.add_field(name="Album", value=entry.album.name)
            elif isinstance(entry, TimestampEntry):
                sub_entry = entry.current_sub_entry
                index = sub_entry["index"] + 1
                progress_ratio = sub_entry["progress"] / sub_entry["duration"]
                desc = "{} `[{}/{}]`".format(
                    create_bar(progress_ratio, length=20),
                    to_timestamp(sub_entry["progress"]),
                    to_timestamp(sub_entry["duration"])
                )
                foot = "{}{} sub-entry of \"{}\" [{}/{}]".format(
                    index,
                    ordinal(index),
                    entry.whole_title,
                    to_timestamp(player.progress),
                    to_timestamp(entry.end_seconds)
                )

                em = Embed(
                    title=sub_entry["name"],
                    description=desc,
                    url=entry.url,
                    colour=hex_to_dec("#00FFFF")
                )

                em.set_footer(text=foot)
                em.set_thumbnail(url=entry.thumbnail)
                if "playlist" in entry.meta:
                    pl = entry.meta["playlist"]
                    em.set_author(name=pl["name"])
                elif "author" in entry.meta:
                    author = entry.meta["author"]
                    em.set_author(
                        name=author.display_name,
                        icon_url=author.avatar_url
                    )
            elif isinstance(entry, YoutubeEntry):
                progress_ratio = player.progress / entry.end_seconds
                desc = "{} `[{}/{}]`".format(
                    create_bar(progress_ratio, length=20),
                    to_timestamp(player.progress),
                    to_timestamp(entry.end_seconds)
                )

                em = Embed(
                    title=entry.title,
                    description=desc,
                    url=entry.url,
                    colour=hex_to_dec("#a9b244")
                )

                em.set_thumbnail(url=entry.thumbnail)
                if "playlist" in entry.meta:
                    pl = entry.meta["playlist"]
                    em.set_author(name=pl["name"].title())
                elif "author" in entry.meta:
                    author = entry.meta["author"]
                    em.set_author(
                        name=author.display_name,
                        icon_url=author.avatar_url
                    )

            if em:
                self.server_specific_data[server]["last_np_msg"] = await self.safe_send_message(channel, embed=em)
        else:
            return Response(
                "There are no songs queued! Queue something with {}play.".
                format(self.config.command_prefix))

    async def cmd_summon(self, channel, author, voice_channel):
        """
        Usage:
            {command_prefix}summon

        Call the bot to the summoner's voice channel.
        """

        if not author.voice_channel:
            raise exceptions.CommandError('You are not in a voice channel!')

        voice_client = self.the_voice_clients.get(channel.server.id, None)
        if voice_client and voice_client.channel.server == author.voice_channel.server:
            await self.move_voice_client(author.voice_channel)
            return

        # move to _verify_vc_perms?
        chperms = author.voice_channel.permissions_for(
            author.voice_channel.server.me)

        if not chperms.connect:
            print("Cannot join channel \"%s\", no permission." %
                  author.voice_channel.name)
            return Response(
                "```Cannot join channel \"%s\", no permission.```" %
                author.voice_channel.name,
                delete_after=25)

        elif not chperms.speak:
            print("Will not join channel \"%s\", no permission to speak." %
                  author.voice_channel.name)
            return Response(
                "```Will not join channel \"%s\", no permission to speak.```" %
                author.voice_channel.name,
                delete_after=25)

        player = await self.get_player(author.voice_channel, create=True)

        if player.is_stopped:
            player.play()

        if self.config.auto_playlist:
            await self.on_player_finished_playing(player)

    @command_info("1.0.0", 1477180800, {
        "3.5.2": (1497712233, "Updated documentaion for this command"),
        "3.8.9": (1499461647, "Part of the `Giesenesis` rewrite")
    })
    async def cmd_pause(self, player):
        """
        ///|Usage
        `{command_prefix}pause`
        ///|Explanation
        Pause playback of the current song.
        """

        if player.is_playing:
            player.pause()

        else:
            return Response("Cannot pause what is not playing")

    @command_info("1.0.0", 1477180800, {
        "3.5.2": (1497712233, "Updated documentaion for this command"),
        "3.8.9": (1499461647, "Part of the `Giesenesis` rewrite")
    })
    async def cmd_resume(self, player):
        """
        ///|Usage
        `{command_prefix}resume`
        ///|Explanation
        Resumes playback of the current song.
        """

        if player.is_paused:
            player.resume()

        else:
            return Response("Hard to unpause something that's not pause, amirite?")

    async def cmd_shuffle(self, channel, player):
        """
        ///|Usage
        `{command_prefix}shuffle`
        ///|Explanation
        Shuffles the queue.
        """

        player.playlist.shuffle()

        cards = [":spades:", ":clubs:", ":hearts:", ":diamonds:"]
        hand = await self.send_message(channel, " ".join(cards))

        for x in range(4):
            await asyncio.sleep(0.6)
            shuffle(cards)
            await self.safe_edit_message(hand, " ".join(cards))

        await self.safe_delete_message(hand, quiet=True)
        return Response(":ok_hand:")

    @command_info("1.0.0", 1477180800, {
        "3.5.2": (1497712233, "Updated documentaion for this command")
    })
    async def cmd_clear(self, player, author):
        """
        ///|Usage
        `{command_prefix}clear`
        ///|Explanation
        Clears the queue.
        """

        player.playlist.clear()
        return Response(':put_litter_in_its_place:')

    @command_info("1.0.0", 1477180800, {
        "3.3.7": (1497471674, "adapted the new \"seek\" command instead of \"skipto\""),
        "3.5.2": (1497714839, "Removed all the useless permission stuff and updated help text"),
        "3.8.9": (1499461647, "Part of the `Giesenesis` rewrite")
    })
    async def cmd_skip(self, player, skip_amount=None):
        """
        ///|Usage
        `{command_prefix}skip [all]`
        ///|Explanation
        Skips the current song.
        When given the keyword "all", skips all timestamped-entries in the current timestamp-entry.
        """

        if player.is_stopped:
            return Response("Can't skip! The player is not playing!")

        if not player.current_entry:
            if player.playlist.peek():
                if player.playlist.peek()._is_downloading:
                    # print(player.playlist.peek()._waiting_futures[0].__dict__)
                    return Response("The next song ({}) is downloading, please wait.".format(player.playlist.peek().title))

                elif player.playlist.peek().is_downloaded:
                    return Response("Something strange is happening.")
                else:
                    return Response("Something odd is happening.")
            else:
                return Response("Something strange is happening.")

        if isinstance(player.current_entry, TimestampEntry) and (not skip_amount or skip_amount.lower() != "all"):
            return await self.cmd_seek(
                player,
                str(player.current_entry.current_sub_entry["end"])
            )

        player.skip()

    @command_info("1.0.0", 1477180800, {
        "3.5.2": (1497712233, "Updated documentaion for this command"),
        "3.8.8": (1499421755, "improved volume bar")
    })
    async def cmd_volume(self, message, player, leftover_args):
        """
        ///|Usage
        `{command_prefix}volume [+ | -][volume]`
        ///|Explanation
        Sets the playback volume. Accepted values are from 1 to 100.
        Putting + or - before the volume will make the volume change relative to the current volume.
        """

        new_volume = "".join(leftover_args)

        if not new_volume:
            bar_len = 20
            return Response("Current volume: {}%\n{}".format(
                int(player.volume * 100), create_bar(player.volume, 20)))

        relative = False
        special_operation = None
        if new_volume[0] in '+-':
            relative = True
        if new_volume[0] in '*/%':
            special_operation = new_volume[0]
            new_volume = new_volume[1:]

        try:
            new_volume = int(new_volume)

        except ValueError:
            raise exceptions.CommandError(
                '{} is not a valid number'.format(new_volume), expire_in=20)

        if relative:
            vol_change = new_volume
            new_volume += (player.volume * 100)

        if special_operation is not None:
            operations = {
                "*": lambda x, y: x * y,
                "/": lambda x, y: x / y,
                "%": lambda x, y: x % y,
            }
            op = operations[special_operation]
            new_volume = op(player.volume * 100, new_volume)

        old_volume = int(player.volume * 100)

        if 0 < new_volume <= 100:
            player.volume = new_volume / 100.0

            return Response(
                'updated volume from %d to %d' % (old_volume, new_volume),
                reply=True,
                delete_after=20)

        else:
            if relative:
                raise exceptions.CommandError(
                    'Unreasonable volume change provided: {}{:+} -> {}%.  Provide a change between {} and {:+}.'.
                    format(old_volume, vol_change, old_volume + vol_change,
                           1 - old_volume, 100 - old_volume),
                    expire_in=20)
            else:
                raise exceptions.CommandError(
                    'Unreasonable volume provided: {}%. Provide a value between 1 and 100.'.
                    format(new_volume),
                    expire_in=20)

    @command_info("1.0.0", 1477180800, {
        "3.5.1": (1497706997, "Queue doesn't show the current entry anymore, always shows the whole queue and a bit of cleanup"),
        "3.5.5": (1497795534, "Total time takes current entry into account"),
        "3.5.8": (1497825017, "Doesn't show the whole queue right away anymore, instead the queue command takes a quantity argument which defaults to 15"),
        "3.8.0": (1499110875, "Displaying real index of sub-entries (timestamp-entry)"),
        "3.8.9": (1499461647, "Part of the `Giesenesis` rewrite")
    })
    async def cmd_queue(self, channel, player, num="15"):
        """
        ///|Usage
        {command_prefix}queue [quantity]
        ///|Explanation
        Show the first 15 entries of the current song queue.
        One can specify the amount of entries to be shown.
        """

        try:
            quantity = int(num)

            if quantity < 1:
                return Response("Please provide a reasonable quantity")
        except ValueError:
            if num.lower() == "all":
                quantity = len(player.playlist.entries)
            else:
                return Response("Quantity must be a number")

        lines = []

        lines.append("**QUEUE**\n")

        if player.current_entry and isinstance(player.current_entry, TimestampEntry):
            sub_queue = player.current_entry.sub_queue
            sub_queue = [sub_entry for sub_entry in sub_queue if sub_entry[
                "start"] >= player.progress]
            for item in sub_queue:
                lines.append(
                    "            ►`{}.` **{}**".format(
                        item["index"] + 1,
                        nice_cut(item["name"], 35)
                    )
                )

        entries = list(player.playlist.entries)[:quantity]
        for i, item in enumerate(entries, 1):
            origin_text = ""
            if "playlist" in item.meta:
                origin_text = "from playlist **{}**".format(
                    item.meta["playlist"]["name"].title()
                )
            elif "author" in item.meta:
                origin_text = "by **{}**".format(
                    item.meta["author"].name
                )

            lines.append("`{}.` **{}** {}".format(
                i, nice_cut(item.title, 40), origin_text))

        if len(lines) < 2:
            return Response(
                "There are no songs queued! Use `{}help` to find out how to queue something.".
                format(self.config.command_prefix))

        total_time = sum(
            [entry.end_seconds for entry in player.playlist.entries])
        if player.current_entry:
            total_time += player.current_entry.end_seconds - player.progress

        lines.append(
            "\nShowing {} out of {} entr{}".format(
                len(entries),
                len(player.playlist.entries),
                "y" if len(entries) == 1 else "ies"
            )
        )
        lines.append(
            "**Total duration:** `{}`".format(
                format_time(total_time, True, 5, 2)
            )
        )

        return Response("\n".join(lines))

    @command_info("3.3.3", 1497197957, {
        "3.3.8": (1497474312, "added failsafe for player not currently playing something"),
        "3.5.8": (1497825334, "Adjusted design to look more like `queue`'s style"),
        "3.8.9": (1499465102, "Part of the `Giesenesis` rewrite"),
        "4.0.1": (1500346108, "Quantity parameter. Increased history limit")
    })
    async def cmd_history(self, channel, player, num="15"):
        """
        ///|Usage
        {command_prefix}history [quantity]
        ///|Explanation
        Show the last [quantity] songs. If [quantity] isn't provided, show back to 15 songs
        """

        try:
            quantity = int(num)

            if quantity < 1:
                return Response("Please provide a reasonable quantity")
        except ValueError:
            if num.lower() == "all":
                quantity = len(player.playlist.entries)
            else:
                return Response("Quantity must be a number")

        if not player.playlist.history:
            return Response("There IS no history")

        seconds_passed = player.progress if player.current_entry else 0

        lines = []
        for ind, entry in enumerate(player.playlist.history[:quantity], 1):
            finish_time = entry.meta.get("finish_time", None)
            if finish_time:
                seconds_passed = time.time() - finish_time
            lines.append(
                "`{}.` **{}** {} ago".format(
                    ind,
                    nice_cut(entry.title, 40),
                    format_time(seconds_passed, max_specifications=2)
                )
            )
            seconds_passed += entry.end_seconds

        return Response("\n".join(lines))

    async def cmd_clean(self,
                        message,
                        channel,
                        server,
                        author,
                        search_range=50):
        """
        Usage:
            {command_prefix}clean [range]

        Removes up to [range] messages the bot has posted in chat. Default: 50, Max: 1000
        """

        try:
            float(search_range)  # lazy check
            search_range = min(int(search_range) + 1, 1000)
        except:
            return Response(
                "enter a number.  NUMBER.  That means digits.  `15`.  Etc.",
                reply=True,
                delete_after=8)

        await self.safe_delete_message(message, quiet=True)

        def is_possible_command_invoke(entry):
            valid_call = any(
                entry.content.startswith(prefix)
                for prefix in [self.config.command_prefix])  # can be expanded
            return valid_call and not entry.content[1:2].isspace()

        delete_invokes = True
        delete_all = channel.permissions_for(
            author).manage_messages or self.config.owner_id == author.id

        def check(message):
            if is_possible_command_invoke(message) and delete_invokes:
                return delete_all or message.author == author
            return message.author == self.user

        if self.user.bot:
            if channel.permissions_for(server.me).manage_messages:
                deleted = await self.purge_from(
                    channel, check=check, limit=search_range, before=message)
                return Response(
                    'Cleaned up {} message{}.'.format(
                        len(deleted), 's' * bool(deleted)),
                    delete_after=15)

        deleted = 0
        async for entry in self.logs_from(
                channel, search_range, before=message):
            if entry == self.server_specific_data[channel.server][
                    'last_np_msg']:
                continue

            if entry.author == self.user:
                await self.safe_delete_message(entry)
                deleted += 1
                await asyncio.sleep(0.21)

            if is_possible_command_invoke(entry) and delete_invokes:
                if delete_all or entry.author == author:
                    try:
                        await self.delete_message(entry)
                        await asyncio.sleep(0.21)
                        deleted += 1

                    except discord.Forbidden:
                        delete_invokes = False
                    except discord.HTTPException:
                        pass

        return Response(
            'Cleaned up {} message{}.'.format(deleted, 's' * bool(deleted)),
            delete_after=15)

    async def cmd_listids(self, server, author, leftover_args, cat='all'):
        """
        Usage:
            {command_prefix}listids [categories]

        Lists the ids for various things.  Categories are:
           all, users, roles, channels
        """

        cats = ['channels', 'roles', 'users']

        if cat not in cats and cat != 'all':
            return Response(
                "Valid categories: " + ' '.join(['`%s`' % c for c in cats]),
                reply=True,
                delete_after=25)

        if cat == 'all':
            requested_cats = cats
        else:
            requested_cats = [cat] + [c.strip(',') for c in leftover_args]

        data = ['Your ID: %s' % author.id]

        for cur_cat in requested_cats:
            rawudata = None

            if cur_cat == 'users':
                data.append("\nUser IDs:")
                rawudata = [
                    '%s #%s: %s' % (m.name, m.discriminator, m.id)
                    for m in server.members
                ]

            elif cur_cat == 'roles':
                data.append("\nRole IDs:")
                rawudata = ['%s: %s' % (r.name, r.id) for r in server.roles]

            elif cur_cat == 'channels':
                data.append("\nText Channel IDs:")
                tchans = [
                    c for c in server.channels
                    if c.type == discord.ChannelType.text
                ]
                rawudata = ['%s: %s' % (c.name, c.id) for c in tchans]

                rawudata.append("\nVoice Channel IDs:")
                vchans = [
                    c for c in server.channels
                    if c.type == discord.ChannelType.voice
                ]
                rawudata.extend('%s: %s' % (c.name, c.id) for c in vchans)

            if rawudata:
                data.extend(rawudata)

        with BytesIO() as sdata:
            sdata.writelines(d.encode('utf8') + b'\n' for d in data)
            sdata.seek(0)

            await self.send_file(
                author,
                sdata,
                filename='%s-ids-%s.txt' % (server.name.replace(' ', '_'),
                                            cat))

        return Response(":mailbox_with_mail:")

    @owner_only
    async def cmd_setname(self, leftover_args, name):
        """
        Usage:
            {command_prefix}setname name

        Changes the bot's username.
        Note: This operation is limited by discord to twice per hour.
        """

        name = ' '.join([name, *leftover_args])

        try:
            await self.edit_profile(username=name)
        except Exception as e:
            raise exceptions.CommandError(e, expire_in=20)

        return Response(":ok_hand:")

    @owner_only
    async def cmd_setnick(self, server, channel, leftover_args, nick):
        """
        Usage:
            {command_prefix}setnick nick

        Changes the bot's nickname.
        """

        if not channel.permissions_for(server.me).change_nickname:
            raise exceptions.CommandError(
                "Unable to change nickname: no permission.")

        nick = ' '.join([nick, *leftover_args])

        try:
            await self.change_nickname(server.me, nick)
        except Exception as e:
            raise exceptions.CommandError(e, expire_in=20)

        return Response(":ok_hand:")

    @owner_only
    async def cmd_setavatar(self, message, url=None):
        """
        Usage:
            {command_prefix}setavatar [url]

        Changes the bot's avatar.
        Attaching a file and leaving the url parameter blank also works.
        """

        if message.attachments:
            thing = message.attachments[0]['url']
        else:
            thing = url.strip('<>')

        try:
            with aiohttp.Timeout(10):
                async with self.aiosession.get(thing) as res:
                    await self.edit_profile(avatar=await res.read())

        except Exception as e:
            raise exceptions.CommandError(
                "Unable to change avatar: %s" % e, expire_in=20)

        return Response(":ok_hand:")

    async def cmd_autoplay(self, player):
        """
        Usage:
            {command_prefix}autoplay
        Play from the autoplaylist.
        """

        if not self.config.auto_playlist:
            self.config.auto_playlist = True
            await self.on_player_finished_playing(player)
            return Response("Playing from the autoplaylist")
        else:
            self.config.auto_playlist = False
            return Response(
                "Won't play from the autoplaylist anymore")

        # await self.safe_send_message (channel, msgState)

    async def cmd_say(self, channel, message, leftover_args):
        """
        Usage:
            {command_prefix}say <message>
        Make the bot say something
        """

        await self.safe_delete_message(message)
        await self.safe_send_message(channel, " ".join(leftover_args))
        print(message.author.name + " made me say: \"" +
              " ".join(leftover_args) + "\"")

    async def cmd_c(self, author, channel, leftover_args):
        """
        Usage:
            {command_prefix}c <message>

        have a chat
        """
        if len(leftover_args) < 1:
            return Response("You need to actually say something...")

        cb, nick = self.chatters.get(author.id, (None, None))
        if cb is None:
            cb = CleverWrap("CCC8n_IXK43aOV38rcWUILmYUBQ")
            nick = random_line(ConfigDefaults.name_list).strip().title()
            self.chatters[author.id] = (cb, nick)

        await self.send_typing(channel)
        msgContent = " ".join(leftover_args)

        while True:
            answer = cb.say(msgContent)
            answer = re.sub(r"\b[C|c]leverbot\b", "you", answer)
            answer = re.sub(r"\b[C|c][B|b]\b", "you", answer)
            base_answer = re.sub("[^a-z| ]+|\s{2,}", "", answer.lower())
            if base_answer not in "whats your name;what is your name;tell me your name".split(
                    ";") and not any(
                        q in base_answer
                        for q in
                        "whats your name; what is your name;tell me your name".
                        split(";")):
                break

        await asyncio.sleep(len(answer) / 5.5)
        print("<" + str(author.name) + "> " + msgContent + "\n<Bot> " +
              answer + "\n")
        return Response(answer)

    @command_info("1.9.5", 1477774380, {
        "3.6.1": (1497971656, "Fixed broken line wrap")
    })
    async def cmd_ask(self, author, channel, message, leftover_args):
        """
        ///|Usage
        `{command_prefix}ask <query>`
        ///|Explanation
        You can ask anything from science, maths, to culture
        """

        await self.send_typing(channel)
        msgContent = " ".join(leftover_args)

        col = choice(
            [9699539, 4915330, 255, 65280, 16776960, 16744192, 16711680])

        client = Tungsten("EH8PUT-67PJ967LG8")
        res = client.query(msgContent)
        if not res.success:
            await self.safe_send_message(
                channel,
                "Nothing found!"
            )

        for pod in res.pods:
            em = Embed(title=pod.title, colour=col)
            em.set_image(url=pod.format["img"][0]["url"])
            em.set_footer(text=pod.format["img"][0]["alt"])
            await self.send_message(channel, embed=em)

    @command_info("1.0.0", 1477180800, {
        "2.0.2": (1481827560, "Can now use @mentions to \"goto\" a user")
    })
    async def cmd_goto(self, server, channel, user_mentions, author, leftover_args):
        """
        Usage:
            {command_prefix}goto <id | name | @mention>

        Call the bot to a channel.
        """

        channelID = " ".join(leftover_args)
        if channelID.lower() == "home":
            await self.goto_home(server)
            return Response("yep")

        targetChannel = self.get_channel(channelID)
        if targetChannel is None:
            for chnl in server.channels:
                if chnl.name == channelID and chnl.type == ChannelType.voice:
                    targetChannel = chnl
                    break
            else:
                if user_mentions:
                    for ch in server.channels:
                        for user in ch.voice_members:
                            if user in user_mentions:
                                targetChannel = ch
                    if targetChannel is None:
                        return Response(
                            "Cannot find **{}** in any voice channel".format(
                                ", ".join([x.mention for x in user_mentions])))
                else:
                    print("Cannot find channel \"%s\"" % channelID)
                    return Response(
                        "```Cannot find channel \"%s\"```" % channelID)

        voice_client = await self.get_voice_client(targetChannel)
        print("Will join channel \"%s\"" % targetChannel.name)
        await self.move_voice_client(targetChannel)

        # move to _verify_vc_perms?
        chperms = targetChannel.permissions_for(targetChannel.server.me)

        if not chperms.connect:
            print("Cannot join channel \"%s\", no permission." %
                  targetChannel.name)
            return Response(
                "```Cannot join channel \"%s\", no permission.```" %
                targetChannel.name)

        elif not chperms.speak:
            print("Will not join channel \"%s\", no permission to speak." %
                  targetChannel.name)
            return Response(
                "```Will not join channel \"%s\", no permission to speak.```" %
                targetChannel.name)

        player = await self.get_player(targetChannel, create=True)

        if player.is_stopped:
            player.play()

        if self.config.auto_playlist:
            await self.on_player_finished_playing(player)

        return Response("Joined the channel **{}**".format(targetChannel.name))

    async def goto_home(self, server, join=True):
        channel = find(lambda c: c.type == ChannelType.voice and any(x in c.name.lower().split(
        ) for x in ["giesela", "musicbot", "bot", "music", "reign"]), server.channels)
        if channel is None:
            channel = choice(
                filter(lambda c: c.type == ChannelType.voice, server.channels))
        if join:
            await self.get_player(channel, create=True)
        return channel

    @command_info("1.9.5", 1477774380, {
        "3.4.2":
        (1497552134,
         "Added a way to not only replay the current song, but also the last one"
         ),
        "3.4.8": (1497649772, "Fixed the issue which blocked Giesela from replaying the last song"),
        "3.5.2": (1497714171, "Can now replay an index from the history"),
        "3.5.9": (1497899132, "Now showing the tile of the entry that is going to be replayed"),
        "3.6.0": (1497903889, "Replay <index> didn't work correctly"),
        "3.8.9": (1499466672, "Part of the `Giesenesis` rewrite")
    })
    async def cmd_replay(self, player, choose_last=""):
        """
        ///|Usage
        `{command_prefix}replay [last]`
        ///|Replay history
        `{command_prefix}replay <index>`
        Replay a song from the history
        ///|Explanation
        Replay the currently playing song. If there's nothing playing, or the \"last\" keyword is given, replay the last song
        """

        try:
            index = int(choose_last) - 1
            if index >= len(player.playlist.history):
                return Response("History doesn't go back that far.")
            if index < 0:
                return Response(
                    "Am I supposed to replay the future or what...?")

            replay_entry = player.playlist.history[index]
            player.playlist._add_entry_next(replay_entry)
            return Response("Replaying **{}**".format(replay_entry.title))
        except:
            pass

        replay_entry = player.current_entry
        if (not player.current_entry) or choose_last.lower() == "last":
            if not player.playlist.history:
                return Response(
                    "Cannot replay the last song as there is no last song")

            replay_entry = player.playlist.history[0]

        if not replay_entry:
            return Response("There's nothing for me to replay")

        player.playlist._add_entry_next(replay_entry)
        return Response("Replaying **{}**".format(replay_entry.title))

    @block_user
    @command_info("2.0.3", 1486054560, {
        "3.7.2": (1498252803, "no arguments provided crash Fixed")
    })
    async def cmd_random(self, channel, author, leftover_args):
        """
        ///|Basic
        `{command_prefix}random <item1>, <item2>, [item3], [item4]`
        ///|Use an existing set
        `{command_prefix}random <setname>`
        ///|List all the existing sets
        `{command_prefix}random list`
        ///|Creation
        `{command_prefix}random create <name>, <option1>, <option2>, [option3], [option4]`
        ///|Editing
        `{command_prefix}random edit <name>, [add | remove | replace], <item> [, item2, item3]`
        ///|Removal
        `{command_prefix}random remove <name>`
        ///|Explanation
        Choose a random item out of a list or use a pre-defined list.
        """

        if not leftover_args:
            return Response("Why u gotta be stupid?")

        items = [x.strip()
                 for x in " ".join(leftover_args).split(",") if x is not ""]

        if items[0].split()[0].lower().strip() == "create":
            if len(items) < 2:
                return Response(
                    "Can't create a set with the given arguments",
                    delete_after=20)

            set_name = "_".join(items[0].split()[1:]).lower().strip()
            set_items = items[1:]
            if self.random_sets.create_set(set_name, set_items):
                return Response(
                    "Created set **{0}**\nUse `{1}random {0}` to use it!".format(
                        set_name, self.config.command_prefix),
                    delete_after=60)
            else:
                return Response(
                    "OMG, shit went bad quickly! Everything's burning!\nDUCK there he goes again, the dragon's coming. Eat HIM not me. PLEEEEEEEEEEEEEASE!"
                )
        elif items[0].split()[0].lower().strip() == "list":
            return_string = ""
            for s in self.random_sets.get_sets():
                return_string += "**{}**\n```\n{}```\n\n".format(
                    s[0], ", ".join(s[1]))

            return Response(return_string)
        elif items[0].split()[0].lower().strip() == "edit":
            if len(items[0].split()) < 2:
                return Response(
                    "Please provide the name of the list you wish to edit!",
                    delete_after=20)

            set_name = "_".join(items[0].split()[1:]).lower().strip()

            existing_items = self.random_sets.get_set(set_name)
            if existing_items is None:
                return Response("This set does not exist!")

            edit_mode = items[1].strip().lower() if len(items) > 1 else None
            if edit_mode is None:
                return Response(
                    "You need to provide the way you want to edit the list",
                    delete_after=20)

            if len(items) < 3:
                return Response(
                    "You have to specify the items you want to add/remove or set as the new items"
                )

            if edit_mode == "add":
                for option in items[2:]:
                    self.random_sets.add_option(set_name, option)
            elif edit_mode == "remove":
                for option in items[2:]:
                    self.random_sets.remove_option(set_name, option)
            elif edit_mode == "replace":
                self.random_sets.replace_options(set_name, items[2:])
            else:
                return Response(
                    "This is not a valid edit mode!")

            return Response("Edited your set!")
        elif items[0].split()[0].lower().strip() == "remove":
            set_name = "_".join(items[0].split()[1:]).lower().strip()
            set_items = items[1:]
            res = self.random_sets.remove_set(set_name, set_items)
            if res:
                return Response("Removed set!")
            elif res is None:
                return Response("No such set!")
            else:
                return Response(
                    "OMG, shit went bad quickly! Everything's burning!\nDUCK there he goes again, the dragon's coming. Eat HIM not me. PLEEEEEEEEEEEEEASE!"
                )

        if len(items) <= 0 or items is None:
            return Response(
                "Is your name \"{0}\" by any chance?\n(This is not how this command works. Use `{1}help random` to find out how not to be a stupid **{0}** anymore)".
                format(author.name, self.config.command_prefix),
                delete_after=30)

        if len(items) <= 1:
            # return Response("Only you could use `{1}random` for one item...
            # Well done, {0}!".format(author.name, self.config.command_prefix),
            # delete_after=30)

            query = "_".join(items[0].split())
            items = self.random_sets.get_set(query.lower().strip())
            if items is None:
                return Response("Something went wrong")

        await self.safe_send_message(channel,
                                     "I choose **" + choice(items) + "**")

    async def cmd_broadcast(self, server, message, leftover_args):
        """
        Usage:
            {command_prefix}broadcast message

        Broadcast a message to every user of the server
        """

        targetMembers = []
        msg = ""

        if len(message.mentions) > 0:
            print("Found mentions!")
            msg = " ".join(leftover_args[len(message.mentions):])
            for target in message.mentions:
                print("User " + str(target) + " added to recipients")
                targetMembers.append(target)

        for role in server.roles:
            if role.name == leftover_args[0] or role.id == leftover_args[0]:
                print("Found " + role.name +
                      " and will send the message to them")
                msg = " ".join(leftover_args[1:])

                for member in server.members:
                    for mRole in member.roles:
                        if member not in targetMembers and (
                                mRole.name == leftover_args[0] or
                                mRole.id == leftover_args[0]):
                            print("User " + str(member) +
                                  " added to recipients")
                            targetMembers.append(member)
                            break
                break

        if len(targetMembers) < 1:
            print(
                "Didn't find a recipient. Will send the message to everyone")
            targetMembers = server.members
            msg = " ".join(leftover_args)

        for m in targetMembers:
            if m.bot:
                continue

            print("Sent \"" + msg + "\" to " + str(m))
            await self.safe_send_message(m, msg)

    async def cmd_getvideolink(self, player, message, channel, author,
                               leftover_args):
        """
        Usage:
            {command_prefix}getvideolink (optional: pause video)

        Sends the video link that gets you to the current location of the bot. Use "pause video" as argument to help you sync up the video.
        """

        if not player.current_entry:
            await self.safe_send_message(
                channel,
                "Can't give you a link for FUCKING NOTHING",
                expire_in=15)
            return

        if "pause video" in " ".join(leftover_args).lower():
            player.pause()
            minutes, seconds = divmod(player.progress, 60)
            await self.safe_send_message(
                channel, player.current_entry.url + "#t={0}m{1}s".format(
                    minutes, seconds))
            msg = await self.safe_send_message(
                channel, "Resuming video in a few seconds!")
            await asyncio.sleep(1.5)

            for i in range(5, 0, -1):
                newMsg = "** %s **" if i <= 3 else "%s"
                newMsg %= str(i)

                msg = await self.safe_edit_message(
                    msg, newMsg, send_if_fail=True)
                await asyncio.sleep(1)

            msg = await self.safe_edit_message(
                msg, "Let's continue!", send_if_fail=True)
            player.resume()

        else:
            minutes, seconds = divmod(player.progress + 3, 60)
            await self.safe_send_message(
                channel, player.current_entry.url + "#t={0}m{1}s".format(
                    minutes, seconds))

    @command_info("1.0.0", 1477180800, {
        "3.8.9": (1499466672, "Part of the `Giesenesis` rewrite")
    })
    async def cmd_remove(self, player, message, channel, author, leftover_args):
        """
        Usage:
            {command_prefix}remove <index | start index | url> [end index]

        Remove a index or a url from the queue.
        """

        if not leftover_args:
            leftover_args = ["0"]

        if len(player.playlist.entries) < 0:
            return Response("There are no entries in the queue!")

        if len(leftover_args) >= 2:
            indices = (
                int(leftover_args[0]) - 1,
                int(leftover_args[1]) - 1
            )

            start_index = min(indices)
            end_index = max(indices)

            if start_index >= len(player.playlist.entries) or start_index < 0:
                return Response("The start index is out of bounds")
            if end_index >= len(player.playlist.entries) or end_index < 0:
                return Response("The end index is out of bounds")

            for i in range(end_index, start_index - 1, -1):
                del player.playlist.entries[i]

            return Response(
                "Removed {} entries from the queue".format(
                    end_index - start_index + 1)
            )

        try:
            index = int(leftover_args[0]) - 1

            if index > len(player.playlist.entries) - 1 or index < 0:
                return Response("This index cannot be found in the queue")

            video = player.playlist.entries[index].title
            del player.playlist.entries[index]
            return Response("Removed **{0}** from the queue".format(video))

        except:
            strindex = leftover_args[0]
            iteration = 1

            for entry in player.playlist.entries:
                print(
                    "Looking at {0}. [{1}]".format(entry.title, entry.url))

                if entry.title == strindex or entry.url == strindex:
                    print("Found {0} and will remove it".format(
                        leftover_args[0]))
                    await self.cmd_remove(player, message, channel, author,
                                          [iteration])
                    return
                iteration += 1

        return Response("Didn't find anything that goes by {0}".format(leftover_args[0]))

    @block_user
    async def cmd_cah(self, message, channel, author, leftover_args):
        """
        Usage:
            {command_prefix}cah create
            {command_prefix}cah join <token>
            {command_prefix}cah leave <token>

            {command_prefix}cah start <token>
            {command_prefix}cah stop <token>

        Play a cards against humanity game

        References:
            {command_prefix}help cards
                -learn how to create/edit cards
            {command_prefix}help qcards
                -learn about how to create/edit question cards
        """

        argument = leftover_args[0].lower() if len(leftover_args) > 0 else None

        if argument == "create":
            if self.cah.is_user_in_game(author.id):
                g = self.cah.get_game(author.id)
                return Response(
                    "You can't host a game if you're already in one\nUse `{}cah leave {}` to leave your current game".
                    format(self.config.command_prefix, g.token),
                    delete_after=15)

            token = self.cah.new_game(author.id)
            return Response(
                "Created a new game.\nUse `{0}cah join {1}` to join this game and\nwhen everyone's in use `{0}cah start {1}`".
                format(self.config.command_prefix, token),
                delete_after=1000)
        elif argument == "join":
            token = leftover_args[
                1].lower() if len(leftover_args) > 1 else None
            if token is None:
                return Response("You need to provide a token")

            if self.cah.is_user_in_game(author.id):
                g = self.cah.get_game_from_user_id(author.id)
                return Response(
                    "You can only be part of one game at a time!\nUse `{}cah leave {}` to leave your current game".
                    format(self.config.command_prefix, g.token),
                    delete_after=15)

            g = self.cah.get_game(token)

            if g is None:
                return Response(
                    "This game does not exist *shrugs*")

            if g.in_game(author.id):
                return Response(
                    "You're already in this game!")

            if self.cah.user_join_game(author.id, token):
                return Response("Successfully joined the game **{}**".format(
                    token.upper()))
            else:
                return Response(
                    "Failed to join game **{}**".format(token.upper()))
        elif argument == "leave":
            token = leftover_args[
                1].lower() if len(leftover_args) > 1 else None
            if token is None:
                return Response("You need to provide a token")

            g = self.cah.get_game(token)

            if g is None:
                return Response(
                    "This game does not exist *shrugs*")

            if not g.in_game(author.id):
                return Response(
                    "You're not part of this game!")

            if self.cah.player_leave_game(author.id, token):
                return Response(
                    "Successfully left the game **{}**".format(token.upper()))
            else:
                return Response(
                    "Failed to leave game **{}**".format(token.upper()))
        elif argument == "start":
            token = leftover_args[
                1].lower() if len(leftover_args) > 1 else None
            if token is None:
                return Response("You need to provide a token")

            g = self.cah.get_game(token)
            if g is None:
                return Response("This game does not exist!")

            if not g.is_owner(author.id):
                return Response(
                    "Only the owner may start a game!")

            if not g.enough_players():
                return Response(
                    "There are not enough players to start this game.\nUse `{}cah join {}` to join a game".
                    format(self.config.command_prefix, g.token),
                    delete_after=15)

            if not g.start_game():
                return Response(
                    "This game has already started!")
        elif argument == "stop":
            token = leftover_args[
                1].lower() if len(leftover_args) > 1 else None
            g = self.cah.get_game(token)
            if g is None:
                return Response("This game does not exist!")

            if not g.is_owner(author.id):
                return Response(
                    "Only the owner may stop a game!")

            self.cah.stop_game(g.token)
            return Response(
                "Stopped the game **{}**".format(token))

    @block_user
    async def cmd_cards(self, server, channel, author, message, leftover_args):
        """
        Usage:
            {command_prefix}cards list [@mention] [text | likes | occurences | date | random | id | author | none]
                -list all the available cards
            {command_prefix}cards create <text>
                -create a new card with text
            {command_prefix}cards edit <id> <new_text>
                -edit a card by its id
            {command_prefix}cards info <id>
                -Get more detailed information about a card
            {command_prefix}cards search <query>
                -Search for a card
            {command_prefix}cards delete <id>
                -Delete a question card

        Here you manage the non question cards
        """

        argument = leftover_args[0].lower() if len(leftover_args) > 0 else None

        if argument == "list":
            sort_modes = {"text": (lambda entry: entry.text, False, lambda entry: None), "random": None, "occurences": (lambda entry: entry.occurences, True, lambda entry: entry.occurences), "date": (
                lambda entry: entry.creation_date, True, lambda entry: prettydate(entry.creation_date)), "author": (lambda entry: entry.creator_id, False, lambda entry: self.get_global_user(entry.creator_id).name), "id": (lambda entry: entry.id, False, lambda entry: None), "likes": (lambda entry: entry.like_dislike_ratio, True, lambda entry: "{}%".format(int(entry.like_dislike_ratio * 100)))}

            cards = self.cah.cards.cards.copy(
            ) if message.mentions is None or len(message.mentions) < 1 else [
                x for x in self.cah.cards.cards.copy()
                if x.creator_id in [u.id for u in message.mentions]
            ]
            sort_mode = leftover_args[1].lower(
            ) if len(leftover_args) > 1 and leftover_args[1].lower(
            ) in sort_modes.keys() else "none"

            display_info = None

            if sort_mode == "random":
                shuffle(cards)
            elif sort_mode != "none":
                cards = sorted(
                    cards,
                    key=sort_modes[sort_mode][0],
                    reverse=sort_modes[sort_mode][1])
                display_info = sort_modes[sort_mode][2]

            await self.card_viewer(channel, author, cards, display_info)
        elif argument == "search":
            search_query = " ".join(
                leftover_args[1:]) if len(leftover_args) > 1 else None

            if search_query is None:
                return Response(
                    "You need to provide a query to search for!",
                    delete_after=15)

            results = self.cah.cards.search_card(search_query, 3)

            if len(results) < 1:
                return Response("**Didn't find any cards!**")

            card_string = "{0.id}. \"{1}\""
            cards = []
            for card in results:
                cards.append(
                    card_string.format(card, card.text.replace("$", "_____")))

            return Response(
                "**I found the following cards:**\n\n" + "\n".join(cards),
                delete_after=40)
        elif argument == "info":
            card_id = leftover_args[
                1].lower().strip() if len(leftover_args) > 1 else None

            card = self.cah.cards.get_card(card_id)
            if card is not None:
                info = "Card **{0.id}** by {1}\n```\n\"{0.text}\"\nused {0.occurences} time{2}\ndrawn {0.picked_up_count} time{5}\nliked by {6}% of players\ncreated {3}```\nUse `{4}cards edit {0.id}` to edit this card"
                return Response(
                    info.format(card,
                                self.get_global_user(card.creator_id).mention,
                                "s" if card.occurences != 1 else "",
                                prettydate(card.creation_date), self.config.
                                command_prefix, "s" if card.picked_up_count !=
                                1 else "", int(card.like_dislike_ratio * 100)))

            return Response(
                "There's no card with that id. Use `{}cards list` to list all the possible cards".
                format(self.config.command_prefix))
        elif argument == "create":
            text = " ".join(
                leftover_args[1:]) if len(leftover_args) > 1 else None
            if text is None:
                return Response(
                    "You might want to actually add some text to your card",
                    delete_after=20)
            if len(text) < 3:
                return Response(
                    "I think that's a bit too short...")
            if len(text) > 140:
                return Response("Maybe a bit too long?")

            already_has_card, card = self.cah.cards.card_with_text(text)
            if already_has_card:
                return Response(
                    "There's already a card with a fairly similar content. <{0}>\nUse `{1}cards info {0}` to find out more about this card".
                    format(card.id, self.config.command_prefix))

            card_id = self.cah.cards.add_card(text, author.id)
            return Response("Successfully created card **{}**".format(card_id))
        elif argument == "edit":
            card_id = leftover_args[
                1].lower().strip() if len(leftover_args) > 1 else None

            try:
                card_id_value = int(card_id)
            except:
                return Response("An id must be a number")

            if card_id is None:
                return Response(
                    "You need to provide the card's id!")

            text = " ".join(
                leftover_args[2:]) if len(leftover_args) > 1 else None
            if text is None:
                return Response(
                    "You might want to actually add some text to your card",
                    delete_after=20)
            if len(text) < 3:
                return Response(
                    "I think that's a bit too short...")
            if len(text) > 140:
                return Response("Maybe a bit too long?")

            already_has_card, card = self.cah.cards.card_with_text(text)
            if already_has_card and card.id != card_id_value:
                return Response(
                    "There's already a card with a fairly similar content. <{0}>\nUse `{1}cards info {0}` to find out more about this card".
                    format(card.id, self.config.command_prefix))

            if self.cah.cards.edit_card(card_id, text):
                return Response(
                    "Edited card <**{}**>".format(card_id))
            else:
                return Response(
                    "There's no card with that id")
        elif argument == "delete":
            card_id = leftover_args[
                1].lower().strip() if len(leftover_args) > 1 else None

            if card_id is None:
                return Response(
                    "You must specify the card id")

            if self.cah.cards.remove_card(card_id):
                return Response(
                    "Deleted card <**{}**>".format(card_id))
            else:
                return Response(
                    "Could not remove card <**{}**>".format(card_id),
                    delete_after=15)
        else:
            return await self.cmd_help(channel, ["cards"])

    async def card_viewer(self,
                          channel,
                          author,
                          cards,
                          display_additional=None):
        cmds = ("n", "p", "exit")
        site_interface = "**Cards | Page {0} of {1}**\n```\n{2}\n```\nShit you can do:\n`n`: Switch to the next page\n`p`: Switch to the previous page\n`exit`: Exit the viewer"
        card_string = "<{}> [{}]{}"

        items_per_page = 20
        timeout = 60
        current_page = 0

        total_pages, items_on_last_page = divmod(
            len(cards) - 1, items_per_page)

        def msg_check(msg):
            return msg.content.lower().strip().startswith(cmds)

        while True:
            start_index = current_page * items_per_page
            end_index = start_index + \
                (items_per_page - 1 if current_page <
                 total_pages else items_on_last_page)
            page_cards = cards[start_index:end_index]

            page_cards_texts = []
            for p_c in page_cards:
                page_cards_texts.append(
                    card_string.format(
                        p_c.id, p_c.text, "" if display_additional is None or
                        display_additional(p_c) is None else " | {}".format(
                            display_additional(p_c))))

            interface_msg = await self.safe_send_message(
                channel,
                site_interface.format(current_page + 1, total_pages + 1,
                                      "\n".join(page_cards_texts)))
            user_msg = await self.wait_for_message(
                timeout, author=author, channel=channel, check=msg_check)

            if not user_msg:
                await self.safe_delete_message(interface_msg)
                break

            content = user_msg.content.lower().strip()

            if content.startswith("n"):
                await self.safe_delete_message(interface_msg)
                await self.safe_delete_message(user_msg)
                current_page = (current_page + 1) % (total_pages + 1)
            elif content.startswith("p"):
                await self.safe_delete_message(interface_msg)
                await self.safe_delete_message(user_msg)
                current_page = (current_page - 1) % (total_pages + 1)
            elif content.startswith("exit"):
                await self.safe_delete_message(interface_msg)
                await self.safe_delete_message(user_msg)
                break

        await self.safe_send_message(
            channel, "Closed the card viewer!", expire_in=20)

    @block_user
    async def cmd_qcards(self, server, channel, author, message,
                         leftover_args):
        """
        Usage:
            {command_prefix}qcards list [@mention] [text | likes | occurences | date | author | id | blanks | random | none]
                -list all the available question cards
            {command_prefix}qcards create <text (use $ for blanks)>
                -create a new question card with text and if you want the number of cards to draw
            {command_prefix}qcards edit <id> <new_text>
                -edit a question card by its id
            {command_prefix}qcards info <id>
                -Get more detailed information about a question card
            {command_prefix}qcards search <query>
                -Search for a question card
            {command_prefix}qcards delete <id>
                -Delete a question card

        Here you manage the question cards
        """

        argument = leftover_args[0].lower() if len(leftover_args) > 0 else None

        if argument == "list":
            sort_modes = {"text": (lambda entry: entry.text, False, lambda entry: None), "random": None, "occurences": (lambda entry: entry.occurences, True, lambda entry: entry.occurences), "date": (lambda entry: entry.creation_date, True, lambda entry: prettydate(entry.creation_date)), "author": (lambda entry: entry.creator_id, False, lambda entry: self.get_global_user(
                entry.creator_id).name), "id": (lambda entry: entry.id, False, lambda entry: None), "blanks": (lambda entry: entry.number_of_blanks, True, lambda entry: entry.number_of_blanks), "likes": (lambda entry: entry.like_dislike_ratio, True, lambda entry: "{}%".format(int(entry.like_dislike_ratio * 100)))}

            cards = self.cah.cards.question_cards.copy(
            ) if message.mentions is None or len(message.mentions) < 1 else [
                x for x in self.cah.cards.question_cards.copy()
                if x.creator_id in [u.id for u in message.mentions]
            ]
            sort_mode = leftover_args[1].lower(
            ) if len(leftover_args) > 1 and leftover_args[1].lower(
            ) in sort_modes.keys() else "none"

            display_info = None

            if sort_mode == "random":
                shuffle(cards)
            elif sort_mode != "none":
                cards = sorted(
                    cards,
                    key=sort_modes[sort_mode][0],
                    reverse=sort_modes[sort_mode][1])
                display_info = sort_modes[sort_mode][2]

            await self.qcard_viewer(channel, author, cards, display_info)
        elif argument == "search":
            search_query = " ".join(
                leftover_args[1:]) if len(leftover_args) > 1 else None

            if search_query is None:
                return Response(
                    "You need to provide a query to search for!",
                    delete_after=15)

            results = self.cah.cards.search_question_card(search_query, 3)

            if len(results) < 1:
                return Response(
                    "**Didn't find any question cards!**")

            card_string = "{0.id}. \"{1}\""
            cards = []
            for card in results:
                cards.append(
                    card_string.format(card,
                                       card.text.replace("$", "\_\_\_\_\_")))

            return Response(
                "**I found the following question cards:**\n\n" +
                "\n".join(cards),
                delete_after=40)
        elif argument == "info":
            card_id = leftover_args[
                1].lower().strip() if len(leftover_args) > 1 else None

            card = self.cah.cards.get_question_card(card_id)
            if card is not None:
                info = "Question Card **{0.id}** by {1}\n```\n\"{0.text}\"\nused {0.occurences} time{2}\ncreated {3}```\nUse `{4}cards edit {0.id}` to edit this card`"
                return Response(
                    info.format(card,
                                self.get_global_user(card.creator_id).mention,
                                "s" if card.occurences != 1 else "",
                                prettydate(card.creation_date),
                                self.config.command_prefix))
        elif argument == "create":
            text = " ".join(
                leftover_args[1:]) if len(leftover_args) > 1 else None
            if text is None:
                return Response(
                    "You might want to actually add some text to your card",
                    delete_after=20)
            if len(text) < 3:
                return Response(
                    "I think that's a bit too short...")
            if len(text) > 500:
                return Response("Maybe a bit too long?")

            if text.count("$") < 1:
                return Response(
                    "You need to have at least one blank ($) space",
                    delete_after=20)

            already_has_card, card = self.cah.cards.question_card_with_text(
                text)
            if already_has_card:
                return Response(
                    "There's already a question card with a fairly similar content. <{0}>\nUse `{1}qcards info {0}` to find out more about this card".
                    format(card.id, self.config.command_prefix))

            card_id = self.cah.cards.add_question_card(text, author.id)
            return Response(
                "Successfully created question card **{}**".format(card_id))
        elif argument == "edit":
            card_id = leftover_args[
                1].lower().strip() if len(leftover_args) > 1 else None

            try:
                card_id_value = int(card_id)
            except:
                return Response("An id must be a number")

            if card_id is None:
                return Response(
                    "You need to provide the question card's id!",
                    delete_after=20)

            text = " ".join(
                leftover_args[2:]) if len(leftover_args) > 2 else None
            if text is None:
                return Response(
                    "You might want to actually add some text to your question card",
                    delete_after=20)
            if len(text) < 3:
                return Response(
                    "I think that's a bit too short...")
            if len(text) > 500:
                return Response("Maybe a bit too long?")

            if text.count("$") < 1:
                return Response(
                    "You need to have at least one blank ($) space",
                    delete_after=20)

            already_has_card, card = self.cah.cards.question_card_with_text(
                text)
            if already_has_card and card.id != card_id_value:
                return Response(
                    "There's already a question card with a fairly similar content. <{0}>\nUse `{1}qcards info {0}` to find out more about this question card".
                    format(card.id, self.config.command_prefix))

            if self.cah.cards.edit_question_card(card_id, text):
                return Response(
                    "Edited question card <**{}**>".format(card_id),
                    delete_after=15)
            else:
                return Response(
                    "There's no question card with that id")
        elif argument == "delete":
            card_id = leftover_args[
                1].lower().strip() if len(leftover_args) > 1 else None

            if card_id is None:
                return Response(
                    "You must specify the question card id")

            if self.cah.cards.remove_question_card(card_id):
                return Response(
                    "Deleted question card <**{}**>".format(card_id),
                    delete_after=15)
            else:
                return Response(
                    "Could not remove question card <**{}**>".format(card_id),
                    delete_after=15)
        else:
            return await self.cmd_help(channel, ["qcards"])

    async def qcard_viewer(self,
                           channel,
                           author,
                           cards,
                           display_additional=None):
        cmds = ("n", "p", "exit")
        site_interface = "**Question Cards | Page {0} of {1}**\n```\n{2}\n```\nShit you can do:\n`n`: Switch to the next page\n`p`: Switch to the previous page\n`exit`: Exit the viewer"
        card_string = "<{}> \"{}\"{}"

        items_per_page = 20
        timeout = 60
        current_page = 0

        total_pages, items_on_last_page = divmod(
            len(cards) - 1, items_per_page)

        def msg_check(msg):
            return msg.content.lower().strip().startswith(cmds)

        while True:
            start_index = current_page * items_per_page
            end_index = start_index + \
                (items_per_page - 1 if current_page <
                 total_pages else items_on_last_page)
            page_cards = cards[start_index:end_index]

            page_cards_texts = []
            for p_c in page_cards:
                page_cards_texts.append(
                    card_string.format(
                        p_c.id,
                        p_c.text.replace("$", "_____"), "" if
                        display_additional is None or display_additional(p_c)
                        is None else " | {}".format(display_additional(p_c))))

            interface_msg = await self.safe_send_message(
                channel,
                site_interface.format(current_page + 1, total_pages + 1,
                                      "\n".join(page_cards_texts)))
            user_msg = await self.wait_for_message(
                timeout, author=author, channel=channel, check=msg_check)

            if not user_msg:
                await self.safe_delete_message(interface_msg)
                break

            content = user_msg.content.lower().strip()

            if content.startswith("n"):
                await self.safe_delete_message(interface_msg)
                await self.safe_delete_message(user_msg)
                current_page = (current_page + 1) % (total_pages + 1)
            elif content.startswith("p"):
                await self.safe_delete_message(interface_msg)
                await self.safe_delete_message(user_msg)
                current_page = (current_page - 1) % (total_pages + 1)
            elif content.startswith("exit"):
                await self.safe_delete_message(interface_msg)
                await self.safe_delete_message(user_msg)
                break

        await self.safe_send_message(
            channel, "Closed the question card viewer!", expire_in=20)

    @block_user
    @command_info("1.9.5", 1478998740, {
        "2.0.2": (1481387640,
                  "Added Hangman game and generalised game hub command"),
        "3.5.2": (1497712233, "Updated documentaion for this command")
    })
    async def cmd_game(self,
                       message,
                       channel,
                       author,
                       leftover_args,
                       game=None):
        """
        ///|Usage
        `{command_prefix}game [name]`
        ///|Explanation
        Play a game
        ///|References
        Cards against humanity can be played with the `cah` command.
        Use `{command_prefix}help cah` to learn more
        """

        all_funcs = dir(self)
        all_games = list(filter(lambda x: re.search("^g_\w+", x), all_funcs))
        all_game_names = [x[2:] for x in all_games]
        game_list = [{
            "name":
            x[2:],
            "handler":
            getattr(self, x, None),
            "description":
            getattr(self, x, None).__doc__.strip(' \t\n\r')
        } for x in all_games]

        if message.mentions is not None and len(message.mentions) > 0:
            author = message.mentions[0]

        if game is None:
            shuffle(game_list)

            def check(m):
                return (m.content.lower() in ["y", "n", "exit"])

            for current_game in game_list:
                msg = await self.safe_send_message(
                    channel,
                    "How about this game:\n\n**{}**\n{}\n\nType `y`, `n` or `exit`".
                    format(current_game["name"], current_game["description"]))
                response = await self.wait_for_message(
                    100, author=author, channel=channel, check=check)

                if not response or response.content.startswith(
                        self.config.command_prefix) or response.content.lower(
                ).startswith('exit'):
                    await self.safe_delete_message(msg)
                    await self.safe_delete_message(response)
                    await self.safe_send_message(channel, "Nevermind then.")
                    return

                if response.content.lower() == "y":
                    await self.safe_delete_message(msg)
                    await self.safe_delete_message(response)
                    game = current_game["name"]
                    break

                await self.safe_delete_message(msg)
                await self.safe_delete_message(response)

            if game is None:
                await self.safe_send_message(
                    channel, "That was all of them.", expire_in=20)
                return

        game = game.lower()
        handler = getattr(self, "g_" + game.title(), None)
        if handler is None:
            return Response("There's no game like that...")

        await handler(author, channel, leftover_args)

    async def g_2048(self, author, channel, additional_args):
        """
        Join the same numbers and get to the 2048 tile!
        """

        save_code = additional_args[0] if len(additional_args) > 0 else None
        size = additional_args[1] if len(additional_args) > 1 else 5

        game = Game2048(size, save_code)
        game_running = True
        turn_index = 1
        cache_location = "cache/pictures/g2048_img" + str(author.id)

        def check(reaction, user):
            if reaction.custom_emoji:
                # self.log (str (reaction.emoji) + " is a custom emoji")
                # print("Ignoring my own reaction")
                return False

            if (str(reaction.emoji) in ("⬇", "➡", "⬆", "⬅") or
                    str(reaction.emoji).startswith("📽") or
                    str(reaction.emoji).startswith("💾")
                ) and reaction.count > 1 and user == author:
                return True

            # self.log (str (reaction.emoji) + " was the wrong type of
            # emoji")
            return False

        while game_running:
            direction = None
            turn_information = ""
            # self.log (str (game))

            await self.send_typing(channel)

            while direction is None:
                msg = await self.send_file(
                    channel,
                    game.getImage(cache_location) + ".png",
                    content="**2048**\n{} turn {}".format(
                        str(turn_index) +
                        ("th" if 4 <= turn_index % 100 <= 20 else {
                            1: "st",
                            2: "nd",
                            3: "rd"
                        }.get(turn_index % 10, "th")), turn_information))
                turn_information = ""
                await self.add_reaction(msg, "⬅")
                await self.add_reaction(msg, "⬆")
                await self.add_reaction(msg, "➡")
                await self.add_reaction(msg, "⬇")
                await self.add_reaction(msg, "📽")
                await self.add_reaction(msg, "💾")

                reaction, user = await self.wait_for_reaction(
                    check=check, message=msg)
                msg = reaction.message  # for some reason this has to be like this
                # self.log ("User accepted. There are " + str (len
                # (msg.reactions)) + " reactions. [" + ", ".join ([str
                # (r.count) for r in msg.reactions]) + "]")

                for reaction in msg.reactions:
                    if str(reaction.emoji) == "📽" and reaction.count > 1:
                        await self.send_file(
                            user,
                            game.getImage(cache_location) + ".gif",
                            content="**2048**\nYour replay:")
                        turn_information = "| *replay has been sent*"

                    if str(reaction.emoji) == "💾" and reaction.count > 1:
                        await self.safe_send_message(
                            user,
                            "The save code is: **{0}**\nUse `{1}game 2048 {2}` to continue your current game".
                            format(
                                escape_dis(game.get_save()),
                                self.config.command_prefix, game.get_save()))
                        turn_information = "| *save code has been sent*"

                    if str(reaction.emoji) in ("⬇", "➡", "⬆",
                                               "⬅") and reaction.count > 1:
                        direction = ("⬇", "➡", "⬆",
                                     "⬅").index(str(reaction.emoji))

                    # self.log ("This did not match a direction: " + str
                    # (reaction.emoji))

                if direction is None:
                    await self.safe_delete_message(msg)
                    turn_information = "| You didn't specifiy the direction" if turn_information is not "" else turn_information

            # self.log ("Chose the direction " + str (direction))
            game.move(direction)
            turn_index += 1
            await self.safe_delete_message(msg)

            if game.won():
                await self.safe_send_message(
                    channel,
                    "**2048**\nCongratulations, you won after {} turns".format(
                        str(turn_index)))
                game_running = False

            if game.lost():
                await self.safe_send_message(
                    channel, "**2048**\nYou lost after {} turns".format(
                        str(turn_index)))
                game_running = False

        await self.send_file(
            channel,
            game.getImage(cache_location) + ".gif",
            content="**2048**\nYour replay:")
        await self.safe_delete_message(msg)

    async def g_Hangman(self, author, channel, additional_args):
        """
        Guess a word by guessing each and every letter
        """

        tries = additional_args[0] if len(additional_args) > 0 else 10

        word = additional_args[1] if len(additional_args) > 1 else re.sub(
            '[^a-zA-Z]', '', random_line(ConfigDefaults.hangman_wordlist))

        alphabet = list("abcdefghijklmnopqrstuvwxyz")
        print("Started a Hangman game with \"" + word + "\"")

        game = GameHangman(word, tries)
        running = True

        def check(m):
            return (m.content.lower() in alphabet or
                    m.content.lower() == word or m.content.lower() == "exit")

        while running:
            current_status = game.get_beautified_string()
            msg = await self.safe_send_message(
                channel,
                "**Hangman**\n{} trie{} left\n\n{}\n\n`Send the letter you want to guess or type \"exit\" to exit.`".
                format(game.tries_left, "s"
                       if game.tries_left != 1 else "", current_status))
            response = await self.wait_for_message(
                300, author=author, channel=channel, check=check)

            if not response or response.content.lower().startswith(
                    self.config.command_prefix) or response.content.lower(
            ).startswith('exit'):
                await self.safe_delete_message(msg)
                await self.safe_send_message(
                    channel, "Aborting this Hangman game. Thanks for playing!")
                running = False

            if response.content.lower() == word:
                await self.safe_send_message(
                    channel,
                    "Congratulations, you got it!\nThe word is: *{}*".format(
                        word))
                return

            letter = response.content[0]
            game.guess(letter)

            if game.won:
                await self.safe_send_message(
                    channel,
                    "Congratulations, you got it!\nThe word is: *{}*".format(
                        word))
                running = False

            if game.lost:
                await self.safe_send_message(channel, "You lost!")
                running = False

            await self.safe_delete_message(msg)
            await self.safe_delete_message(response)

    async def cmd_9gag(self, channel, author, post_id):
        """
        ///|Usage
        `{command_prefix}9gag <id>`
        ///|Explanation
        Display the 9gag post with the specified id
        """

        post = get_post(post_id)
        if not post:
            return Response("couldn't find that 9gag post, sorreyyyy!")

        if post.content_type == ContentType.IMAGE:
            em = Embed(title=post.title, url=post.hyperlink, colour=9316352)
            em.set_author(name=author.display_name, icon_url=author.avatar_url)
            em.set_image(url=post.content_url)
            em.set_footer(text="{} upvotes | {} comments".format(
                post.upvotes, post.comment_count))

            await self.send_message(channel, embed=em)
        else:
            saveloc = "cache/pictures/9gag.gif"
            resp = requests.get(post.content_url)
            with open(saveloc, "wb+") as f:
                f.write(resp.content)
            clip = editor.VideoFileClip(saveloc)
            # clip.resize(.5)
            clip = video.fx.all.resize(clip, newsize=.55)
            clip.write_gif("cache/pictures/9gag.gif", fps=10)
            saveloc = "cache/pictures/9gag.gif"

            em = Embed(title=post.title, url=post.hyperlink, colour=9316352)
            em.set_author(name=author.display_name, icon_url=author.avatar_url)
            em.set_footer(text="{} upvotes | {} comments".format(
                post.upvotes, post.comment_count))

            await self.send_message(channel, embed=em)
            await self.send_file(channel, saveloc)

        for comment in post.comments[:3]:
            em = Embed(
                timestamp=comment.timestamp,
                colour=11946278,
                url=comment.permalink)
            em.set_author(
                name=comment.name,
                icon_url=comment.avatar,
                url=comment.profile_url)
            em.set_footer(text="{} upvotes | {} replies".format(
                comment.score, comment.reply_count))
            if comment.content_type == ContentType.TEXT:
                em.description = comment.content
            elif comment.content_type in (ContentType.IMAGE, ContentType.GIF):
                em.set_image(url=comment.content)

            await self.send_message(channel, embed=em)

    async def cmd_repeat(self, player):
        """
        ///|Usage
        `{command_prefix}repeat`
        ///|Explanation
        Cycles through the repeat options. Default is no repeat, switchable to repeat all or repeat current song.
        """

        if player.is_stopped:
            raise exceptions.CommandError(
                "Can't change repeat mode! The player is not playing!",
                expire_in=20)

        player.repeat()

        if player.is_repeatNone:
            return Response(":play_pause: Repeat mode: None")
        if player.is_repeatAll:
            return Response(":repeat: Repeat mode: All")
        if player.is_repeatSingle:
            return Response(
                ":repeat_one: Repeat mode: Single")

    async def cmd_promote(self, player, position=None):
        """
        ///|Usage
        `{command_prefix}promote [song position]`
        ///|Explanation
        Promotes the last song in the queue to the front.
        If you specify a position, it promotes the song at that position to the front.
        """

        if player.is_stopped:
            raise exceptions.CommandError(
                "Can't modify the queue! The player is not playing!",
                expire_in=20)

        length = len(player.playlist.entries)

        if length < 2:
            raise exceptions.CommandError(
                "Can't promote! Please add at least 2 songs to the queue!",
                expire_in=20)

        if not position:
            entry = player.playlist.promote_last()
        else:
            try:
                position = int(position)
            except ValueError:
                raise exceptions.CommandError(
                    "This is not a valid song number! Please choose a song \
                    number between 2 and %s!" % length,
                    expire_in=20)

            if position == 1:
                raise exceptions.CommandError(
                    "This song is already at the top of the queue!",
                    expire_in=20)
            if position < 1 or position > length:
                raise exceptions.CommandError(
                    "Can't promote a song not in the queue! Please choose a song \
                    number between 2 and %s!" % length,
                    expire_in=20)

            entry = player.playlist.promote_position(position)

        reply_text = "Promoted **{}** to the :top: of the queue. Estimated time until playing: {}"
        btext = entry.title

        try:
            time_until = await player.playlist.estimate_time_until(1, player)
        except:
            traceback.print_exc()
            time_until = ""

        return Response(reply_text.format(btext, time_until))

    @block_user
    @command_info("1.9.5", 1479599760, {
        "3.4.6": (1497617827, "when Giesela can't add the entry to the playlist she tries to figure out **why** it didn't work"),
        "3.4.7": (1497619770, "Fixed an annoying bug in which the builder wouldn't show any entries if the amount of entries was a multiple of 20"),
        "3.5.1": (1497706811, "Giesela finally keeps track whether a certain entry comes from a playlist or not"),
        "3.5.8": (1497827857, "Default sort mode when loading playlists is now random and removing an entry in the playlist builder no longer messes with the current page."),
        "3.6.1": (1497969463, "when saving a playlist, list all changes"),
        "3.6.8": (1498162378, "checking whether start and end indices are numbers"),
        "3.6.9": (1498163686, "Special handling for sorting in playlist builder"),
        "3.7.0": (1498233256, "Changelog bug fixes"),
        "3.8.5": (1499279145, "Added \"rebuild\" extra command to clean and fix a playlist"),
        "3.8.7": (1499290119, "Due to a mistake \"rebuild\" always led to the deletion of the first entry."),
        "3.8.9": (1499525669, "Part of the `Giesenesis` rewrite"),
        "3.9.3": (1499712451, "Fixed a bug in the playlist builder search command."),
        "4.0.0": (1499978910, "Forgot to implement progress message properly and as a result it could bug out and spam itself.")
    })
    async def cmd_playlist(self, channel, author, server, player, leftover_args):
        """
        ///|Load
        `{command_prefix}playlist load <savename> [add | replace] [none | random] [startindex] [endindex (inclusive)]`\n\nTrust me, it's more complicated than it looks
        ///(NL)|List all playlists
        `{command_prefix}playlist showall [alphabetical | author | entries | playtime | random | replays]`
        ///(NL)|Build a new playlist
        `{command_prefix}playlist builder <savename>`
        ///(NL)|Save the current queue
        `{command_prefix}playlist save <savename>`
        ///(NL)|Clone
        `{command_prefix}playlist clone <fromname> <savename> [startindex | endindex (inclusive)]`
        ///(NL)|Delete a playlist
        `{command_prefix}playlist delete <savename>`
        ///(NL)|Information
        `{command_prefix}playlist <savename>`
        """

        argument = leftover_args[0].lower() if len(leftover_args) > 0 else ""
        savename = re.sub("\W", "", leftover_args[1].lower()) if len(
            leftover_args) > 1 else ""
        load_mode = leftover_args[2].lower() if len(
            leftover_args) > 2 else "add"
        additional_args = leftover_args[2:] if len(leftover_args) > 2 else []

        forbidden_savenames = [
            "showall", "savename", "save", "load", "delete", "builder",
            "extras", "add", "remove", "save", "exit", "clone", "rename",
            "extras", "alphabetical", "author", "entries", "playtime", "random"
        ]

        if argument == "save":
            if savename in self.playlists.saved_playlists:
                return Response(
                    "Can't save the queue, there's already a playlist with this name.")
            if len(savename) < 3:
                return Response(
                    "Can't save the queue, the name must be longer than 3 characters")
            if savename in forbidden_savenames:
                return Response(
                    "Can't save the queue, this name is forbidden!")
            if len(player.playlist.entries) < 1:
                return Response(
                    "Can't save the queue, there are no entries in the queue!")

            if self.playlists.set_playlist(
                [player.current_entry] + list(player.playlist.entries),
                    savename, author.id):
                return Response("Saved the current queue...")

            return Response(
                "Uhm, something went wrong I guess :D")

        elif argument == "load":
            if savename not in self.playlists.saved_playlists:
                return Response(
                    "Can't load this playlist, there's no playlist with this name.")

            playlist = self.playlists.get_playlist(
                savename, player.playlist, channel=channel)
            clone_entries = playlist["entries"]
            broken_entries = playlist["broken_entries"]

            if not clone_entries:
                if broken_entries:
                    return Response("Can't play `{0}`, there are **{1}** broken entr{2} in this playlist.\nOpen the playlist builder to fix {3} (`{4}playlist builder {0}`)".format(
                        savename.title(),
                        len(broken_entries),
                        "y" if len(broken_entries) == 1 else "ies",
                        "it" if len(broken_entries) == 1 else "them",
                        self.config.command_prefix
                    ))
                else:
                    return Response("There's nothing in `{}` to play".format(savename.title()))

            if load_mode == "replace":
                player.playlist.clear()
                if player.current_entry is not None:
                    player.skip()

            try:
                from_index = int(
                    additional_args[2]) - 1 if len(additional_args) > 2 else 0
                if from_index >= len(clone_entries) or from_index < 0:
                    return Response("Can't load the playlist starting from entry {}. This value is out of bounds.".format(from_index))
            except ValueError:
                return Response("Start index must be a number")

            try:
                to_index = int(additional_args[3]) if len(
                    additional_args) > 3 else len(clone_entries)
                if to_index > len(clone_entries) or to_index < 0:
                    return Response("Can't load the playlist from the {}. to the {}. entry. These values are out of bounds.".format(from_index, to_index))
            except ValueError:
                return Response("End index must be a number")

            if to_index - from_index <= 0:
                return Response("No songs to play. RIP.")

            clone_entries = clone_entries[from_index:to_index]

            sort_modes = {
                "alphabetical": (lambda entry: entry.title, False),
                "random": None,
                "length": (lambda entry: entry.duration, True)
            }

            sort_mode = additional_args[1].lower(
            ) if len(additional_args) > 1 and additional_args[1].lower(
            ) in sort_modes.keys() else "random"

            if sort_mode == "random":
                shuffle(clone_entries)
            elif sort_mode != "none":
                clone_entries = sorted(
                    clone_entries,
                    key=sort_modes[sort_mode][0],
                    reverse=sort_modes[sort_mode][1])

            player.playlist.add_entries(clone_entries)
            self.playlists.bump_replay_count(savename)

            if not broken_entries:
                return Response("Loaded `{}`".format(savename.title()))
            else:
                text = "Loaded {0} entr{1} from `{2}`. **{3}** entr{4} couldn't be loaded.\nOpen the playlist builder to repair {5}. (`{6}playlist builder {2}`)"
                return Response(text.format(
                    len(clone_entries),
                    "y" if len(clone_entries) == 1 else "ies",
                    savename.title(),
                    len(broken_entries),
                    "y" if len(broken_entries) == 1 else "ies",
                    "it" if len(broken_entries) == 1 else "them",
                    self.config.command_prefix
                ))

        elif argument == "delete":
            if savename not in self.playlists.saved_playlists:
                return Response(
                    "Can't delete this playlist, there's no playlist with this name.",
                    delete_after=20)

            self.playlists.remove_playlist(savename)
            return Response(
                "*{}* has been deleted".format(savename))

        elif argument == "clone":
            if savename not in self.playlists.saved_playlists:
                return Response(
                    "Can't clone this playlist, there's no playlist with this name.",
                    delete_after=20)
            clone_playlist = self.playlists.get_playlist(
                savename, player.playlist)
            clone_entries = clone_playlist["entries"]
            extend_existing = False

            if additional_args is None:
                return Response(
                    "Please provide a name to save the playlist to",
                    delete_after=20)

            if additional_args[0].lower() in self.playlists.saved_playlists:
                extend_existing = True
            if len(additional_args[0]) < 3:
                return Response(
                    "This is not a valid playlist name, the name must be longer than 3 characters",
                    delete_after=20)
            if additional_args[0].lower() in forbidden_savenames:
                return Response(
                    "This is not a valid playlist name, this name is forbidden!",
                    delete_after=20)

            from_index = int(additional_args[1]) - \
                1 if len(additional_args) > 1 else 0
            if from_index >= len(clone_entries) or from_index < 0:
                return Response(
                    "Can't clone the playlist starting from entry {}. This entry is out of bounds.".
                    format(from_index),
                    delete_after=20)

            to_index = int(additional_args[
                2]) if len(additional_args) > 2 else len(clone_entries)
            if to_index > len(clone_entries) or to_index < 0:
                return Response(
                    "Can't clone the playlist from the {}. to the {}. entry. These values are out of bounds.".
                    format(from_index, to_index),
                    delete_after=20)

            if to_index - from_index <= 0:
                return Response(
                    "That's not enough entries to create a new playlist.",
                    delete_after=20)

            clone_entries = clone_entries[from_index:to_index]
            if extend_existing:
                self.playlists.edit_playlist(
                    additional_args[0].lower(),
                    player.playlist,
                    new_entries=clone_entries)
            else:
                self.playlists.set_playlist(
                    clone_entries, additional_args[0].lower(), author.id)

            return Response(
                "**{}** {}has been cloned to **{}**".format(
                    savename, "(from the {}. to the {}. index) ".format(
                        str(from_index + 1), str(to_index + 1)) if
                    from_index is not 0 or to_index is not len(clone_entries)
                    else "", additional_args[0].lower()),
                delete_after=20)

        elif argument == "showall":
            if len(self.playlists.saved_playlists) < 1:
                return Response(
                    "There are no saved playlists.\n**You** could add one though. Type `{}help playlist` to see how!".format(
                        self.config.command_prefix),
                    delete_after=40)

            response_text = "**Found the following playlists:**\n\n"
            iteration = 1

            sort_modes = {
                "alphabetical": (lambda playlist: playlist, False),
                "entries": (
                    lambda playlist: int(self.playlists.get_playlist(
                        playlist, player.playlist)["entry_count"]),
                    True
                ),
                "author": (
                    lambda playlist: self.get_global_user(
                        self.playlists.get_playlist(playlist, player.playlist)["author"]).name,
                    False
                ),
                "random": None,
                "playtime": (
                    lambda playlist: sum([x.duration for x in self.playlists.get_playlist(
                        playlist, player.playlist)["entries"]]),
                    True
                ),
                "replays": (
                    lambda playlist: self.playlists.get_playlist(
                        playlist, player.playlist)["replay_count"],
                    True
                )
            }

            sort_mode = leftover_args[1].lower(
            ) if len(leftover_args) > 1 and leftover_args[1].lower(
            ) in sort_modes.keys() else "random"

            if sort_mode == "random":
                sorted_saved_playlists = self.playlists.saved_playlists
                shuffle(sorted_saved_playlists)
            else:
                sorted_saved_playlists = sorted(
                    self.playlists.saved_playlists,
                    key=sort_modes[sort_mode][0],
                    reverse=sort_modes[sort_mode][1])

            for pl in sorted_saved_playlists:
                infos = self.playlists.get_playlist(pl, player.playlist)
                response_text += "**{}.** **\"{}\"** by {}\n```\n  {} entr{}\n  played {} time{}\n  {}```\n\n".format(
                    iteration,
                    pl.replace("_", " ").title(),
                    self.get_global_user(infos["author"]).mention,
                    str(infos["entry_count"]), "ies"
                    if int(infos["entry_count"]) is not 1 else "y",
                    infos["replay_count"], "s"
                    if int(infos["replay_count"]) != 1 else "",
                    format_time(
                        sum([x.duration for x in infos["entries"]]),
                        round_seconds=True,
                        max_specifications=2))
                iteration += 1

            # self.log (response_text)
            return Response(response_text)

        elif argument == "builder":
            if len(savename) < 3:
                return Response(
                    "Can't build on this playlist, the name must be longer than 3 characters",
                    delete_after=20)
            if savename in forbidden_savenames:
                return Response(
                    "Can't build on this playlist, this name is forbidden!",
                    delete_after=20)

            print("Starting the playlist builder")
            response = await self.playlist_builder(channel, author, server,
                                                   player, savename)
            return response

        elif argument in self.playlists.saved_playlists:
            infos = self.playlists.get_playlist(argument.lower(),
                                                player.playlist)
            entries = infos["entries"]

            desc_text = "{} entr{}\n{} long".format(
                str(infos["entry_count"]), "ies"
                if int(infos["entry_count"]) is not 1 else "y",
                format_time(
                    sum([x.duration for x in entries]),
                    round_seconds=True,
                    combine_with_and=True,
                    replace_one=True,
                    max_specifications=2))
            em = Embed(
                title=argument.replace("_", " ").title(),
                description=desc_text)
            pl_author = self.get_global_user(infos["author"])
            em.set_author(
                name=pl_author.display_name, icon_url=pl_author.avatar_url)

            for i in range(min(len(entries), 20)):
                em.add_field(
                    name="{0:>3}. {1:<50}".format(i + 1,
                                                  entries[i].title[:50]),
                    value="duration: " + format_time(
                        entries[i].duration,
                        round_seconds=True,
                        round_base=1,
                        max_specifications=2),
                    inline=False)

            if len(entries) > 20:
                em.add_field(
                    name="**And {} more**".format(len(entries) - 20),
                    value="To view them, open the playlist builder")

            em.set_footer(
                text="To edit this playlist type \"{}playlist builder {}\"".
                format(self.config.command_prefix, argument))

            await self.send_message(channel, embed=em)

            return

        return await self.cmd_help(channel, ["playlist"])

    async def playlist_builder(self, channel, author, server, player, _savename):
        if _savename not in self.playlists.saved_playlists:
            self.playlists.set_playlist([], _savename, author.id)

        def check(m):
            return (m.content.split()[0].lower() in ["add", "remove", "rename", "exit", "p", "n", "save", "extras", "search"])

        async def _get_entries_from_urls(urls, message):
            entries = []
            removed_entries = []

            entry_generator = player.playlist.get_entries_from_urls_gen(
                *urls)

            total_entries = len(urls)
            progress_message = await self.safe_send_message(channel, "{}\n{} [0%]".format(message.format(entries_left=total_entries), create_bar(0, length=20)))
            times = []
            start_time = time.time()

            progress_message_future = None

            async for ind, entry in entry_generator:
                if entry:
                    entries.append(entry)
                else:
                    removed_entries.append(ind)

                times.append(time.time() - start_time)
                start_time = time.time()

                if not progress_message_future or progress_message_future.done():
                    entries_left = total_entries - ind - 1
                    avg_time = sum(times) / float(len(times))
                    expected_time = avg_time * entries_left

                    if progress_message_future:
                        progress_message = progress_message_future.result()

                    progress_message_future = asyncio.ensure_future(self.safe_edit_message(
                        progress_message,
                        "{}\n{} [{}%]\n{} remaining".format(
                            message.format(entries_left=entries_left),
                            create_bar((ind + 1) / total_entries, length=20),
                            round(100 * (ind + 1) / total_entries),
                            format_time(
                                expected_time,
                                max_specifications=1,
                                combine_with_and=True,
                                unit_length=1
                            )
                        ),
                        keep_at_bottom=True
                    ))

            await progress_message_future
            await self.safe_delete_message(progress_message)
            return entries, removed_entries

        abort = False
        save = False
        entries_page = 0
        pl_changes = {
            "remove_entries": [],  # used for changelog
            "added_entries": [],  # changelog
            "order": None,  # changelog
            "new_name": None
        }
        savename = _savename
        user_savename = savename

        interface_string = "**{}** by **{}** ({} song{} with a total length of {})\n\n{}\n\n**You can use the following commands:**\n`add <query>`: Add a video to the playlist (this command works like the normal `{}play` command)\n`remove <index> [index 2] [index 3] [index 4]`: Remove a song from the playlist by it's index\n`rename <newname>`: rename the current playlist\n`search <query>`: search for an entry\n`extras`: see the special functions\n\n`p`: previous page\n`n`: next page\n`save`: save and close the builder\n`exit`: leave the builder without saving"

        extras_string = "**{}** by **{}** ({} song{} with a total length of {})\n\n**Extra functions:**\n`sort <alphabetical | length | random>`: sort the playlist (default is alphabetical)\n`removeduplicates`: remove all duplicates from the playlist\n`rebuild`: clean the playlist by removing broken videos\n\n`abort`: return to main screen"

        playlist = self.playlists.get_playlist(_savename, player.playlist)

        if playlist["broken_entries"]:
            broken_entries = playlist["broken_entries"]
            if len(broken_entries) > 1:
                m = "There are {entries_left} broken/outdated entries in this playlist. I'm going to fix them, please stand by."
                new_entries, hopeless_entries = await _get_entries_from_urls([entry["url"] for entry in broken_entries], m)
                playlist["entries"].extend(new_entries)
                if hopeless_entries:
                    await self.safe_send_message(channel, "I couldn't save the following entries\n{}".format(
                        "\n".join(
                            "**" + broken_entries[entry_index]["title"] + "**" for entry_index in hopeless_entries
                        )
                    ))

            else:
                broken_entry = broken_entries[0]
                info = await self.safe_send_message(channel, "**{}** is broken, please wait while I fix it for ya.".format(broken_entry["title"]))
                new_entry = await player.playlist.get_entry(broken_entry["url"])
                if not new_entry:
                    await self.safe_send_message(channel, "Couldn't safe **{}**".format(broken_entry["title"]))
                else:
                    playlist["entries"].append(new_entry)
                    await self.safe_delete_message(info)

        while (not abort) and (not save):
            entries = playlist["entries"]
            entries_text = ""

            items_per_page = 20
            iterations, overflow = divmod(len(entries), items_per_page)

            if iterations > 0 and overflow == 0:
                iterations -= 1
                overflow += items_per_page

            start = (entries_page * items_per_page)
            end = (start + (overflow if entries_page >= iterations else
                            items_per_page)) if len(entries) > 0 else 0

            for i in range(start, end):
                entries_text += str(i + 1) + ". " + entries[i].title + "\n"
            entries_text += "\nPage {} of {}".format(entries_page + 1,
                                                     iterations + 1)

            interface_message = await self.safe_send_message(
                channel,
                interface_string.format(
                    user_savename.replace("_", " ").title(),
                    self.get_global_user(playlist["author"]).mention,
                    playlist["entry_count"],
                    "s" if int(playlist["entry_count"]) is not 1 else "",
                    format_time(sum([x.duration for x in entries])),
                    entries_text,
                    self.config.command_prefix
                )
            )
            response_message = await self.wait_for_message(
                author=author, channel=channel, check=check)

            if not response_message:
                await self.safe_delete_message(interface_message)
                abort = True
                break

            elif response_message.content.lower().startswith(self.config.command_prefix) or response_message.content.lower().startswith('exit'):
                abort = True

            elif response_message.content.lower().startswith("save"):
                save = True

            split_message = response_message.content.split()
            arguments = split_message[1:] if len(split_message) > 1 else None

            if split_message[0].lower() == "add":
                if arguments is not None:
                    msg = await self.safe_send_message(channel, "I'm working on it.")
                    query = " ".join(arguments)
                    try:
                        start_time = datetime.now()
                        entry = await player.playlist.get_entry_from_query(query)
                        if isinstance(entry, list):
                            entries, _ = await _get_entries_from_urls(entry, "Parsing {entries_left} entries")
                        else:
                            entries = [entry, ]
                        if (datetime.now() - start_time).total_seconds() > 40:
                            await self.safe_send_message(author, "Wow, that took quite a while.\nI'm done now though so come check it out!")

                        pl_changes["added_entries"].extend(
                            entries)  # just for the changelog
                        playlist["entries"].extend(entries)
                        playlist["entry_count"] = str(
                            int(playlist["entry_count"]) + len(entries))
                        it, ov = divmod(
                            int(playlist["entry_count"]), items_per_page)
                        entries_page = it - 1 if ov == 0 else it
                    except Exception as e:
                        await self.safe_send_message(
                            channel,
                            "**Something went terribly wrong there:**\n```\n{}\n```".format(
                                e)
                        )
                    await self.safe_delete_message(msg)

            elif split_message[0].lower() == "remove":
                if arguments is not None:
                    indices = []
                    for arg in arguments:
                        try:
                            index = int(arg) - 1
                        except:
                            index = -1

                        if index >= 0 and index < int(playlist["entry_count"]):
                            indices.append(index)

                    pl_changes["remove_entries"].extend(
                        [(ind, playlist["entries"][ind]) for ind in indices])  # for the changelog
                    playlist["entry_count"] = str(
                        int(playlist["entry_count"]) - len(indices))
                    playlist["entries"] = [
                        playlist["entries"][x]
                        for x in range(len(playlist["entries"]))
                        if x not in indices
                    ]

            elif split_message[0].lower() == "rename":
                if arguments is not None and len(
                        arguments[0]
                ) >= 3 and arguments[0] not in self.playlists.saved_playlists:
                    pl_changes["new_name"] = re.sub("\W", "",
                                                    arguments[0].lower())
                    user_savename = pl_changes["new_name"]

            elif split_message[0].lower() == "search":
                if not arguments:
                    msg = await self.safe_send_message(channel, "Please provide a query to search for!")
                    asyncio.sleep(3)
                    await self.safe_delete_message(msg)
                    continue

                query = " ".join(arguments)
                results = self.playlists.search_entries_in_playlist(
                    player.playlist, playlist, query, certainty_threshold=.55)

                if not results:
                    msg = await self.safe_send_message(channel, "**Didn't find anything**")
                    asyncio.sleep(4)
                    await self.safe_delete_message(msg)
                    continue

                lines = []
                for certainty, entry in results[:5]:
                    entry_index = entries.index(entry)
                    lines.append(
                        "`{}.` **{}**".format(entry_index, entry.title))

                msg = "**Found the following entries:**\n" + \
                    "\n".join(lines) + \
                    "\n*Send any message to close this message*"
                msg = await self.safe_send_message(channel, msg)

                resp = await self.wait_for_message(timeout=60, author=author, channel=channel)
                if resp:
                    await self.safe_delete_message(resp)
                await self.safe_delete_message(msg)

                continue

            elif split_message[0].lower() == "extras":

                def extras_check(m):
                    return (m.content.split()[0].lower() in [
                        "abort", "sort", "removeduplicates", "rebuild"
                    ])

                extras_message = await self.safe_send_message(
                    channel,
                    extras_string.format(
                        user_savename.replace("_", " ").title(),
                        self.get_global_user(playlist["author"]).mention,
                        playlist["entry_count"], "s"
                        if int(playlist["entry_count"]) is not 1 else "",
                        format_time(sum([x.duration for x in entries]))))
                resp = await self.wait_for_message(
                    author=author, channel=channel, check=extras_check)

                if not resp.content.lower().startswith(self.config.command_prefix) and not resp.content.lower().startswith('abort'):
                    _cmd = resp.content.split()
                    cmd = _cmd[0].lower()
                    args = _cmd[1:] if len(_cmd) > 1 else None

                    if cmd == "sort":
                        sort_method = args[0].lower() if args is not None and args[0].lower() in [
                            "alphabetical", "length", "random"] else "alphabetical"

                        if sort_method == "alphabetical":
                            playlist["entries"] = sorted(
                                entries, key=lambda entry: entry.title)
                        elif sort_method == "length":
                            playlist["entries"] = sorted(
                                entries, key=lambda entry: entry.duration)
                        elif sort_method == "random":
                            new_ordered = entries
                            shuffle(new_ordered)
                            playlist["entries"] = new_ordered

                        # bodge for changelog
                        pl_changes["order"] = sort_method

                    if cmd == "removeduplicates":
                        urls = []
                        new_list = []
                        for entry in entries:
                            if entry.url not in urls:
                                urls.append(entry.url)
                                new_list.append(entry)

                        playlist["entries"] = new_list

                    if cmd == "rebuild":
                        entry_urls = [entry.url for entry in entries]
                        rebuild_safe_entries, rebuild_removed_entries = await _get_entries_from_urls(entry_urls, "Rebuilding the playlist. This might take a while, please hold on.")

                        pl_changes["remove_entries"].extend(
                            [(ind, playlist["entries"][ind]) for ind in rebuild_removed_entries])  # for the changelog
                        playlist["entries"] = rebuild_safe_entries
                        playlist["entry_count"] = str(
                            len(rebuild_safe_entries))
                        it, ov = divmod(
                            int(playlist["entry_count"]), items_per_page)
                        entries_page = it - 1 if ov == 0 else it
                await self.safe_delete_message(extras_message)
                await self.safe_delete_message(resp)
                await self.safe_delete_message(response_message)
                await self.safe_delete_message(interface_message)
                continue

            elif split_message[0].lower() == "p":
                entries_page = (entries_page - 1) % (iterations + 1)

            elif split_message[0].lower() == "n":
                entries_page = (entries_page + 1) % (iterations + 1)

            await self.safe_delete_message(response_message)
            await self.safe_delete_message(interface_message)

        if abort:
            return Response("Closed **{}** without saving".format(savename))
            print("Closed the playlist builder")

        if save:
            if pl_changes["added_entries"] or pl_changes["remove_entries"] or pl_changes["new_name"] or pl_changes["order"]:
                c_log = "**CHANGES**\n\n"
                if pl_changes["added_entries"]:
                    new_entries_string = "\n".join(["    `{}.` {}".format(ind, nice_cut(
                        entry.title, 40)) for ind, entry in enumerate(pl_changes["added_entries"], 1)])
                    c_log += "**New entries**\n{}\n".format(new_entries_string)
                if pl_changes["remove_entries"]:
                    removed_entries_string = "\n".join(
                        ["    `{}.` {}".format(ind + 1, nice_cut(entry.title, 40)) for ind, entry in pl_changes["remove_entries"]])
                    c_log += "**Removed entries**\n{}\n".format(
                        removed_entries_string)
                if pl_changes["order"]:
                    c_log += "**Changed order**\n    To `{}`".format(
                        pl_changes["order"])
                if pl_changes["new_name"]:
                    c_log += "**Renamed playlist**\n    From `{}` to `{}`".format(
                        savename.title(), pl_changes["new_name"].title())
            else:
                c_log = "No changes were made"

            self.playlists.edit_playlist(
                savename,
                player.playlist,
                all_entries=playlist["entries"],
                new_name=pl_changes["new_name"])
            print("Closed the playlist builder and saved the playlist")

            return Response("Successfully saved **{}**\n\n{}".format(
                user_savename.replace("_", " ").title(), c_log))

    @command_info("2.9.2", 1479945600, {
        "3.3.6": (1497387101, "added the missing \"s\", should be working again"),
        "3.4.4": (1497611753, "Changed command name from \"addplayingtoplaylist\" to \"addtoplaylist\", thanks Paulo"),
        "3.5.5": (1497792167, "Now displaying what entry has been added to the playlist"),
        "3.5.8": (1497826743, "Even more information displaying"),
        "3.6.1": (1497972538, "now accepts a query parameter which adds a song to the playlist like the `play` command does so for the queue"),
        "3.8.9": (1499516220, "Part of the `Giesenesis` rewrite")
    })
    async def cmd_addtoplaylist(self, channel, author, player, playlistname, query=None):
        """
        ///|Usage
        `{command_prefix}addtoplaylist <playlistname> [link | name]`
        ///|Explanation
        Add the current entry to a playlist.
        If you either provide a link or a name, that song is added to the queue.
        """

        if playlistname is None:
            return Response(
                "Please specify the playlist's name!")

        playlistname = playlistname.lower()

        await self.send_typing(channel)

        if query:
            add_entry = await player.playlist.get_entry_from_query(query, channel=channel, author=author)
        else:
            if not player.current_entry:
                return Response(
                    "There's nothing playing right now so I can't add it to your playlist..."
                )

            add_entry = player.current_entry
            if isinstance(add_entry, TimestampEntry):
                current_timestamp = add_entry.current_sub_entry["name"]
                # this looks ugly but eh, it works
                try:
                    add_entry = await player.playlist.get_entry_from_query(current_timestamp, channel=channel, author=author)
                except:
                    pass  # just go ahead and add the whole thing, what do I care :3

        if playlistname not in self.playlists.saved_playlists:
            if len(playlistname) < 3:
                return Response(
                    "Your name is too short. Please choose one with at least three letters."
                )
            self.playlists.set_playlist([add_entry], playlistname, author.id)
            return Response("Created a new playlist `{}` and added **{}**.".format(playlistname.title(),
                                                                                   add_entry.title))

        res = self.playlists.in_playlist(
            player.playlist, playlistname, add_entry)
        if res:
            notification = await self.safe_send_message(
                channel,
                "There's already an entry similar to this one in `{}`\n**{}**\n\nDo you still want to add **{}**? `yes`/`no`".format(
                    playlistname.title(),
                    res.title,
                    add_entry.title
                )
            )
            response = await self.wait_for_message(timeout=60, channel=channel, author=author, check=lambda msg: msg.content.lower().strip() in ["y", "yes", "no", "n"])
            await self.safe_delete_message(notification)
            if response:
                await self.safe_delete_message(response)

            if not (response and response.content.lower().strip().startswith("y")):
                return Response("Didn't add **{}** to `{}`".format(add_entry.title, playlistname.title()))

        self.playlists.edit_playlist(
            playlistname, player.playlist, new_entries=[add_entry])
        return Response("Added **{}** to playlist `{}`.".format(add_entry.title, playlistname.title()))

    @command_info("2.9.2", 1479945600, {
        "3.3.6": (1497387101, "added the missing \"s\", should be working again"),
        "3.4.4": (1497611753, "Changed command name from \"removeplayingfromplaylist\" to \"removefromplaylist\", thanks Paulo"),
        "3.5.8": (1497826917, "Now displaying the names of the song and the playlist"),
        "3.6.5": (1498152365, "Don't require a playlistname argument anymore but take it from the entry itself"),
        "3.8.9": (1499516220, "Part of the `Giesenesis` rewrite")
    })
    async def cmd_removefromplaylist(self, channel, author, player, playlistname=None):
        """
        ///|Usage
        `{command_prefix}removefromplaylist [playlistname]`
        ///|Explanation
        Remove the current entry from its playlist or from the specified playlist.
        """

        if not player.current_entry:
            return Response("There's nothing playing right now so I can hardly remove it from your playlist...")

        if not playlistname:
            if "playlist" in player.current_entry.meta:
                # because why make it easy when you can have it complicated
                playlist_name, index = (
                    player.current_entry.meta["playlist"][x] for x in ["name", "index"])

                self.playlists.edit_playlist(
                    playlist_name, player.playlist, remove_entries=[player.current_entry, ])
                return Response("Removed **{}** from playlist `{}`".format(player.current_entry.title, playlist_name.title()))
            else:
                return Response("Please specify the playlist's name!")

        playlistname = playlistname.lower()

        remove_entry = player.current_entry
        if isinstance(remove_entry, TimestampEntry):
            current_timestamp = remove_entry.current_sub_entry["name"]
            remove_entry = await player.playlist.get_entry_from_query(current_timestamp, channel=channel, author=author)

        if playlistname not in self.playlists.saved_playlists:
            return Response("There's no playlist `{}`.".format(playlistname.title()))

        self.playlists.edit_playlist(
            playlistname, player.playlist, remove_entries=[remove_entry])
        return Response("Removed **{}** from playlist `{}`.".format(remove_entry.title, playlistname))

    @block_user
    async def cmd_reminder(self, channel, author, player, server,
                           leftover_args):
        """
        Usage:
            {command_prefix}reminder create
            {command_prefix}reminder list

        Create a reminder!
        """

        if len(leftover_args) < 1:
            return Response("Please git gud!")

        command = leftover_args[0].lower().strip()

        if (command == "create"):
            import parsedatetime
            cal = parsedatetime.Calendar()

            reminder_name = None
            reminder_due = None
            reminder_repeat = None
            reminder_end = None
            reminder_action = None

            # find out the name
            def check(m):
                return len(m.content) > 3

            msg = await self.safe_send_message(
                channel, "How do you want to call your reminder?")
            response = await self.wait_for_message(
                author=author, channel=channel, check=check)
            reminder_name = response.content
            await self.safe_delete_message(msg)
            await self.safe_delete_message(response)

            # find out the due date
            while True:
                msg = await self.safe_send_message(channel, "When is it due?")
                response = await self.wait_for_message(
                    author=author, channel=channel)

                reminder_due = datetime(
                    *cal.parse(response.content.strip().lower())[0][:6])
                await self.safe_delete_message(msg)
                if reminder_due is not None:
                    await self.safe_delete_message(response)
                    break

                await self.safe_delete_message(response)

            # repeated reminder
            while True:
                msg = await self.safe_send_message(
                    channel,
                    "When should this reminder be repeated? (\"never\" if not at all)"
                )
                response = await self.wait_for_message(
                    author=author, channel=channel)
                await self.safe_delete_message(msg)
                if (response.content.lower().strip() in ("n", "no", "nope",
                                                         "never")):
                    await self.safe_delete_message(response)
                    reminder_repeat = None
                    break

                reminder_repeat = datetime(*cal.parse(
                    response.content.strip().lower())[0][:6]) - datetime.now()
                if reminder_repeat is not None:
                    await self.safe_delete_message(response)
                    break

                await self.safe_delete_message(response)

            # reminder end
            if reminder_repeat is not None:
                while True:
                    msg = await self.safe_send_message(
                        channel,
                        "When should this reminder stop being repeated? (\"never\" if not at all)"
                    )
                    response = await self.wait_for_message(
                        author=author, channel=channel)
                    await self.safe_delete_message(msg)
                    if (response.content.lower().strip() in ("n", "no", "nope",
                                                             "never")):
                        await self.safe_delete_message(response)
                        reminder_end = None
                        break

                    reminder_end = datetime(
                        *cal.parse(response.content.strip().lower())[0][:6])
                    if reminder_end is not None:
                        await self.safe_delete_message(response)
                        break

                    await self.safe_delete_message(response)

            # action
            def check(m):
                try:
                    if 4 > int(m.content) > 0:
                        return True
                    else:
                        return False
                except:
                    return False

            selected_action = 0

            while True:
                msg = await self.safe_send_message(
                    channel,
                    "**Select one:**\n```\n1: Send a message\n2: Play a video\n3: Play an alarm sound```"
                )
                response = await self.wait_for_message(
                    author=author, channel=channel)
                await self.safe_delete_message(msg)
                selected_action = int(response.content)

                if selected_action is not None:
                    await self.safe_delete_message(response)
                    break

                await self.safe_delete_message(response)

            # action 1 (message)
            if selected_action == 1:
                action_message = "Your reminder **{reminder.name}** is due"
                action_channel = None
                action_delete_after = 0
                action_delete_previous = False

                # find message
                msg = await self.safe_send_message(
                    channel, "What should the message say?")
                response = await self.wait_for_message(
                    author=author, channel=channel)
                action_message = response.content
                await self.safe_delete_message(msg)
                await self.safe_delete_message(response)

                # find channel
                while action_channel is None:
                    msg = await self.safe_send_message(
                        channel,
                        "To which channel should the message be sent?\nPossible inputs:\n\n:white_small_square: Channel id or channel name\n:white_small_square: \"me\" for a private message\n:white_small_square: \"this\" to select the current channel\n:white_small_square: You can also @mention people or #mention a channel"
                    )
                    response = await self.wait_for_message(
                        author=author, channel=channel)

                    if len(response.channel_mentions) > 0:
                        action_channel = response.channel_mentions[0]
                    elif len(response.mentions) > 0:
                        action_channel = response.mentions[0]
                    elif response.content.lower().strip() == "me":
                        action_channel = author
                    elif response.content.lower().strip() == "this":
                        action_channel = channel
                    else:
                        return Response("not yet implemented :P")

                    await self.safe_delete_message(msg)
                    await self.safe_delete_message(response)

                # find delete after time
                def check(m):
                    try:
                        if m.content.lower().strip() in [
                                "never", "no"
                        ] or int(m.content.strip()) >= 0:
                            return True
                        else:
                            return False
                    except:
                        return False

                msg = await self.safe_send_message(
                    channel,
                    "After how many seconds should the message be deleted? (\"never\" for not at all)"
                )
                response = await self.wait_for_message(
                    author=author, channel=channel, check=check)
                if response.content.lower().strip() in ["never", "no"]:
                    action_delete_after = 0
                else:
                    action_delete_after = int(response.content.strip())

                await self.safe_delete_message(msg)
                await self.safe_delete_message(response)

                # find if delete old message
                if reminder_repeat is not None:
                    msg = await self.safe_send_message(
                        channel,
                        "Before sending a new message, should the old one be deleted?"
                    )
                    response = await self.wait_for_message(
                        author=author, channel=channel)
                    if response.content.lower().strip() in ["y", "yes"]:
                        action_delete_previous = True

                    await self.safe_delete_message(msg)
                    await self.safe_delete_message(response)

                reminder_action = Action(
                    channel=action_channel,
                    msg_content=action_message,
                    delete_msg_after=action_delete_after,
                    delete_old_message=action_delete_previous)

            # action 2 (play url)
            elif selected_action == 2:
                action_source_url = ""
                action_voice_channel = None

                # find video url
                msg = await self.safe_send_message(
                    channel, "What's the url of the video you want to play?")
                response = await self.wait_for_message(
                    author=author, channel=channel)
                action_source_url = response.content
                await self.safe_delete_message(msg)
                await self.safe_delete_message(response)

                # find playback channel
                msg = await self.safe_send_message(
                    channel,
                    "To which channel should the video be played?\nPossible inputs:\n\n:white_small_square: Channel id or channel name\n:white_small_square: \"this\" to select your current channel"
                )
                response = await self.wait_for_message(
                    author=author, channel=channel)

                if response.content.lower().strip() == "this":
                    return Response("not yet implemented :P")
                else:
                    return Response("not yet implemented :P")

            # action 3 (play predefined)
            elif selected_action == 3:
                pass

            # finalizing
            self.calendar.create_reminder(
                reminder_name,
                reminder_due,
                reminder_action,
                repeat_every=reminder_repeat,
                repeat_end=reminder_end)
            return Response(
                "Created a reminder called **{}**\ndue: {}\nrepeat: {}\nrepeat end: {}\naction: {}".
                format(reminder_name, reminder_due, reminder_repeat,
                       reminder_end, reminder_action))

        elif (command == "list"):
            if len(self.calendar.reminders) < 1:
                return Response("There are no reminders")

            text = ""
            for reminder in self.calendar.reminders:
                text += "*{.name}*\n".format(reminder)

            return Response(text)

    @command_info("2.0.3", 1485516420, {
        "3.7.5": (1481827320, "The command finally works like it should"),
        "3.9.9": (1499977057, "moving Giesela too")
    })
    async def cmd_moveus(self, channel, server, author, message, leftover_args):
        """
        ///|Usage
        `{command_prefix}moveus <channel name>`
        ///|Explanation
        Move everyone in your current channel to another one!
        """

        if len(leftover_args) < 1:
            return Response("You need to provide a target channel")

        search_channel = " ".join(leftover_args)
        if search_channel.lower().strip() == "home":
            search_channel = "Giesela's reign"

        if author.voice.voice_channel is None:
            return Response(
                "You're incredibly incompetent to do such a thing!")

        author_channel = author.voice.voice_channel
        voice_members = author_channel.voice_members
        move_myself = False
        if server.me in voice_members:
            voice_members.remove(server.me)
            move_myself = True

        target_channel = self.get_channel(search_channel)
        if target_channel is None:
            for chnl in server.channels:
                if chnl.name == search_channel and chnl.type == ChannelType.voice:
                    target_channel = chnl
                    break

        if target_channel is None:
            return Response(
                "Can't resolve the target channel!")

        s = 0
        for voice_member in voice_members:
            await self.move_member(voice_member, target_channel)
            s += 1

        if move_myself:
            print("moving myself")
            await self.move_voice_client(target_channel)

    @block_user
    @command_info("2.0.2", 1484676180, {
        "3.8.3": (1499184914, "Can now use multiline statements without having to use tricks like /n/"),
        "3.8.5": (1499279145, "Better code display"),
        "3.9.6": (1499889309, "Escaping the result and adding the shortcut entry for player.current_entry")
    })
    async def cmd_execute(self, channel, author, server, raw_content, player=None):
        statement = raw_content.strip()
        beautiful_statement = "```python\n{}\n```".format(statement)

        statement = "async def func():\n{}".format(indent(statement, "\t"))
        await self.safe_send_message(channel, "**RUNNING CODE**\n{}".format(beautiful_statement))

        env = {}
        env.update(globals())
        env.update(locals())
        env.update(entry=player.current_entry)

        try:
            exec(statement, env)
        except SyntaxError as e:
            return Response(
                "**While compiling the statement the following error occured**\n{}\n{}".
                format(traceback.format_exc(), str(e)))

        func = env["func"]

        try:
            ret = await func()
        except Exception as e:
            return Response(
                "**While executing the statement the following error occured**\n{}\n{}".
                format(traceback.format_exc(), str(e)))

        return Response("**RESULT**\n```python\n{}\n```".format(escape_dis(str(ret))))

    @command_info("2.0.3", 1487538840, {
        "3.3.7": (1497471402, "changed command from \"skipto\" to \"seek\"")
    })
    async def cmd_seek(self, player, timestamp):
        """
        ///|Usage
        `{command_prefix}seek <timestamp>`
        ///|Explanation
        Go to the given timestamp formatted (minutes:seconds)
        """

        secs = parse_timestamp(timestamp)
        if secs is None:
            return Response(
                "Please provide a valid timestamp")

        if player.current_entry is None:
            return Response("Nothing playing!")

        if not player.goto_seconds(secs):
            return Response(
                "Timestamp exceeds song duration!")

    @command_info("2.2.1", 1493975700, {
        "3.8.9": (1499516220, "Part of the `Giesenesis` rewrite")
    })
    async def cmd_fwd(self, player, timestamp):
        """
        ///|Usage
        `{command_prefix}fwd <timestamp>`
        ///|Explanation
        Forward <timestamp> into the current entry
        """

        secs = parse_timestamp(timestamp)
        if secs is None:
            return Response(
                "Please provide a valid timestamp")

        if player.current_entry is None:
            return Response("Nothing playing!")

        if not player.goto_seconds(player.progress + secs):
            return Response(
                "Timestamp exceeds song duration!")

    @command_info("2.2.1", 1493975700, {
        "3.4.3": (1497609912, "Can now rewind past the current song"),
        "3.8.9": (1499516220, "Part of the `Giesenesis` rewrite")
    })
    async def cmd_rwd(self, player, timestamp=None):
        """
        ///|Usage
        `{command_prefix}rwd [timestamp]`
        ///|Explanation
        Rewind <timestamp> into the current entry or if the current entry is a timestamp-entry, rewind to the previous song
        """

        if player.current_entry is None:
            return Response("Nothing playing!")

        if timestamp is None:
            if isinstance(player.current_entry, TimestampEntry):
                current_song = player.current_entry.current_sub_entry
                ind = current_song["index"]
                progress = current_song["progress"]

                if ind == 0:
                    secs = 0
                else:
                    if progress < 15:
                        secs = player.current_entry.sub_queue[ind - 1]["start"]
                    else:
                        secs = current_song["start"]

            else:
                return Response("Please provide a valid timestamp")
        else:
            secs = player.progress - parse_timestamp(timestamp)

        if not secs:
            if not player.playlist.history:
                return Response(
                    "Please provide a valid timestamp (no history to rewind into)")
            else:
                # just replay the last entry
                last_entry = player.playlist.history[0]
                player.play_entry(last_entry)
                return

        if secs < 0:
            if not player.playlist.history:
                secs = 0
            else:
                last_entry = player.playlist.history[0]
                # since secs is negative I can just add it
                if not last_entry.set_start(last_entry.end_seconds + secs):
                    # mostly because I'm lazy
                    return Response(
                        "I won't go further back than one song, that's just mad"
                    )
                player.play_entry(last_entry)
                return

        if not player.goto_seconds(secs):
            return Response(
                "Timestamp exceeds song duration!")

    async def cmd_disconnect(self, server):
        """
        Usage:
            {command_prefix}disconnect

        Make the bot leave his current voice channel.
        """
        await self.disconnect_voice_client(server)
        return Response(":hear_no_evil:")

    async def cmd_restart(self, channel):
        await self.safe_send_message(channel, ":wave:")
        await self.disconnect_all_voice_clients()
        raise exceptions.RestartSignal

    @owner_only
    async def cmd_countmsgs(self, server, author, channel_id, number):
        alphabet = list("abcdefghijklmnopqrstuvwxyz")

        def index_to_alphabet(ind):
            if ind < len(alphabet):
                return alphabet[ind].upper()

            remainder = ind % len(alphabet)
            return index_to_alphabet(ind -
                                     remainder) + alphabet[remainder].upper()

        msgs_by_member = {}
        msgs_by_date = OrderedDict()
        answers_by_date = OrderedDict()
        channel = server.get_channel(channel_id)
        last_msg = None
        last_answer = None
        spam = 0

        async for msg in self.logs_from(channel, limit=int(number)):
            increment = 1
            if last_msg is not None and msg.author.id == last_msg.author.id and abs(
                    (last_msg.timestamp - msg.timestamp).total_seconds()) < 10:
                spam += 1
                last_msg = msg
                increment = 0

            if last_answer is None or last_answer.author != msg.author:
                dt = answers_by_date.get(
                    "{0.day:0>2}/{0.month:0>2}/{0.year:0>4}".format(
                        msg.timestamp), {})
                dt[msg.author.id] = dt.get(msg.author.id, 0) + increment
                answers_by_date["{0.day:0>2}/{0.month:0>2}/{0.year:0>4}".
                                format(msg.timestamp)] = dt
                last_answer = msg

            existing_msgs = msgs_by_member.get(msg.author.id, [0, 0])
            existing_msgs[0] += increment
            existing_msgs[1] += len(re.sub(r"\W", r"", msg.content))
            msgs_by_member[msg.author.id] = existing_msgs
            dt = msgs_by_date.get(
                "{0.day:0>2}/{0.month:0>2}/{0.year:0>4}".format(msg.timestamp),
                {})
            dt[msg.author.id] = dt.get(msg.author.id, 0) + increment
            msgs_by_date["{0.day:0>2}/{0.month:0>2}/{0.year:0>4}".format(
                msg.timestamp)] = dt
            last_msg = msg

        wb = Workbook()
        ws = wb.active
        ws.title = "Messages"
        ws2 = wb.create_sheet("Answers")
        ws["A2"] = "TOTAL"
        sorted_user_index = {}
        i = 1
        for member in sorted(msgs_by_member):
            data = msgs_by_member[member]
            ws["{}{}".format("A", i)] = server.get_member(
                member
            ).name if server.get_member(member) is not None else "Unknown"
            ws["{}{}".format("B", i)] = data[0]
            ws["{}{}".format("C", i)] = data[1]
            sorted_user_index[member] = index_to_alphabet(i)
            i += 1

        i += 1
        for date in reversed(msgs_by_date.keys()):
            ws["A" + str(i)] = date
            for mem in msgs_by_date[date]:
                ws["{}{}".format(sorted_user_index.get(mem),
                                 i)] = msgs_by_date[date][mem]
            i += 1

        i = 1
        for date in reversed(answers_by_date.keys()):
            ws2["A" + str(i)] = date
            for mem in answers_by_date[date]:
                ws2["{}{}".format(sorted_user_index.get(mem),
                                  i)] = answers_by_date[date][mem]
            i += 1

        wb.save("cache/last_data.xlsx")

        await self.send_file(
            author,
            open("cache/last_data.xlsx", "rb"),
            filename='%s-msgs.xlsx' % (server.name.replace(' ', '_')))

    async def cmd_archivechat(self,
                              server,
                              author,
                              message,
                              placeholder=None,
                              number=1000000):
        if message.channel_mentions is None or len(
                message.channel_mentions) < 1:
            return Response("Stupid duck")

        channel = message.channel_mentions[0]
        msgs = []
        async for msg in self.logs_from(channel, limit=int(number)):
            msg_data = {
                "name": msg.author.name,
                "timestamp": str(round(msg.timestamp.timestamp())),
                "content": msg.content,
                "attachments": msg.attachments
            }
            msgs.append(msg_data)

        json.dump(msgs[::-1], open("cache/last_message_archive.json", "w+"))
        await self.send_file(
            author,
            open("cache/last_message_archive.json", "rb"),
            filename='%s-msg-archive.json' % (server.name.replace(' ', '_')))

    @owner_only
    async def cmd_surveyserver(self, server):
        if self.online_loggers.get(server.id, None) is not None:
            return Response("I'm already looking at this server")
        else:
            online_logger = OnlineLogger(self)
            self.online_loggers[server.id] = online_logger
            Settings["online_loggers"] = list(self.online_loggers.keys())
            return Response("okay, okay!")

    def load_online_loggers(self):
        for server_id in Settings.get_setting("online_loggers", default=[]):
            online_logger = OnlineLogger(self)
            self.online_loggers[server_id] = online_logger
            for listener in Settings.get_setting(
                    "online_logger_listeners_" + server_id, default=[]):
                online_logger.add_listener(listener)

    @owner_only
    async def cmd_evalsurvey(self, server, author):
        online_logger = self.online_loggers.get(server.id, None)
        if online_logger is None:
            return Response("I'm not even spying here")
        online_logger.create_output()
        await self.send_file(
            author,
            open("cache/last_survey_data.xlsx", "rb"),
            filename='%s-survey.xlsx' % (server.name.replace(' ', '_')))
        return Response("There you go, fam")

    @owner_only
    async def cmd_resetsurvey(self, server):
        online_logger = self.online_loggers.get(server.id, None)
        if online_logger is None:
            return Response("I'm not even spying here")
        online_logger.reset()
        return Response("Well then")

    async def cmd_notifyme(self, server, author):
        """
        Usage:
            {command_prefix}notifyme

        Get notified when someone starts playing
        """
        online_logger = self.online_loggers.get(server.id, None)
        if online_logger is None:
            return Response("I'm not even spying here")
        if online_logger.add_listener(author.id):
            Settings["online_logger_listeners_" + server.id] = [
                *Settings.get_setting(
                    "online_logger_listeners_" + server.id, default=[]),
                author.id
            ]
            return Response("Got'cha!")
        else:
            try:
                Settings["online_logger_listeners_" + server.id] = [
                    x
                    for x in Settings.get_setting(
                        "online_logger_listeners_" + server.id, default=[])
                    if x != author.id
                ]
            except ValueError:
                pass

            return Response("Nevermore you shall be annoyed!")

    @command_info("2.2.1", 1493757540, {
        "3.7.8": (1499019245, "Fixed quoting by content.")
    })
    async def cmd_quote(self, author, channel, message, leftover_args):
        """
        ///|Usage
        `{command_prefix}quote [#channel] <message id> [message id...]`
        `{command_prefix}quote [#channel] [@mention] \"<message content>\"`
        ///|Explanation
        Quote a message
        """

        quote_to_channel = channel
        target_author = None

        if message.channel_mentions:
            channel = message.channel_mentions[0]
            leftover_args = leftover_args[1:]

        if message.mentions:
            target_author = message.mentions[0]
            leftover_args = leftover_args[1:]

        if len(leftover_args) < 1:
            return Response("Please specify the message you want to quote")

        message_content = " ".join(leftover_args)
        if (message_content[0] == "\"" and message_content[-1] == "\"") or re.search(r"\D", message_content) is not None:
            message_content = message_content.replace("\"", "")
            async for msg in self.logs_from(channel, limit=3000):
                if msg.id != message.id and message_content.lower().strip() in msg.content.lower().strip():
                    if target_author is None or target_author.id == msg.author.id:
                        leftover_args = [msg.id, ]
                        break
            else:
                if target_author is not None:
                    return Response("Didn't find a message with that content from {}".format(target_author.mention))
                else:
                    return Response("Didn't find a message with that content")

        await self.safe_delete_message(message)
        for message_id in leftover_args:
            try:
                quote_message = await self.get_message(channel, message_id)
            except:
                return Response("Didn't find a message with the id `{}`".
                                format(message_id))

            author_data = {
                "name": quote_message.author.display_name,
                "icon_url": quote_message.author.avatar_url
            }
            embed_data = {
                "description": quote_message.content,
                "timestamp": quote_message.timestamp,
                "colour": quote_message.author.colour
            }
            em = Embed(**embed_data)
            em.set_author(**author_data)
            await self.send_message(quote_to_channel, embed=em)
        return

    @command_info("3.2.5", 1496428380, {
        "3.3.9": (1497521393, "Added edit sub-command"),
        "3.4.1": (1497550771, "Added the filter \"mine\" to the listing function"),
        "3.4.6": (1497617827, "when listing bookmarks, they musn't be \"inline\"."),
        "3.5.8": (1497827057, "Editing bookmarks now works as expected")
    })
    async def cmd_bookmark(self, author, player, leftover_args):
        """
        ///|Creation
        {command_prefix}bookmark [name] [timestamp]
        ///|Explanation
        Create a new bookmark for the current entry. If no name is provided the entry's title will be used and if there's no timestamp provided the current timestamp will be used.
        ///|Using
        {command_prefix}bookmark <id | name>
        ///|Editing
        {command_prefix}bookmark edit <id> [new name] [new timestamp]
        ///|Listing
        {command_prefix}bookmark list [mine]
        ///|Removal
        {command_prefix}bookmark remove <id | name>
        """
        if len(leftover_args) > 0:
            arg = leftover_args[0].lower()
            if arg in ["list", "showall"]:
                em = Embed(title="Bookmarks")
                bookmarks = bookmark.all_bookmarks

                if "mine" in leftover_args:
                    bookmarks = filter(
                        lambda x: bookmark.get_bookmark(
                            x)["author_id"] == author.id,
                        bookmarks)

                for bm in bookmarks:
                    bm_name = bm["name"]
                    bm_author = self.get_global_user(
                        bm["author_id"]).display_name
                    bm_timestamp = to_timestamp(bm["timestamp"])
                    bm_id = bm["id"]
                    t = "**{}**".format(bm_name)
                    v = "`{}` starting at `{}` *by* **{}**".format(
                        bm_id, bm_timestamp, bm_author)
                    em.add_field(name=t, value=v, inline=False)
                return Response(embed=em)
            elif arg in ["remove", "delete"]:
                if len(leftover_args) < 2:
                    return Response("Please provide an id or a name")
                bm = bookmark.get_bookmark(" ".join(leftover_args[1:]))
                if not bm:
                    return Response("Didn't find a bookmark with that query")
                if bookmark.remove_bookmark(bm["id"]):
                    return Response("Removed bookmark `{}`".format(bm["name"]))
                else:
                    return Response("Something went wrong")
            elif arg in ["edit", "change"]:
                if len(leftover_args) < 2:
                    return Response("Please provide an id")

                bm_id = leftover_args[1]
                if bm_id not in bookmark:
                    return Response(
                        "No bookmark with id `{}` found".format(bm_id))

                if len(leftover_args) < 3:
                    return Response(
                        "Please also specify what you want to change")

                new_timestamp = parse_timestamp(leftover_args[-1])
                if new_timestamp is not None:  # 0 evaluates to false so I need to check this oldschool-like
                    new_name = " ".join(
                        leftover_args[2:-1]) if len(leftover_args) > 3 else None
                else:
                    new_name = " ".join(leftover_args[2:])

                if bookmark.edit_bookmark(bm_id, new_name, new_timestamp):
                    return Response(
                        "Successfully edited bookmark `{}`".format(bm_id))
                else:
                    return Response("Something went wrong while editing `{}`".
                                    format(bm_id))
            else:
                bm = bookmark.get_bookmark(" ".join(leftover_args))
                if bm:
                    player.playlist._add_entry(
                        URLPlaylistEntry.from_dict(player.playlist, bm[
                            "entry"]))
                    return Response("Loaded bookmark `{0}` by **{1}**".
                                    format(bm["name"],
                                           self.get_global_user(
                                               bm["author_id"]).display_name))
                else:
                    bm_timestamp = player.progress
                    bm_name = None
                    if len(leftover_args) > 1:
                        timestamp = parse_timestamp(leftover_args[-1])
                        if timestamp:
                            bm_timestamp = timestamp
                        bm_name = " ".join(
                            leftover_args[:-1]) if timestamp else " ".join(
                                leftover_args)
                    else:
                        timestamp = parse_timestamp(leftover_args[-1])
                        if timestamp:
                            bm_timestamp = timestamp
                        else:
                            bm_name = " ".join(leftover_args)

                    id = bookmark.add_bookmark(
                        player.current_entry, bm_timestamp, author.id, bm_name)
                    return Response(
                        "Created a new bookmark with the id `{0}` (\"{2}\", `{3}`)\nUse `{1}bookmark {0}` to load it ".
                        format(id, self.config.command_prefix, bm_name,
                               to_timestamp(bm_timestamp)))

        else:
            if player.current_entry:
                id = bookmark.add_bookmark(player.current_entry,
                                           player.progress, author.id)
                return Response(
                    "Created a new bookmark with the id `{0}`\nUse `{1}bookmark {0}` to load it ".
                    format(id, self.config.command_prefix))
            else:
                return await self.cmd_bookmark(author, player, [
                    "list",
                ])

    @owner_only
    @command_info("3.1.6", 1498672140, {
        "3.6.4": (1498146841, "Can now specify the required arguments in order to block a command"),
        "3.9.8": (1499976133, "Saving the blocked commands")
    })
    async def cmd_blockcommand(self, command, leftover_args):
        """
        ///|Usage
        `{command_prefix}blockcommand <command> [args] <"reason">`
        ///|Explanation
        Block a command
        """
        if command.lower() in self.blocked_commands:
            self.blocked_commands.pop(command.lower())
            Settings["blocked_commands"] = self.blocked_commands
            return Response("Block lifted")
        else:
            if len(leftover_args) < 1:
                return Response("Reason plz")

            args = []

            for i, el in enumerate(leftover_args):
                if not el.startswith("\""):
                    args.append(el)
                else:
                    reason = " ".join(leftover_args[i:]).strip("\"")
                    break

            if not reason:
                return Response("Put your reason in quotes, idiot!")

            self.blocked_commands[command.lower()] = (args, reason)
            Settings["blocked_commands"] = self.blocked_commands
            return Response("Blocked command `{} {}`".format(command, " ".join(args)))

    @command_info("3.4.0", 1497533758, {
        "3.4.8": (1497650090, "When showing changelogs, two logs can't be on the same line anymore")
    })
    async def cmd_commandinfo(self, command):
        """
        ///|Usage
        `{command_prefix}commandinfo <command>`
        ///|Explanation
        More information on a command
        """

        c_info = getattr(self, "cmd_" + command, None)
        if not c_info:
            return Response(
                "Couldn't find a command called \"{}\"".format(command))

        try:
            em = Embed(title=command.upper(), colour=hex_to_dec("#ffd700"))
            em.add_field(
                name="Version `{}`".format(c_info.version),
                value="`{}`\nCommand has been added".format(c_info.timestamp),
                inline=False)

            for cl in c_info.changelog:
                v, t, l = cl
                em.add_field(
                    name="Version `{}`".format(v),
                    value="`{}`\n{}".format(t, l),
                    inline=False)

            return Response(embed=em)
        except:
            return Response(
                "Couldn't find any information on the `{}` command".format(
                    command))

    @command_info("3.5.6", 1497819288, {
        "3.6.2": (1497978696, "references are now clickable"),
        "3.7.6": (1498947694, "fixed a bug which would stop Giesela from executing the command because of underscores in the version name")
    })
    async def cmd_version(self, channel):
        """
        ///|Usage
        `{command_prefix}version`
        ///|Explanation
        Some more informat about the current version and what's to come.
        """

        await self.send_typing(channel)
        v_code, v_name = BOTVERSION.split("_", 1)
        dev_code, dev_name = get_dev_version()
        changelog = get_dev_changelog()

        desc = "Current Version is `{}`\nDevelopment is at `{}`\n\n**What's to come:**\n\n".format(
            BOTVERSION, dev_code + "_" + dev_name)
        desc += "\n".join("● " + l for l in changelog)
        em = Embed(title="Version " + v_name, description=desc,
                   url="https://siku2.github.io/Giesela", colour=hex_to_dec("#67BE2E"))

        return Response(embed=em)

    @command_info("3.5.7", 1497823283, {
        "3.8.9": (1499645741, "`Giesenesis` rewrite was here")
    })
    async def cmd_interact(self, channel, message):
        """
        ///|Usage
        `{command_prefix}interact <query>`
        ///|Explanation
        Use everyday language to control Giesela
        ///|Disclaimer
        **Help out with the development of a "smarter" Giesela by testing out this new feature!**
        """

        await self.send_typing(channel)

        matcher = "^\{}?interact".format(self.config.command_prefix)
        query = re.sub(matcher, "", message.content,
                       flags=re.MULTILINE).strip()
        if not query:
            return Response("Please provide a query for me to work with")

        print("[INTERACT] \"{}\"".format(query))

        params = {
            "v": date.today().strftime("%d/%m/%y"),
            "q": query
        }
        headers = {"Authorization": "Bearer HVSTOLU3UQLR7YOYXCONQCCIQNHXZYDM"}
        resp = requests.get("https://api.wit.ai/message",
                            params=params, headers=headers)
        data = resp.json()
        entities = data["entities"]

        msg = ""

        for entity, data in entities.items():
            d = data[0]
            msg += "**{}** [{}] ({}% sure)\n".format(entity,
                                                     d["value"], round(d["confidence"] * 100, 1))

        return Response(msg)

    @command_info("3.7.3", 1498306682, {
        "3.7.4": (1498312423, "Fixed severe bug and added musixmatch as a source"),
        "3.9.2": (1499709472, "Fixed typo")
    })
    async def cmd_lyrics(self, player, channel):
        """
        ///|Usage
        `{command_prefix}lyrics`
        ///|Explanation
        Try to find lyrics for the current entry and display 'em
        """

        await self.send_typing(channel)

        if not player.current_entry:
            return Response("There's no way for me to find lyrics for something that doesn't even exist!")

        title = player.current_entry.title
        lyrics = search_for_lyrics(title)

        if not lyrics:
            return Response("Couldn't find any lyrics for **{}**".format(title))
        else:
            return Response("**{}**\n\n{}".format(title, lyrics))

    @command_info("3.8.1", 1499116644)
    async def cmd_register(self, server, author, token):
        """
        ///|Usage
        `{command_prefix}register <token>`
        ///|Explanation
        Use this command in order to use the [Giesela-Website](http://giesela.org).
        """

        if GieselaServer.register_information(server.id, author.id, token.lower()):
            return Response("You've successfully registered yourself. Go back to your browser and check it out")
        else:
            return Response("Something went wrong while registering. It could be that your code `{}` is wrong. Please make sure that you've entered it correctly.".format(token.upper()))

    @command_info("4.0.2", 1500360351)
    async def cmd_explode(self, player, channel, author, leftover_args):
        """
        ///|Usage
        `{command_prefix}explode [playlist link]`
        ///|Explanation
        Split a timestamp-entry into its sub-entries.
        """

        await self.send_typing(channel)

        if leftover_args:
            query = " ".join(leftover_args)
            entry = await player.playlist.get_entry_from_query(query, channel=channel, author=author)
        elif player.current_entry:
            entry = player.current_entry
        else:
            return Response("Can't explode what's not there")

            entry = player.current_entry

        if not isinstance(entry, TimestampEntry):
            return Response("Can only explode timestamp-entries")

        sub_queue = entry.sub_queue

        progress_message = await self.safe_send_message(channel, "Exploding {} entr{}".format(
            len(sub_queue),
            "y" if len(sub_queue) == 1 else "ies"
        ))

        for ind, sub_entry in enumerate(sub_queue, 1):
            add_entry = await player.playlist.get_entry_from_query(sub_entry["name"], author=author, channel=channel)
            player.playlist._add_entry(add_entry)

            prg = ind / len(sub_queue)

            progress_message = await self.safe_edit_message(
                progress_message,
                "Explosion in progress\n{} `{}%`".format(
                    create_bar(prg, length=20),
                    round(100 * prg)
                ),
                keep_at_bottom=True
            )

        await self.safe_delete_message(progress_message)

        return Response("Exploded **{}** into {} entr{}".format(
            entry.whole_title,
            len(sub_queue),
            "y" if len(sub_queue) == 1 else "ies"
        ))

    @owner_only
    async def cmd_shutdown(self, channel):
        await self.safe_send_message(channel, ":wave:")
        await self.disconnect_all_voice_clients()
        raise exceptions.TerminateSignal

    async def on_message(self, message):
        await self.wait_until_ready()

        message_content = message.content.strip()

        nine_gag_match = re.match(
            r".+?[http|https]:\/\/(?:m\.)?9gag.com\/gag\/(\w+)(?:\?.+)",
            message_content)
        if nine_gag_match:
            post_id = nine_gag_match.group(1)
            await self.cmd_9gag(message.channel, message.author, post_id)
            await self.safe_delete_message(message)
            return

        if message.author.id in self.users_in_menu:
            print("{} is currently in a menu. Ignoring \"{}\"".format(
                message.author, message_content))
            return

        if not message_content.startswith(self.config.command_prefix):
            if not self.config.owned_channels or message.channel.id not in self.config.owned_channels:
                return

        # don't react to own messages or messages from bots
        if message.author == self.user or message.author.bot:
            # print("Ignoring command from myself (%s)" %
            #          message.content)
            return

        raw_command, *args = message_content.split()
        command = raw_command[len(self.config.command_prefix):].lower().strip(
        ) if message_content.startswith(
            self.config.command_prefix) else raw_command.lower().strip()

        handler = getattr(self, 'cmd_%s' % command, None)
        if not handler:
            return

        if command in self.blocked_commands:
            required_args, reason = self.blocked_commands[command]
            if all(arg in args for arg in required_args):
                await self.send_message(message.channel, reason)
                return

        if message.channel.is_private:
            if not (message.author.id == self.config.owner_id and command ==
                    'joinserver') and not command in self.config.private_chat_commands:
                await self.send_message(
                    message.channel,
                    'You cannot use this command in private messages.')
                return

        if message.author.id in self.blacklist and message.author.id != self.config.owner_id:
            print("[User blacklisted] {0.id}/{0.name} ({1})".format(
                message.author, message_content))
            return

        else:
            print("[Command] {0.id}/{0.name} ({1})".format(
                message.author, message_content))

        argspec = inspect.signature(handler)
        params = argspec.parameters.copy()
        raw_content = message_content[len(raw_command):]

        # noinspection PyBroadException
        try:
            handler_kwargs = {}
            if params.pop('message', None):
                handler_kwargs['message'] = message

            if params.pop("raw_content", None):
                handler_kwargs["raw_content"] = raw_content

            if params.pop('channel', None):
                handler_kwargs['channel'] = message.channel

            if params.pop('author', None):
                handler_kwargs['author'] = message.author

            if params.pop('server', None):
                handler_kwargs['server'] = message.server

            if params.pop('player', None):
                handler_kwargs['player'] = await self.get_player(
                    message.channel)

            if params.pop('user_mentions', None):
                handler_kwargs['user_mentions'] = list(
                    map(message.server.get_member, message.raw_mentions))

            if params.pop('channel_mentions', None):
                handler_kwargs['channel_mentions'] = list(
                    map(message.server.get_channel,
                        message.raw_channel_mentions))

            if params.pop('voice_channel', None):
                handler_kwargs[
                    'voice_channel'] = message.server.me.voice_channel

            if params.pop('leftover_args', None):
                handler_kwargs['leftover_args'] = args

            args_expected = []
            for key, param in list(params.items()):
                doc_key = '[%s=%s]' % (
                    key, param.default
                ) if param.default is not inspect.Parameter.empty else key
                args_expected.append(doc_key)

                if not args and param.default is not inspect.Parameter.empty:
                    params.pop(key)
                    continue

                if args:
                    arg_value = args.pop(0)
                    handler_kwargs[key] = arg_value
                    params.pop(key)

            if params:
                return await self.cmd_help(message.channel, [
                    command,
                ])

            response = await handler(**handler_kwargs)
            if response and isinstance(response, Response):
                content = response.content
                if content and response.reply:
                    content = '%s, %s' % (message.author.mention, content)

                sentmsg = await self.safe_send_message(
                    message.channel,
                    content,
                    expire_in=response.delete_after
                    if self.config.delete_messages else 0,
                    also_delete=message
                    if self.config.delete_invoking else None,
                    embed=response.embed)

        except (exceptions.CommandError, exceptions.HelpfulError,
                exceptions.ExtractionError) as e:
            print("{0.__class__}: {0.message}".format(e))

            expirein = e.expire_in if self.config.delete_messages else None
            alsodelete = message if self.config.delete_invoking else None

            await self.safe_send_message(
                message.channel,
                '```\n%s\n```' % e.message,
                expire_in=expirein,
                also_delete=alsodelete)

        except exceptions.Signal:
            raise

        except Exception:
            traceback.print_exc()
            if self.config.debug_mode:
                await self.safe_send_message(
                    message.channel, '```\n%s\n```' % traceback.format_exc())

    async def on_reaction_add(self, reaction, user):
        if reaction.message.author == self.user:
            await self.safe_send_message(user, "I hate you too!")
            return

    async def autopause(self, before, after):
        if not all([before, after]):
            return

        if before.voice_channel == after.voice_channel:
            return

        if before.server.id not in self.players:
            return

        # This should always work, right?
        my_voice_channel = after.server.me.voice_channel

        if not my_voice_channel:
            return

        if before.voice_channel == my_voice_channel:
            joining = False
        elif after.voice_channel == my_voice_channel:
            joining = True
        else:
            return  # Not my channel

        moving = before == before.server.me

        player = await self.get_player(my_voice_channel)

        if after == after.server.me and after.voice_channel:
            player.voice_client.channel = after.voice_channel

        if not self.config.auto_pause:
            return

        vm_count = sum(
            1 for m in my_voice_channel.voice_members if m != after.server.me)
        if vm_count == 0:
            if player.is_playing:
                print("[AUTOPAUSE] Pausing")
                player.pause()
        elif vm_count == 1 and joining:
            if player.is_paused:
                print("[AUTOPAUSE] Unpausing")
                player.resume()

    async def on_server_update(self,
                               before: discord.Server,
                               after: discord.Server):
        if before.region != after.region:
            print("[Servers] \"%s\" changed regions: %s -> %s" %
                  (after.name, before.region, after.region))

            await self.reconnect_voice_client(after)

    async def on_member_update(self, before, after):
        await self.on_any_update(before, after)
        if before.server.id in self.online_loggers:
            self.online_loggers[before.server.id].update_stats(
                after.id, after.status == discord.Status.online, after.game)

    async def on_voice_state_update(self, before, after):
        await self.on_any_update(before, after)
        await self.autopause(before, after)

    async def on_any_update(self, before, after):
        if before.server.id in self.online_loggers:
            timestamp = "{0.hour:0>2}:{0.minute:0>2}".format(datetime.now())
            notification = None
            mem_name = "\"{}\"".format(after.display_name) if len(
                after.display_name.split()) > 1 else after.display_name
            if before.status != after.status:
                notification = "`{}` {} {}".format(timestamp, mem_name, {
                    discord.Status.online:
                    "came **online**",
                    discord.Status.offline:
                    "went **offline**",
                    discord.Status.idle:
                    "went **away**",
                    discord.Status.dnd:
                    "doesn't want to be disturbed"
                }[after.status])
            if before.game != after.game:
                text = ""
                if after.game is None:
                    text = "stopped playing **{}**".format(before.game.name)
                else:
                    text = "started playing **{}**".format(after.game.name)
                if notification is None:
                    notification = "`{}` {} {}".format(timestamp, mem_name,
                                                       text)
                else:
                    notification += "\nand {}".format(text)

            if before.voice.voice_channel != after.voice.voice_channel:
                text = ""
                if after.voice.voice_channel is None:
                    text = "quit **{}** (voice channel)".format(
                        before.voice.voice_channel.name)
                else:
                    text = "joined **{}** (voice channel)".format(
                        after.voice.voice_channel.name)
                if notification is None:
                    notification = "`{}` {} {}".format(timestamp, mem_name,
                                                       text)
                else:
                    notification += "\nand {}".format(text)

            if notification is not None:
                for listener in self.online_loggers[
                        before.server.id].listeners:
                    if before.id == listener:
                        continue

                    mem = self.get_global_user(listener)
                    await self.safe_send_message(mem, notification)


if __name__ == '__main__':
    bot = MusicBot()
    bot.run()
