from datetime import datetime
from string import ascii_lowercase as alphabet

from musicbot.utils import format_time
from openpyxl import Workbook


class OnlineLogger:

    def __init__(self, musicbot):
        self.member_data = {}
        self.ongoing_online_phases = {}
        self.ongoing_playing_phases = {}
        self.listeners = []
        self.musicbot = musicbot

    def add_listener(self, id):
        if id in self.listeners:
            self.remove_listener(id)
            return False
        self.listeners.append(id)
        return True

    def remove_listener(self, id):
        self.listeners.remove(id)

    def create_output(self):

        def index_to_alphabet(ind):
            if ind < len(alphabet):
                return alphabet[ind].upper()

            remainder = ind % len(alphabet)
            return index_to_alphabet(ind - remainder) + alphabet[remainder].upper()

        wb = Workbook()
        all_phases = {}
        for mem in self.ongoing_online_phases:
            phases = all_phases.get(mem, None)
            if phases is None:
                all_phases[mem] = self.ongoing_online_phases[mem]
            else:
                all_phases[mem].extend(self.ongoing_online_phases[mem])
        for mem in self.ongoing_playing_phases:
            phases = all_phases.get(mem, None)
            if phases is None:
                all_phases[mem] = self.ongoing_playing_phases[mem]
            else:
                all_phases[mem].extend(self.ongoing_playing_phases[mem])

        for member in all_phases:
            ws = wb.create_sheet(self.musicbot.get_global_user(member).name)
            index = 1
            for action in sorted(all_phases[member], key=lambda phase: phase.start):
                ws["A{}".format(index)] = action.type_string
                ws["B{}".format(index)] = action.detailed_string
                ws["C{}".format(index)] = action.start_string
                ws["D{}".format(index)] = action.end_string
                ws["E{}".format(index)] = action.duration_string
                index += 1

            for dimension in ws.column_dimensions.values():
                dimension.auto_size = True

        wb.save("cache/last_survey_data.xlsx")

    def reset(self):
        self.ongoing_online_phases = {}
        self.ongoing_playing_phases = {}
        self.member_data = {}

    def update_stats(self, user_id, is_online, game_playing):
        user_data = self.get_user_data(user_id)
        last_online_phase = self.get_last_online_phase(user_id)
        last_playing_phase = self.get_last_playing_phase(user_id)

        if is_online and not user_data.is_online:
            # came online
            self.push_ongoing_online_phase(
                user_id, OnlinePhase(datetime.now()))

        if not is_online and user_data.is_online:
            # went offline
            if last_online_phase is not None:
                last_online_phase.set_end(datetime.now())

        if game_playing is not None and user_data.game_playing is None:
            # started playing
            self.push_ongoing_playing_phase(
                user_id, PlayingPhase(game_playing, datetime.now()))

        if game_playing is None and user_data.game_playing is not None:
            # stopped playing
            if last_playing_phase is not None:
                last_playing_phase.set_end(datetime.now())

        self.member_data[user_id] = MemberStaus(is_online, game_playing)

    def get_user_data(self, user_id):
        user_data = self.member_data.get(user_id, None)

        if user_data is None:
            return MemberStaus()

        return user_data

    def push_ongoing_online_phase(self, user_id, phase):
        phases = self.ongoing_online_phases.get(user_id, None)
        if phases is None:
            self.ongoing_online_phases[user_id] = [phase, ]
        else:
            self.ongoing_online_phases[user_id].append(phase)

    def push_ongoing_playing_phase(self, user_id, phase):
        phases = self.ongoing_playing_phases.get(user_id, None)
        if phases is None:
            self.ongoing_playing_phases[user_id] = [phase, ]
        else:
            self.ongoing_playing_phases[user_id].append(phase)

    def get_last_online_phase(self, user_id):
        online_phases = self.ongoing_online_phases.get(user_id, None)
        if online_phases is None or len(online_phases) < 1:
            return None

        return online_phases[-1]

    def get_last_playing_phase(self, user_id):
        playing_phases = self.ongoing_playing_phases.get(user_id, None)
        if playing_phases is None or len(playing_phases) < 1:
            return None

        return playing_phases[-1]


class MemberStaus:

    def __init__(self, is_online=False, game_playing=None):
        self.is_online = is_online
        self.game_playing = game_playing


class PlayingPhase:

    def __init__(self, game, start, end=None):
        self.game = game
        self.start = start
        self.end = end

    def set_end(self, end):
        self.end = end

    def __str__(self):
        if self.end is None:
            return "Started \"{0}\" at {1.month:0>2}-{1.day:0>2} {1.hour:0>2}:{1.minute:0>2}".format(self.game.name, self.start)

        return "Played \"{0}\" from {1.month:0>2}-{1.day:0>2} {1.hour:0>2}:{1.minute:0>2} to {2.month:0>2}-{2.day:0>2} {2.hour:0>2}:{2.minute:0>2}".format(self.game.name, self.start, self.end)

    @property
    def duration_string(self):
        return format_time(((self.end if self.end is not None else datetime.now()) - self.start).total_seconds(), round_seconds=True, max_specifications=3)

    @property
    def type_string(self):
        return "PLAYING"

    @property
    def detailed_string(self):
        return self.game.name

    @property
    def start_string(self):
        return "{0.year:0>4}/{0.month:0>2}/{0.day:0>2} {0.hour:0>2}:{0.minute:0>2}".format(self.start)

    @property
    def end_string(self):
        return "{0.year:0>4}/{0.month:0>2}/{0.day:0>2} {0.hour:0>2}:{0.minute:0>2}".format(self.end) if self.end is not None else "Until now"


class OnlinePhase:

    def __init__(self, start, end=None):
        self.start = start
        self.end = end

    def set_end(self, end):
        self.end = end

    def __str__(self):
        if self.end is None:
            return "Came online at {0.month:0>2}-{0.day:0>2} {0.hour:0>2}:{0.minute:0>2}".format(self.start)

        return "Was online from {0.month:0>2}-{0.day:0>2} {0.hour:0>2}:{0.minute:0>2} to {1.month:0>2}-{1.day:0>2} {1.hour:0>2}:{1.minute:0>2}".format(self.start, self.end)

    @property
    def duration_string(self):
        return format_time(((self.end if self.end is not None else datetime.now()) - self.start).total_seconds(), round_seconds=True, max_specifications=3)

    @property
    def type_string(self):
        return "ONLINE"

    @property
    def detailed_string(self):
        return ""

    @property
    def start_string(self):
        return "{0.year:0>4}/{0.month:0>2}/{0.day:0>2} {0.hour:0>2}:{0.minute:0>2}".format(self.start)

    @property
    def end_string(self):
        return "{0.year:0>4}/{0.month:0>2}/{0.day:0>2} {0.hour:0>2}:{0.minute:0>2}".format(self.end) if self.end is not None else "Until now"
